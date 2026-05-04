"""8D Raporu Excel üretici — openpyxl"""
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

KOYU_FILL  = PatternFill('solid', fgColor='1E293B')
YESIL_FILL = PatternFill('solid', fgColor='16A34A')
MAVI_FILL  = PatternFill('solid', fgColor='1D4ED8')
KIRMIZI_FILL = PatternFill('solid', fgColor='DC2626')
TURUNCU_FILL = PatternFill('solid', fgColor='EA580C')
GRI_FILL   = PatternFill('solid', fgColor='F8FAFC')
BEYAZ_FILL = PatternFill('solid', fgColor='FFFFFF')

INCE_BORDEl = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)

def _hucre(ws, row, col, value, bold=False, fill=None, font_color='1E293B',
           size=9, align='left', wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=font_color, size=size, name='Calibri')
    c.alignment = Alignment(horizontal=align, vertical='top', wrap_text=wrap)
    c.border = INCE_BORDEl
    if fill:
        c.fill = fill
    return c


def _bolum_satiri(ws, row, baslik, fill, ncol=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    _hucre(ws, row, 1, baslik, bold=True, fill=fill, font_color='FFFFFF', size=10, align='left')
    ws.row_dimensions[row].height = 20
    return row + 1


def _etiket_deger(ws, row, etiket, deger, fill_row=False):
    fill = GRI_FILL if fill_row else BEYAZ_FILL
    _hucre(ws, row, 1, etiket, bold=True, fill=fill, size=8)
    c = _hucre(ws, row, 2, deger, fill=fill, size=8)
    ws.row_dimensions[row].height = max(15, min(60, (len(str(deger or '')) // 60 + 1) * 15))
    return row + 1


def build_sekizd_excel(s) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = s.sekizd_no

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 72

    r = 1
    # Başlık
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _hucre(ws, r, 1, '8D — DÜZELTİCİ VE ÖNLEYİCİ FAALİYET RAPORU',
           bold=True, fill=KOYU_FILL, font_color='FFFFFF', size=14, align='center')
    ws.row_dimensions[r].height = 28
    r += 1

    # Alt başlık
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _hucre(ws, r, 1, s.sekizd_no, bold=True, fill=KOYU_FILL, font_color='FBBF24', size=11, align='center')
    ws.row_dimensions[r].height = 20
    r += 1

    r = _bolum_satiri(ws, r, 'RAPOR BİLGİLERİ', KOYU_FILL)
    pairs = [
        ('Tedarikçi', s.tedarikci.name if s.tedarikci else ''),
        ('Ürün Kodu / Adı', f"{s.urun_kodu or ''} / {s.urun_adi or ''}"),
        ('Tarih', s.tarih.strftime('%d.%m.%Y') if s.tarih else ''),
        ('Revizyon No', s.revizyon_no or '1'),
        ('Durum', {'taslak': 'Taslak', 'gonderildi': 'Gönderildi', 'kapali': 'Kapalı'}.get(s.durum or '', '')),
    ]
    for i, (lbl, val) in enumerate(pairs):
        r = _etiket_deger(ws, r, lbl, val, fill_row=(i % 2 == 1))

    r += 1
    r = _bolum_satiri(ws, r, 'D1 — EKİP OLUŞTURMA', YESIL_FILL)
    r = _etiket_deger(ws, r, 'Ekip Lideri', s.d1_ekip_lideri)
    r = _etiket_deger(ws, r, 'Ekip Üyeleri', s.d1_ekip_uyeleri, fill_row=True)

    r += 1
    r = _bolum_satiri(ws, r, 'D2 — PROBLEM TANIMI (5W2H)', MAVI_FILL)
    d2_pairs = [
        ('Problem Özeti', s.d2_problem_ozeti),
        ('Kim?', s.d2_kim),
        ('Ne?', s.d2_ne),
        ('Nerede?', s.d2_nerede),
        ('Ne Zaman?', s.d2_ne_zaman),
        ('Neden?', s.d2_neden),
        ('Nasıl?', s.d2_nasil),
        ('Ne Kadar?', s.d2_ne_kadar),
        ('İlk Tespit Yeri', s.d2_ilk_tespit),
    ]
    for i, (lbl, val) in enumerate(d2_pairs):
        r = _etiket_deger(ws, r, lbl, val, fill_row=(i % 2 == 1))

    r += 1
    r = _bolum_satiri(ws, r, 'D3 — GEÇİCİ ÖNLEMLER', TURUNCU_FILL)
    r = _etiket_deger(ws, r, 'Önlem', s.d3_onlem)
    r = _etiket_deger(ws, r, 'Sorumlu', s.d3_sorumlu, fill_row=True)
    r = _etiket_deger(ws, r, 'Tarih', s.d3_tarih.strftime('%d.%m.%Y') if s.d3_tarih else '')
    r = _etiket_deger(ws, r, 'Etkinlik Doğrulama', s.d3_etkinlik, fill_row=True)

    r += 1
    r = _bolum_satiri(ws, r, 'D4 — KÖK NEDEN ANALİZİ', KIRMIZI_FILL)
    metod_label = {'5why': '5 Why', 'balik_kilcigi': 'Balık Kılçığı', 'diger': 'Diğer'}
    r = _etiket_deger(ws, r, 'Kök Neden', s.d4_kok_neden)
    r = _etiket_deger(ws, r, 'Analiz Metodu', metod_label.get(s.d4_analiz_metod or '', s.d4_analiz_metod), fill_row=True)

    r += 1
    r = _bolum_satiri(ws, r, 'D5 — KALICI DÜZELTİCİ AKSİYON', YESIL_FILL)
    r = _etiket_deger(ws, r, 'Aksiyon', s.d5_aksiyon)

    r += 1
    r = _bolum_satiri(ws, r, 'D6 — UYGULANMASI', MAVI_FILL)
    r = _etiket_deger(ws, r, 'Uygulama', s.d6_uygulama)
    r = _etiket_deger(ws, r, 'Tarih', s.d6_tarih.strftime('%d.%m.%Y') if s.d6_tarih else '', fill_row=True)
    r = _etiket_deger(ws, r, 'Doğrulama', s.d6_dogrulama)

    r += 1
    r = _bolum_satiri(ws, r, 'D7 — TEKRARı ÖNLEYİCİ TEDBİRLER', TURUNCU_FILL)
    r = _etiket_deger(ws, r, 'Tedbirler', s.d7_onleyici)

    r += 1
    r = _bolum_satiri(ws, r, 'D8 — EKİP TAKDİRİ', YESIL_FILL)
    r = _etiket_deger(ws, r, 'Takdir Notu', s.d8_tadir)

    # İmza alanı
    r += 2
    for col, baslik in enumerate(['Hazırlayan', 'Onaylayan', 'Tedarikçi'], start=1):
        if col <= 2:
            ws.column_dimensions[get_column_letter(col)].width = 36
        c = ws.cell(row=r, column=col if col < 3 else 3, value=baslik)
        c.font = Font(bold=True, size=9, name='Calibri')
        c.alignment = Alignment(horizontal='center')
        c.border = INCE_BORDEl
        c.fill = GRI_FILL
        ws.row_dimensions[r].height = 16
        # İmza satırı
        c2 = ws.cell(row=r+1, column=col if col < 3 else 3, value='')
        c2.border = INCE_BORDEl
        ws.row_dimensions[r+1].height = 40

    # Print settings
    ws.print_title_rows = '1:3'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
