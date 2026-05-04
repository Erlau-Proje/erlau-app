from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.orm import selectinload
from app import db
from app.models import User, UserPermission, UserScopeDepartment, UserScopeStation, Department, TalepFormu, TalepKalem, Tedarikci, Fatura, FaturaKalem, TedarikciSablon, Malzeme, Urun, IsIstasyonu, UretimPersoneli, UretimPlani, UretimPlaniSatir, UretimKaydi, ArizaKaydi, Makine, BakimPlani, BakimKaydi, PlanliBakim, PeriyodikKontrol, TeklifGrubu, TeklifKalem, TeknikResim, KaliteKontrol, DOF, DOFAksiyon
from app.utils import generate_siparis_no, generate_stok_kodu, generate_urun_kodu, generate_plan_no, generate_teklif_no, devir_gunu
from app.permissions import PERMISSIONS, PERMISSION_BY_CODE, ENDPOINT_PERMISSIONS, has_permission, default_permission_allowed
from app.access_profiles import JOB_PROFILES, SCOPE_TYPES, USER_TYPES, PROFILE_LABELS, SCOPE_LABELS, USER_TYPE_LABELS, infer_profile
from datetime import datetime, date, timedelta
from functools import wraps

auth = Blueprint('auth', __name__)
main = Blueprint('main', __name__)
satin_alma = Blueprint('satin_alma', __name__, url_prefix='/satinalma')
admin = Blueprint('admin', __name__, url_prefix='/admin')
muhasebe = Blueprint('muhasebe', __name__, url_prefix='/muhasebe')
uretim = Blueprint('uretim', __name__, url_prefix='/uretim')
bakim = Blueprint('bakim', __name__, url_prefix='/bakim')
planlama = Blueprint('planlama', __name__, url_prefix='/planlama')
teknik_resim_bp = Blueprint('teknik_resim', __name__, url_prefix='/teknik-resim')

_TEKNIK_EDIT_ROLES = ('admin', 'satinalma', 'gm', 'departman_yoneticisi', 'proje')

def teknik_yetki_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not has_permission(current_user, 'technical.manage'):
            flash('Bu işlem için yetkiniz yok.', 'danger')
            return redirect(url_for('teknik_resim.teknik_resim_listesi'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            endpoint_permission = ENDPOINT_PERMISSIONS.get(request.endpoint)
            has_endpoint_permission = endpoint_permission and has_permission(current_user, endpoint_permission)
            if not current_user.is_authenticated or (current_user.role not in roles and not has_endpoint_permission):
                flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
                return redirect(url_for('main.dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@auth.route('/', methods=['GET'])
@auth.route('/portal', methods=['GET'])
def portal():
    if current_user.is_authenticated and request.path == '/':
        return redirect(url_for('main.dashboard'))
    return render_template('portal.html')

@auth.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password) and user.is_active:
            login_user(user)
            user.son_giris = datetime.utcnow()
            db.session.commit()
            return redirect(url_for('main.dashboard'))
        flash('E-posta veya şifre hatalı.', 'danger')
    return render_template('login.html')

@auth.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))

@main.route('/dashboard')
@login_required
def dashboard():
    from app.services import get_gm_dashboard_stats
    bugun = date.today()

    if current_user.role == 'gm':
        stats_data = get_gm_dashboard_stats()

        # ── GECİKMİŞ SİPARİŞLER ──────────────────────────────────────────────
        yolda_talepler = TalepFormu.query.options(
            selectinload(TalepFormu.kalemler)
        ).filter(
            TalepFormu.durum == 'yolda',
            TalepFormu.yolda_tarihi != None
        ).all()

        gecikis_listesi = []
        for t in yolda_talepler:
            if not t.kalemler:
                continue
            termin = max((k.termin_gun or 0) for k in t.kalemler)
            if termin > 0:
                bitis = t.yolda_tarihi.date() + timedelta(days=termin)
                if bitis < bugun:
                    gecikis_listesi.append({'talep': t, 'gecikme': (bugun - bitis).days})
        gecikis_listesi.sort(key=lambda x: x['gecikme'], reverse=True)

        return render_template('dashboard.html',
            gm_stats=stats_data['gm_stats'],
            dept_stats=stats_data['dept_stats'],
            bekleyen_talepler=stats_data['bekleyen_talepler'],
            trend_json=stats_data['trend_json'],
            bekleme_stats=stats_data['bekleme_stats'],
            gecikis_listesi=gecikis_listesi,
            top_tedarikci=stats_data['top_tedarikci'],
            talepler=[], kalan_gunler={}, pagination=None)

    page = request.args.get('page', 1, type=int)
    if current_user.role == 'departman_yoneticisi':
        base_filter = {'department_id': current_user.department_id}
        pagination = TalepFormu.query.options(
            selectinload(TalepFormu.kalemler),
            selectinload(TalepFormu.talep_eden),
        ).filter_by(**base_filter).order_by(TalepFormu.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    else:
        base_filter = {'talep_eden_id': current_user.id}
        pagination = TalepFormu.query.options(
            selectinload(TalepFormu.kalemler),
        ).filter_by(**base_filter).order_by(TalepFormu.created_at.desc()).paginate(page=page, per_page=20, error_out=False)

    talepler = pagination.items

    # Yolda siparişleri tüm kayıtlardan ayrıca çek (sayfa sınırından bağımsız)
    yolda_talepler = TalepFormu.query.options(
        selectinload(TalepFormu.kalemler)
    ).filter_by(**base_filter, durum='yolda').order_by(TalepFormu.created_at.desc()).all()

    kalan_gunler = {}
    for talep in yolda_talepler:
        if talep.yolda_tarihi:
            termin = max((k.termin_gun or 0) for k in talep.kalemler) if talep.kalemler else 0
            if termin > 0:
                bitis = talep.yolda_tarihi.date() + timedelta(days=termin)
                kalan_gunler[talep.id] = (bitis - bugun).days

    # İstatistikler tüm kayıtlar üzerinden
    stats_q = TalepFormu.query.filter_by(**base_filter)
    stats = {
        'toplam': stats_q.count(),
        'bekleyen': stats_q.filter(TalepFormu.durum.in_(('bekliyor', 'fiyatlandirildi'))).count(),
        'yolda': stats_q.filter_by(durum='yolda').count(),
        'onaylandi': stats_q.filter_by(durum='onaylandi').count(),
        'teslim': stats_q.filter_by(durum='teslim_alindi').count(),
    }

    return render_template('dashboard.html',
        talepler=talepler,
        yolda_talepler=yolda_talepler,
        kalan_gunler=kalan_gunler,
        pagination=pagination,
        stats=stats,
        gm_stats=None)

@main.route('/arama')
@login_required
def arama():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    sonuclar = []
    pagination = None
    if q:
        from sqlalchemy import or_
        kalem_query = TalepKalem.query.join(TalepFormu).filter(
            or_(
                TalepKalem.malzeme_adi.ilike(f'%{q}%'),
                TalepKalem.marka_model.ilike(f'%{q}%'),
                TalepKalem.malzeme_turu.ilike(f'%{q}%'),
                TalepKalem.aciklama.ilike(f'%{q}%'),
                TalepKalem.kullanim_amaci.ilike(f'%{q}%'),
                TalepKalem.teknik_resim_kodu.ilike(f'%{q}%'),
                TalepKalem.standart.ilike(f'%{q}%'),
                TalepFormu.siparis_no.ilike(f'%{q}%'),
            )
        )
        if current_user.role not in ['satinalma', 'admin', 'gm']:
            if current_user.role == 'departman_yoneticisi':
                kalem_query = kalem_query.filter(TalepFormu.department_id == current_user.department_id)
            else:
                kalem_query = kalem_query.filter(TalepFormu.talep_eden_id == current_user.id)
        
        pagination = kalem_query.order_by(TalepFormu.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
        sonuclar = pagination.items
        
    return render_template('arama.html', q=q, sonuclar=sonuclar, pagination=pagination)

@main.route('/talep/yeni', methods=['GET', 'POST'])
@login_required
def yeni_talep():
    if request.method == 'POST':
        talep = TalepFormu(
            siparis_no=generate_siparis_no(),
            talep_eden_id=current_user.id,
            department_id=current_user.department_id,
            durum='bekliyor'
        )
        db.session.add(talep)
        db.session.flush()

        malzeme_adlari = request.form.getlist('malzeme_adi[]')
        marka_modeller = request.form.getlist('marka_model[]')
        malzeme_turleri = request.form.getlist('malzeme_turu[]')
        birimler = request.form.getlist('birim[]')
        miktarlar = request.form.getlist('miktar[]')
        hedefler = request.form.getlist('hedef[]')
        kwler = request.form.getlist('kw[]')
        aciklamalar = request.form.getlist('aciklama[]')
        anlik_stoklar = request.form.getlist('anlik_stok[]')
        kullanim_amaclari = request.form.getlist('kullanim_amaci[]')
        kullanilan_alanlar = request.form.getlist('kullanilan_alan[]')
        proje_makineler = request.form.getlist('proje_makine[]')

        for i, ad in enumerate(malzeme_adlari):
            if ad.strip():
                kalem = TalepKalem(
                    talep_id=talep.id,
                    malzeme_adi=ad,
                    marka_model=marka_modeller[i] if i < len(marka_modeller) else '',
                    malzeme_turu=malzeme_turleri[i] if i < len(malzeme_turleri) else '',
                    birim=birimler[i] if i < len(birimler) else 'Adet',
                    miktar=float(miktarlar[i]) if i < len(miktarlar) and miktarlar[i] else 0,
                    hedef=hedefler[i] if i < len(hedefler) else 'siparis',
                    kullanim_amaci=kullanim_amaclari[i] if i < len(kullanim_amaclari) else '',
                    kullanilan_alan=kullanilan_alanlar[i] if i < len(kullanilan_alanlar) else '',
                    proje_makine=proje_makineler[i] if i < len(proje_makineler) else '',
                    kw=kwler[i] if i < len(kwler) else '',
                    aciklama=aciklamalar[i] if i < len(aciklamalar) else '',
                    anlik_stok=anlik_stoklar[i] if i < len(anlik_stoklar) else ''
                )
                db.session.add(kalem)

        db.session.commit()
        flash(f'Talep {talep.siparis_no} başarıyla oluşturuldu!', 'success')
        return redirect(url_for('main.dashboard'))

    return render_template('yeni_talep.html')

@main.route('/talep/<int:talep_id>')
@login_required
def talep_detay(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    return render_template('talep_detay.html', talep=talep)

@main.route('/talep/<int:talep_id>/duzenle', methods=['GET', 'POST'])
@login_required
def talep_duzenle(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    if talep.talep_eden_id != current_user.id and current_user.role != 'admin':
        flash('Bu talebi düzenleme yetkiniz yok.', 'danger')
        return redirect(url_for('main.dashboard'))
    if talep.durum != 'bekliyor':
        flash('Sadece bekleyen talepler düzenlenebilir.', 'danger')
        return redirect(url_for('main.talep_detay', talep_id=talep_id))
    if request.method == 'POST':
        for kalem in list(talep.kalemler):
            db.session.delete(kalem)
        db.session.flush()
        malzeme_adlari = request.form.getlist('malzeme_adi[]')
        marka_modeller = request.form.getlist('marka_model[]')
        malzeme_turleri = request.form.getlist('malzeme_turu[]')
        birimler = request.form.getlist('birim[]')
        miktarlar = request.form.getlist('miktar[]')
        hedefler = request.form.getlist('hedef[]')
        kullanim_amaclari = request.form.getlist('kullanim_amaci[]')
        kullanilan_alanlar = request.form.getlist('kullanilan_alan[]')
        proje_makineler = request.form.getlist('proje_makine[]')
        kwler = request.form.getlist('kw[]')
        aciklamalar = request.form.getlist('aciklama[]')
        anlik_stoklar = request.form.getlist('anlik_stok[]')
        for i, ad in enumerate(malzeme_adlari):
            if ad.strip():
                db.session.add(TalepKalem(
                    talep_id=talep.id,
                    malzeme_adi=ad,
                    marka_model=marka_modeller[i] if i < len(marka_modeller) else '',
                    malzeme_turu=malzeme_turleri[i] if i < len(malzeme_turleri) else '',
                    birim=birimler[i] if i < len(birimler) else 'Adet',
                    miktar=float(miktarlar[i]) if i < len(miktarlar) and miktarlar[i] else 0,
                    hedef=hedefler[i] if i < len(hedefler) else 'siparis',
                    kullanim_amaci=kullanim_amaclari[i] if i < len(kullanim_amaclari) else '',
                    kullanilan_alan=kullanilan_alanlar[i] if i < len(kullanilan_alanlar) else '',
                    proje_makine=proje_makineler[i] if i < len(proje_makineler) else '',
                    kw=kwler[i] if i < len(kwler) else '',
                    aciklama=aciklamalar[i] if i < len(aciklamalar) else '',
                    anlik_stok=anlik_stoklar[i] if i < len(anlik_stoklar) else ''
                ))
        db.session.commit()
        flash('Talep güncellendi.', 'success')
        return redirect(url_for('main.talep_detay', talep_id=talep_id))
    return render_template('talep_duzenle.html', talep=talep)

@main.route('/talep/<int:talep_id>/sil', methods=['POST'])
@login_required
def talep_sil(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    is_satinalma = current_user.role in ('satinalma', 'admin')
    if talep.talep_eden_id != current_user.id and not is_satinalma:
        flash('Bu talebi silme yetkiniz yok.', 'danger')
        return redirect(url_for('main.dashboard'))
    if talep.durum != 'bekliyor' and not is_satinalma:
        flash('Sadece bekleyen talepler silinebilir.', 'danger')
        return redirect(url_for('main.dashboard'))
    db.session.delete(talep)
    db.session.commit()
    flash('Talep silindi.', 'success')
    redirect_url = url_for('satin_alma.panel') if is_satinalma else url_for('main.dashboard')
    return redirect(redirect_url)

@satin_alma.route('/siparis-raporu')
@login_required
def siparis_raporu():
    from sqlalchemy import func

    # Filtreler — tarih boşsa son 6 ay varsayılan
    _varsayilan_bas = (datetime.utcnow() - timedelta(days=182)).strftime('%Y-%m-%d')
    bas_str = request.args.get('bas_tarih', _varsayilan_bas)
    bit_str = request.args.get('bit_tarih', '')
    tur_filtre = request.args.get('malzeme_turu', '')
    ted_filtre = request.args.get('tedarikci_id', '')
    durum_filtre = request.args.get('durum', '')
    hedef_filtre = request.args.get('hedef', '')
    amac_filtre = request.args.get('kullanim_amaci', '')
    alan_filtre = request.args.get('kullanilan_alan', '')
    proje_filtre = request.args.get('proje_makine', '').strip()
    export = request.args.get('export', '')
    page = request.args.get('page', 1, type=int)

    # Base query — options eklenmeden, aggregation sorgularında da kullanılacak
    query = TalepKalem.query.join(TalepFormu)

    # Rol bazlı kapsam
    if current_user.role in ['satinalma', 'admin', 'gm']:
        pass  # hepsini görür
    else:
        query = query.filter(TalepFormu.department_id == current_user.department_id)

    if bas_str:
        try:
            bas = datetime.strptime(bas_str, '%Y-%m-%d')
            query = query.filter(TalepFormu.created_at >= bas)
        except ValueError:
            pass
    if bit_str:
        try:
            bit = datetime.strptime(bit_str, '%Y-%m-%d')
            bit = bit.replace(hour=23, minute=59, second=59)
            query = query.filter(TalepFormu.created_at <= bit)
        except ValueError:
            pass
    if tur_filtre:
        query = query.filter(TalepKalem.malzeme_turu == tur_filtre)
    if ted_filtre:
        query = query.filter(TalepKalem.tedarikci_id == int(ted_filtre))
    if durum_filtre:
        query = query.filter(TalepFormu.durum == durum_filtre)
    if hedef_filtre:
        query = query.filter(TalepKalem.hedef == hedef_filtre)
    if amac_filtre:
        query = query.filter(TalepKalem.kullanim_amaci == amac_filtre)
    if alan_filtre:
        query = query.filter(TalepKalem.kullanilan_alan == alan_filtre)
    if proje_filtre:
        query = query.filter(TalepKalem.proje_makine.ilike(f'%{proje_filtre}%'))

    _eager = [
        selectinload(TalepKalem.talep).selectinload(TalepFormu.talep_eden),
        selectinload(TalepKalem.talep).selectinload(TalepFormu.department),
        selectinload(TalepKalem.tedarikci)
    ]

    # Excel export — filtreli tüm kayıtları al
    if export == 'excel':
        kalemler = query.options(*_eager).order_by(TalepFormu.created_at.desc()).all()
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import make_response
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sipariş Raporu'
        basliklar = ['Sipariş No', 'Tarih', 'Talep Eden', 'Departman', 'Malzeme Adı',
                     'Marka/Model', 'Malzeme Türü', 'Kullanım Amacı', 'Kullanılan Alan',
                     'Proje/Makine', 'Miktar', 'Birim', 'Br. Fiyat', 'Toplam', 'Para Birimi',
                     'Tedarikçi', 'Vade (gün)', 'Termin (gün)', 'Durum']
        yesil = PatternFill(fill_type='solid', fgColor='2d7a3a')
        for col, b in enumerate(basliklar, 1):
            c = ws.cell(row=1, column=col, value=b)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = yesil
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[c.column_letter].width = 18
        for i, k in enumerate(kalemler, 2):
            t = k.talep
            ws.append([
                t.siparis_no,
                t.created_at.strftime('%d.%m.%Y'),
                t.talep_eden.name if t.talep_eden else '',
                t.department.name if t.department else '',
                k.malzeme_adi, k.marka_model or '',
                k.malzeme_turu or '', k.kullanim_amaci or '',
                k.kullanilan_alan or '', k.proje_makine or '',
                k.miktar or '', k.birim or '',
                k.br_fiyat or '', k.toplam_fiyat or '',
                k.para_birimi or '',
                k.tedarikci.name if k.tedarikci else '',
                k.vade_gun or '', k.termin_gun or '',
                t.durum
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import make_response
        tarih = datetime.utcnow().strftime('%Y%m%d')
        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="siparis_raporu_{tarih}.xlsx"'
        return resp

    # Özet: türe göre — tüm filtreli kayıtlardan (DB aggregation)
    tur_agg = query.with_entities(
        TalepKalem.malzeme_turu,
        TalepKalem.para_birimi,
        func.count(TalepKalem.id).label('adet'),
        func.sum(TalepKalem.toplam_fiyat).label('toplam'),
    ).group_by(TalepKalem.malzeme_turu, TalepKalem.para_birimi).all()

    tur_ozet = {}
    for row in tur_agg:
        tur = row.malzeme_turu or 'Belirtilmemiş'
        if tur not in tur_ozet:
            tur_ozet[tur] = {'adet': 0, 'toplam': 0, 'para': ''}
        tur_ozet[tur]['adet'] += row.adet
        tur_ozet[tur]['toplam'] += row.toplam or 0
        if row.para_birimi:
            tur_ozet[tur]['para'] = row.para_birimi

    # Özet: tedarikçiye göre — tüm filtreli kayıtlardan (DB aggregation)
    ted_agg = query.filter(TalepKalem.tedarikci_id.isnot(None)).with_entities(
        TalepKalem.tedarikci_id,
        TalepKalem.para_birimi,
        func.count(TalepKalem.id).label('adet'),
        func.sum(TalepKalem.toplam_fiyat).label('toplam'),
    ).group_by(TalepKalem.tedarikci_id, TalepKalem.para_birimi).all()

    ted_id_set = {r.tedarikci_id for r in ted_agg}
    ted_name_map = {t.id: t.name for t in Tedarikci.query.filter(Tedarikci.id.in_(ted_id_set)).all()} if ted_id_set else {}
    ted_ozet = {}
    for row in ted_agg:
        name = ted_name_map.get(row.tedarikci_id, '—')
        if name not in ted_ozet:
            ted_ozet[name] = {'adet': 0, 'toplam': 0, 'para': '', 'id': row.tedarikci_id}
        ted_ozet[name]['adet'] += row.adet
        ted_ozet[name]['toplam'] += row.toplam or 0
        if row.para_birimi:
            ted_ozet[name]['para'] = row.para_birimi

    toplam_tutar = query.with_entities(func.sum(TalepKalem.toplam_fiyat)).scalar() or 0

    # Sayfa görünümü için pagination (20 per page)
    pagination = query.options(*_eager).order_by(TalepFormu.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    kalemler = pagination.items

    def distinct_vals(col):
        return [v[0] for v in db.session.query(col).filter(col != None, col != '').distinct().order_by(col).all()]

    turler = distinct_vals(TalepKalem.malzeme_turu)
    hedefler = distinct_vals(TalepKalem.hedef)
    amaclar = distinct_vals(TalepKalem.kullanim_amaci)
    alanlar = distinct_vals(TalepKalem.kullanilan_alan)
    tedarikciler = Tedarikci.query.filter_by(is_active=True).order_by(Tedarikci.name).all()

    return render_template('siparis_raporu.html',
        kalemler=kalemler, pagination=pagination, tur_ozet=tur_ozet, ted_ozet=ted_ozet,
        toplam_tutar=toplam_tutar,
        turler=turler, hedefler=hedefler, amaclar=amaclar, alanlar=alanlar,
        tedarikciler=tedarikciler,
        bas_str=bas_str, bit_str=bit_str,
        tur_filtre=tur_filtre, ted_filtre=ted_filtre, durum_filtre=durum_filtre,
        hedef_filtre=hedef_filtre, amac_filtre=amac_filtre,
        alan_filtre=alan_filtre, proje_filtre=proje_filtre
    )

@satin_alma.route('/raporlar')
@login_required
@role_required('satinalma', 'admin', 'gm')
def raporlar():
    from sqlalchemy import func, extract
    from datetime import date

    # Durum özeti
    durum_rows = db.session.query(TalepFormu.durum, func.count(TalepFormu.id)).group_by(TalepFormu.durum).all()
    durum_map = dict(durum_rows)

    # Son 6 ay — talep sayısı
    bugun = date.today()
    aylar, ay_sayilari = [], []
    for i in range(5, -1, -1):
        ay = bugun.month - i
        yil = bugun.year
        while ay <= 0:
            ay += 12
            yil -= 1
        sayi = TalepFormu.query.filter(
            extract('month', TalepFormu.created_at) == ay,
            extract('year', TalepFormu.created_at) == yil
        ).count()
        aylar.append(f"{ay:02d}/{yil}")
        ay_sayilari.append(sayi)

    # Departmana göre talep sayısı
    dept_rows = db.session.query(Department.name, func.count(TalepFormu.id))\
        .join(TalepFormu, TalepFormu.department_id == Department.id)\
        .group_by(Department.name).order_by(func.count(TalepFormu.id).desc()).all()

    # Malzeme türüne göre
    tur_rows = db.session.query(TalepKalem.malzeme_turu, func.count(TalepKalem.id))\
        .filter(TalepKalem.malzeme_turu != None)\
        .group_by(TalepKalem.malzeme_turu).order_by(func.count(TalepKalem.id).desc()).all()

    # Tedarikçiye göre harcama (TL)
    ted_rows = db.session.query(Tedarikci.name, func.coalesce(func.sum(TalepKalem.toplam_fiyat), 0))\
        .join(TalepKalem, TalepKalem.tedarikci_id == Tedarikci.id)\
        .filter(TalepKalem.para_birimi == 'TL')\
        .group_by(Tedarikci.name).order_by(func.sum(TalepKalem.toplam_fiyat).desc()).limit(10).all()

    toplam_talep = TalepFormu.query.count()
    toplam_kalem = TalepKalem.query.count()
    toplam_tutar = db.session.query(func.coalesce(func.sum(TalepKalem.toplam_fiyat), 0))\
        .filter(TalepKalem.para_birimi == 'TL').scalar() or 0

    return render_template('raporlar.html',
        durum_map=durum_map,
        aylar=aylar, ay_sayilari=ay_sayilari,
        dept_rows=dept_rows, tur_rows=tur_rows, ted_rows=ted_rows,
        toplam_talep=toplam_talep, toplam_kalem=toplam_kalem, toplam_tutar=toplam_tutar
    )

@satin_alma.route('/panel')
@login_required
@role_required('satinalma', 'admin', 'gm')
def panel():
    durum = request.args.get('durum', 'hepsi')
    dept = request.args.get('dept', '')
    arama = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    q = TalepFormu.query.options(
        selectinload(TalepFormu.kalemler),
        selectinload(TalepFormu.talep_eden),
        selectinload(TalepFormu.department),
    )
    if durum != 'hepsi':
        q = q.filter_by(durum=durum)
    # Satınalma/admin/gm tüm siparişleri görebilir; diğerleri sadece kendi departmanını
    if current_user.role in ('satinalma', 'admin', 'gm'):
        if dept:
            q = q.filter_by(department_id=dept)
    else:
        q = q.filter_by(department_id=current_user.department_id)
    if arama:
        from sqlalchemy import or_
        q = q.filter(or_(
            TalepFormu.siparis_no.ilike(f'%{arama}%'),
            TalepFormu.kalemler.any(TalepKalem.malzeme_adi.ilike(f'%{arama}%'))
        ))
    pagination = q.order_by(TalepFormu.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    departmanlar = Department.query.all()
    return render_template('satinalma_panel.html',
        talepler=pagination.items,
        pagination=pagination,
        departmanlar=departmanlar,
        secili_durum=durum,
        arama=arama,
        dept_filtre=dept)

@satin_alma.route('/onayla/<int:talep_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def onayla(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    talep.durum = 'onaylandi'
    db.session.commit()
    flash('Talep onaylandı.', 'success')
    # Arka planda malzeme kullanım bilgisini öğren
    try:
        _malzeme_kullanim_ogren(talep)
    except Exception:
        pass
    return redirect(url_for('satin_alma.panel'))


def _malzeme_kullanim_ogren(talep):
    """Onaylanan siparişteki malzemeleri Malzeme tablosuna eşleştir ve kullanım notunu güncelle."""
    import os
    from sqlalchemy import func
    for kalem in talep.kalemler:
        if not kalem.malzeme_adi:
            continue
        # Malzeme tablosunda eşleşme ara
        mal = Malzeme.query.filter(
            Malzeme.malzeme_adi.ilike(f'%{kalem.malzeme_adi[:30]}%'),
            Malzeme.is_active == True
        ).first()
        if not mal:
            continue
        # Kullanım bilgisi oluştur
        proje = kalem.proje_makine or ''
        amac = kalem.kullanim_amaci or ''
        alan = kalem.kullanilan_alan or ''
        yeni_bilgi = f"{talep.siparis_no}"
        if proje: yeni_bilgi += f" / {proje}"
        if amac:  yeni_bilgi += f" / {amac}"
        # Mevcut nota ekle (son 5 kullanım tut)
        mevcut = mal.kullanim_notu or ''
        satirlar = [x.strip() for x in mevcut.split('\n') if x.strip()]
        if yeni_bilgi not in satirlar:
            satirlar.append(yeni_bilgi)
        mal.kullanim_notu = '\n'.join(satirlar[-5:])  # son 5 kullanım
    db.session.commit()

@satin_alma.route('/iptal/<int:talep_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def iptal(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    talep.durum = 'iptal'
    db.session.commit()
    flash('Talep iptal edildi.', 'warning')
    return redirect(url_for('satin_alma.panel'))

@admin.route('/kullanicilar')
@login_required
@role_required('admin')
def kullanicilar():
    users = User.query.order_by(User.is_active.desc(), User.name).all()
    for user in users:
        if not user.job_profile or not user.user_type or not user.scope_type:
            profile, user_type, scope_type = infer_profile(user)
            user.job_profile = user.job_profile or profile
            user.user_type = user.user_type or user_type
            user.scope_type = user.scope_type or scope_type
    departmanlar = Department.query.order_by(Department.name).all()
    istasyonlar = IsIstasyonu.query.filter_by(is_active=True).order_by(IsIstasyonu.istasyon_adi).all()
    uretim_personeller = UretimPersoneli.query.filter_by(is_active=True).order_by(UretimPersoneli.ad, UretimPersoneli.soyad).all()
    roller = [
        ('personel', 'Personel'),
        ('departman_yoneticisi', 'Departman Yöneticisi'),
        ('proje', 'Proje Departmanı'),
        ('satinalma', 'Satınalma'),
        ('gm', 'Genel Müdür'),
        ('muhasebe', 'Muhasebe'),
        ('uretim', 'Üretim'),
        ('planlama', 'Planlama'),
        ('kalite', 'Kalite'),
        ('bakim', 'Bakım'),
        ('admin', 'Admin'),
    ]
    db.session.commit()
    return render_template('admin_kullanicilar_gelismis.html',
        users=users, departmanlar=departmanlar, roller=roller,
        istasyonlar=istasyonlar, uretim_personeller=uretim_personeller,
        job_profiles=JOB_PROFILES, scope_types=SCOPE_TYPES, user_types=USER_TYPES,
        profile_labels=PROFILE_LABELS, scope_labels=SCOPE_LABELS, user_type_labels=USER_TYPE_LABELS)


def _sync_user_scope(user):
    UserScopeDepartment.query.filter_by(user_id=user.id).delete()
    UserScopeStation.query.filter_by(user_id=user.id).delete()
    for department_id in request.form.getlist('scope_department_ids'):
        if department_id:
            db.session.add(UserScopeDepartment(user_id=user.id, department_id=int(department_id)))
    for station_id in request.form.getlist('scope_station_ids'):
        if station_id:
            db.session.add(UserScopeStation(user_id=user.id, station_id=int(station_id)))

@admin.route('/yetki-matrisi')
@login_required
@role_required('admin')
def yetki_matrisi():
    users = User.query.order_by(User.is_active.desc(), User.name).all()
    temel_kullanicilar = [
        user for user in users
        if user.is_active
        and user.role == 'personel'
        and not user.liste_yetki
        and not user.teknik_resim_yetki
    ]

    satirlar = []
    for yetki in PERMISSIONS:
        izinler = {user.id: has_permission(user, yetki['code']) for user in users}
        varsayilanlar = {user.id: default_permission_allowed(user, yetki) for user in users}
        satirlar.append({
            'grup': yetki['grup'],
            'ad': yetki['ad'],
            'code': yetki['code'],
            'tip': yetki['tip'],
            'izinler': izinler,
            'varsayilanlar': varsayilanlar,
            'atanmis_sayisi': sum(1 for izinli in izinler.values() if izinli)
        })
    mudahale_satirlari = [satir for satir in satirlar if satir['tip'] == 'mudahale']
    atanmamis_mudahale_satirlari = [satir for satir in mudahale_satirlari if satir['atanmis_sayisi'] == 0]
    return render_template('admin_yetki_matrisi.html',
        users=users,
        satirlar=satirlar,
        temel_kullanicilar=temel_kullanicilar,
        mudahale_satirlari=mudahale_satirlari,
        atanmamis_mudahale_satirlari=atanmamis_mudahale_satirlari)

@admin.route('/yetki-matrisi/kaydet', methods=['POST'])
@login_required
@role_required('admin')
def yetki_matrisi_kaydet():
    permission_codes = {p['code'] for p in PERMISSIONS}
    users = User.query.all()
    if not has_permission(current_user, 'admin.permissions'):
        flash('Yetki matrisi düzenleme yetkiniz yok.', 'danger')
        return redirect(url_for('admin.yetki_matrisi'))

    for user in users:
        for code in permission_codes:
            field = f'perm__{user.id}__{code}'
            allowed = field in request.form
            if user.id == current_user.id and code in ('admin.permissions', 'admin.user_manage'):
                allowed = True
            kayit = UserPermission.query.filter_by(user_id=user.id, permission_code=code).first()
            if not kayit:
                kayit = UserPermission(user_id=user.id, permission_code=code)
                db.session.add(kayit)
            kayit.allowed = allowed
            kayit.updated_by_id = current_user.id
    db.session.commit()
    flash('Yetki matrisi güncellendi.', 'success')
    return redirect(url_for('admin.yetki_matrisi'))

@admin.route('/kullanici/ekle', methods=['POST'])
@login_required
@role_required('admin')
def kullanici_ekle():
    name = request.form.get('name')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')
    department_id = request.form.get('department_id')
    if User.query.filter_by(email=email).first():
        flash('Bu e-posta zaten kayıtlı.', 'danger')
        return redirect(url_for('admin.kullanicilar'))
    user = User(
        name=name, email=email,
        password=generate_password_hash(password),
        role=role,
        department_id=department_id or None,
        user_type=request.form.get('user_type') or 'office',
        job_profile=request.form.get('job_profile') or 'sadece_goruntuleme',
        scope_type=request.form.get('scope_type') or 'self',
        uretim_personeli_id=request.form.get('uretim_personeli_id') or None,
        unvan=request.form.get('unvan', '').strip(),
        telefon=request.form.get('telefon', '').strip(),
        teknik_resim_yetki=bool(request.form.get('teknik_resim_yetki')),
        liste_yetki=bool(request.form.get('liste_yetki')),
        bildirim_email=bool(request.form.get('bildirim_email')),
        unvan_pdf_goster=bool(request.form.get('unvan_pdf_goster')),
    )
    db.session.add(user)
    db.session.flush()
    _sync_user_scope(user)
    db.session.commit()
    flash(f'{name} başarıyla eklendi.', 'success')
    return redirect(url_for('admin.kullanicilar'))

@admin.route('/kullanici/<int:user_id>/duzenle', methods=['POST'])
@login_required
@role_required('admin')
def kullanici_duzenle(user_id):
    user = db.get_or_404(User, user_id)
    user.name = request.form.get('name')
    user.role = request.form.get('role')
    user.department_id = request.form.get('department_id') or None
    user.user_type = request.form.get('user_type') or 'office'
    user.job_profile = request.form.get('job_profile') or 'sadece_goruntuleme'
    user.scope_type = request.form.get('scope_type') or 'self'
    user.uretim_personeli_id = request.form.get('uretim_personeli_id') or None
    user.unvan = request.form.get('unvan', '').strip()
    user.telefon = request.form.get('telefon', '').strip()
    user.teknik_resim_yetki = bool(request.form.get('teknik_resim_yetki'))
    user.liste_yetki = bool(request.form.get('liste_yetki'))
    user.bildirim_email = bool(request.form.get('bildirim_email'))
    user.unvan_pdf_goster = bool(request.form.get('unvan_pdf_goster'))
    user.is_active = request.form.get('is_active') == '1'
    _sync_user_scope(user)
    db.session.commit()
    flash(f'{user.name} güncellendi.', 'success')
    return redirect(url_for('admin.kullanicilar'))

@admin.route('/kullanici/<int:user_id>/sil', methods=['POST'])
@login_required
@role_required('admin')
def kullanici_sil(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == current_user.id:
        flash('Kendi hesabınızı silemezsiniz.', 'danger')
        return redirect(url_for('admin.kullanicilar'))
    user.is_active = False
    db.session.commit()
    flash(f'{user.name} pasife alındı.', 'warning')
    return redirect(url_for('admin.kullanicilar'))

@admin.route('/tedarikci')
@login_required
@role_required('admin', 'satinalma')
def tedarikci_listesi():
    tedarikciler = Tedarikci.query.all()
    return render_template('tedarikci.html', tedarikciler=tedarikciler)

@admin.route('/tedarikci/sablonu-indir')
@login_required
@role_required('admin', 'satinalma')
def tedarikci_sablon_indir():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import make_response
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tedarikçiler'
    basliklar = ['Firma Adı*', 'Unvan', 'Vergi No', 'E-posta', 'Telefon', 'İletişim Kişisi', 'Adres', 'Para Birimi (TL/USD/EUR)', 'Vade Gün', 'Kategori']
    for col, baslik in enumerate(basliklar, 1):
        hucre = ws.cell(row=1, column=col, value=baslik)
        hucre.font = Font(bold=True, color='FFFFFF')
        hucre.fill = PatternFill(fill_type='solid', fgColor='2d7a3a')
        hucre.alignment = Alignment(horizontal='center')
        ws.column_dimensions[hucre.column_letter].width = 20
    ornek = ['Örnek A.Ş.', 'Makine Tedarikçisi', '1234567890', 'info@ornek.com', '0212 000 00 00', 'Ahmet Yılmaz', 'İstanbul', 'TL', '30', 'Makine']
    for col, deger in enumerate(ornek, 1):
        ws.cell(row=2, column=col, value=deger)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=tedarikci_sablonu.xlsx'
    return resp

@admin.route('/tedarikci/excel-yukle', methods=['POST'])
@login_required
@role_required('admin', 'satinalma')
def tedarikci_excel_yukle():
    from openpyxl import load_workbook
    dosya = request.files.get('excel_dosya')
    if not dosya or not dosya.filename.endswith('.xlsx'):
        flash('Geçerli bir .xlsx dosyası seçin.', 'danger')
        return redirect(url_for('admin.tedarikci_listesi'))
    try:
        wb = load_workbook(dosya)
        ws = wb.active
        eklenen, atlanan = 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            name = str(row[0]).strip()
            vergi_no = str(row[2]).strip() if row[2] else None
            if Tedarikci.query.filter_by(name=name).first():
                atlanan += 1
                continue
            t = Tedarikci(
                name=name,
                unvan=str(row[1]).strip() if row[1] else None,
                vergi_no=vergi_no,
                email=str(row[3]).strip() if row[3] else None,
                telefon=str(row[4]).strip() if row[4] else None,
                iletisim_kisi=str(row[5]).strip() if row[5] else None,
                adres=str(row[6]).strip() if row[6] else None,
                para_birimi=str(row[7]).strip() if row[7] else 'TL',
                vade_gun=int(row[8]) if row[8] else 30,
                kategori=str(row[9]).strip() if row[9] else None,
            )
            db.session.add(t)
            eklenen += 1
        db.session.commit()
        flash(f'{eklenen} tedarikçi eklendi. {atlanan} kayıt zaten mevcut olduğu için atlandı.', 'success')
    except Exception as e:
        flash(f'Dosya okunamadı: {e}', 'danger')
    return redirect(url_for('admin.tedarikci_listesi'))

@admin.route('/tedarikci/ekle', methods=['POST'])
@login_required
@role_required('admin', 'satinalma')
def tedarikci_ekle():
    t = Tedarikci(
        name=request.form.get('name'),
        unvan=request.form.get('unvan'),
        vergi_no=request.form.get('vergi_no'),
        email=request.form.get('email'),
        telefon=request.form.get('telefon'),
        adres=request.form.get('adres'),
        iletisim_kisi=request.form.get('iletisim_kisi'),
        para_birimi=request.form.get('para_birimi', 'TL'),
        vade_gun=int(request.form.get('vade_gun', 30)),
        kategori=request.form.get('kategori')
    )
    db.session.add(t)
    db.session.commit()
    flash('Tedarikçi eklendi.', 'success')
    return redirect(url_for('admin.tedarikci_listesi'))

@admin.route('/tedarikci/<int:t_id>/duzenle', methods=['POST'])
@login_required
@role_required('admin', 'satinalma')
def tedarikci_duzenle(t_id):
    t = db.get_or_404(Tedarikci, t_id)
    t.name = request.form.get('name')
    t.unvan = request.form.get('unvan')
    t.vergi_no = request.form.get('vergi_no')
    t.iletisim_kisi = request.form.get('iletisim_kisi')
    t.email = request.form.get('email')
    t.telefon = request.form.get('telefon')
    t.kategori = request.form.get('kategori')
    t.para_birimi = request.form.get('para_birimi', 'TL')
    t.vade_gun = int(request.form.get('vade_gun', 30))
    t.is_active = request.form.get('is_active') == '1'
    db.session.commit()
    flash(f'{t.name} güncellendi.', 'success')
    return redirect(url_for('admin.tedarikci_listesi'))

@admin.route('/tedarikci/<int:t_id>/sil', methods=['POST'])
@login_required
@role_required('admin', 'satinalma')
def tedarikci_sil(t_id):
    t = db.get_or_404(Tedarikci, t_id)
    t.is_active = False
    db.session.commit()
    flash(f'{t.name} pasife alındı.', 'warning')
    return redirect(url_for('admin.tedarikci_listesi'))

from flask import make_response
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.graphics.shapes import Drawing, String, Line
from reportlab.graphics import renderPDF
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io, os

_fonts_registered = False
def _register_fonts():
    global _fonts_registered
    if _fonts_registered:
        return
    base = os.path.join(os.path.dirname(__file__), 'static')
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', os.path.join(base, 'DejaVuSans.ttf')))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', os.path.join(base, 'DejaVuSans-Bold.ttf')))
        _fonts_registered = True
    except Exception:
        pass

@main.route('/profil', methods=['GET', 'POST'])
@login_required
def profil():
    if request.method == 'POST':
        aksiyon = request.form.get('aksiyon')
        if aksiyon == 'profil_guncelle':
            current_user.name = request.form.get('name') or current_user.name
            current_user.unvan = request.form.get('unvan') or None
            current_user.telefon = request.form.get('telefon') or None
            dogum_str = request.form.get('dogum_tarihi')
            if dogum_str:
                try:
                    from datetime import date
                    current_user.dogum_tarihi = date.fromisoformat(dogum_str)
                except ValueError:
                    pass
            current_user.unvan_pdf_goster = request.form.get('unvan_pdf_goster') == '1'
            current_user.bildirim_email = request.form.get('bildirim_email') == '1'
            pin = request.form.get('tablet_pin', '').strip()
            if pin and pin.isdigit() and len(pin) == 4:
                current_user.tablet_pin = pin
            elif pin == '':
                current_user.tablet_pin = None
            db.session.commit()
            flash('Profil güncellendi.', 'success')
        elif aksiyon == 'sifre_degistir':
            mevcut = request.form.get('mevcut_sifre')
            yeni = request.form.get('yeni_sifre')
            tekrar = request.form.get('yeni_sifre_tekrar')
            if not check_password_hash(current_user.password, mevcut):
                flash('Mevcut şifre hatalı.', 'danger')
            elif len(yeni) < 6:
                flash('Yeni şifre en az 6 karakter olmalıdır.', 'danger')
            elif yeni != tekrar:
                flash('Yeni şifreler eşleşmiyor.', 'danger')
            else:
                current_user.password = generate_password_hash(yeni)
                current_user.sifre_degisim_tarihi = datetime.utcnow()
                db.session.commit()
                flash('Şifreniz başarıyla değiştirildi.', 'success')
    from sqlalchemy import func
    toplam = TalepFormu.query.filter_by(talep_eden_id=current_user.id).count()
    durum_sayilari = dict(
        db.session.query(TalepFormu.durum, func.count(TalepFormu.id))
        .filter_by(talep_eden_id=current_user.id)
        .group_by(TalepFormu.durum).all()
    )
    bu_ay_baslangic = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    bu_ay = TalepFormu.query.filter(
        TalepFormu.talep_eden_id == current_user.id,
        TalepFormu.created_at >= bu_ay_baslangic
    ).count()
    return render_template('profil.html', toplam=toplam, durum_sayilari=durum_sayilari, bu_ay=bu_ay)

@admin.route('/kullanici/<int:user_id>/sifre-sifirla', methods=['POST'])
@login_required
@role_required('admin')
def kullanici_sifre_sifirla(user_id):
    user = db.get_or_404(User, user_id)
    yeni = request.form.get('yeni_sifre')
    if not yeni or len(yeni) < 6:
        flash('Şifre en az 6 karakter olmalıdır.', 'danger')
    else:
        user.password = generate_password_hash(yeni)
        user.sifre_degisim_tarihi = datetime.utcnow()
        db.session.commit()
        flash(f'{user.name} şifresi sıfırlandı.', 'success')
    return redirect(url_for('admin.kullanicilar'))

@main.route('/talep/<int:talep_id>/pdf')
@login_required
def talep_pdf(talep_id):
    from reportlab.pdfbase.pdfmetrics import stringWidth as sw
    _register_fonts()
    FONT = 'DejaVu' if _fonts_registered else 'Helvetica'
    FONT_BOLD = 'DejaVu-Bold' if _fonts_registered else 'Helvetica-Bold'

    talep = db.get_or_404(TalepFormu, talep_id)

    PAGE_W, PAGE_H = landscape(A4)
    LM = RM = 1*cm
    TM = 1.5*cm
    BM = 1*cm
    USABLE_W = PAGE_W - LM - RM

    # İmza tablosu canvas callback ile sayfanın sağ altına çizilecek
    IMZA_COL_W = 7.5*cm
    IMZA_W = IMZA_COL_W * 3
    imza_eden = (talep.talep_eden.name if talep.talep_eden else '') + \
                ('\n' + talep.talep_eden.unvan if talep.talep_eden and talep.talep_eden.unvan
                 and talep.talep_eden.unvan_pdf_goster else '')
    imza_data = [
        ['Talebi Oluşturan', 'Departman Müdürü Onayı', 'Genel Müdür Onayı'],
        ['\n\n\n' + imza_eden, '\n\n\nAd Soyad', '\n\n\nAd Soyad'],
        ['İmza / Tarih', 'İmza / Tarih', 'İmza / Tarih'],
    ]
    imza_style = TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 6),
    ])

    def draw_imza(canvas, doc):
        canvas.saveState()
        t = Table(imza_data, colWidths=[IMZA_COL_W]*3)
        t.setStyle(imza_style)
        w, h = t.wrapOn(canvas, IMZA_W, 6*cm)
        t.drawOn(canvas, PAGE_W - RM - w, BM)
        canvas.restoreState()

    # bottomMargin: imza tablosu için alan bırak (yaklaşık 4cm)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
        rightMargin=RM, leftMargin=LM,
        topMargin=TM, bottomMargin=BM + 4*cm)

    elements = []

    # ── Başlık: logo solda, başlık tam ortada ──
    LOGO_W = 6*cm
    title_style = ParagraphStyle('title', fontSize=13, fontName=FONT_BOLD, alignment=1)
    logo_drawing = Drawing(140, 40)
    logo_drawing.add(String(0, 14, 'ERLAU', fontSize=28, fontName='Helvetica-Bold', fillColor=colors.HexColor('#3a8a00')))
    logo_drawing.add(String(2, 4, 'EINE MARKE DER RUD GRUPPE', fontSize=7, fontName='Helvetica-Bold', fillColor=colors.black))

    header_data = [[logo_drawing, Paragraph('SATIN ALMA TALEP FORMU', title_style), '']]
    header_table = Table(header_data, colWidths=[LOGO_W, USABLE_W - 2*LOGO_W, LOGO_W])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*cm))

    # ── Sipariş bilgileri ──
    info_data = [
        ['Sipariş No:', talep.siparis_no, 'Tarih:', talep.created_at.strftime('%d.%m.%Y')],
        ['Departman:', talep.department.name if talep.department else '-',
         'Talep Eden:', talep.talep_eden.name if talep.talep_eden else '-'],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 7*cm, 3*cm, 7*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,0), (0,-1), FONT_BOLD),
        ('FONTNAME', (2,0), (2,-1), FONT_BOLD),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (2,0), (2,-1), colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    info_table.hAlign = 'LEFT'
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*cm))

    # ── Malzeme tablosu: sütun genişliği içeriğe göre otomatik ──
    HEADERS = ['#', 'Malzeme Adı', 'Marka/Model', 'Tür', 'Birim', 'Miktar', 'Hedef', 'KW', 'Açıklama', 'Son Alım', 'Anlık Stok']
    PAD = 10  # hücre dolgusu (point)

    table_data = [HEADERS]
    for i, kalem in enumerate(talep.kalemler):
        son_alim = kalem.son_alinma_tarihi.strftime('%d.%m.%Y') if kalem.son_alinma_tarihi else '-'
        table_data.append([
            str(i+1),
            kalem.malzeme_adi or '',
            kalem.marka_model or '',
            kalem.malzeme_turu or '',
            kalem.birim or '',
            str(kalem.miktar or ''),
            kalem.hedef or '',
            kalem.kw or '',
            kalem.aciklama or '',
            son_alim,
            kalem.anlik_stok or '',
        ])

    # Her sütun için maksimum içerik genişliği hesapla
    col_count = len(HEADERS)
    col_widths = []
    for j in range(col_count):
        max_w = 0
        for r, row in enumerate(table_data):
            fn = FONT_BOLD if r == 0 else FONT
            w = sw(str(row[j]), fn, 8) + PAD
            max_w = max(max_w, w)
        col_widths.append(max_w)

    # Her zaman tam sayfa genişliğine orantılı ölçekle
    total = sum(col_widths)
    col_widths = [w * USABLE_W / total for w in col_widths]

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2d7a3a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('PADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    main_table.hAlign = 'LEFT'
    elements.append(main_table)

    doc.build(elements, onFirstPage=draw_imza, onLaterPages=draw_imza)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=talep_{talep.siparis_no}.pdf'
    return response

@satin_alma.route('/fiyatlandir/<int:talep_id>', methods=['GET', 'POST'])
@login_required
@role_required('satinalma', 'admin')
def fiyatlandir(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    tedarikciler = Tedarikci.query.filter_by(is_active=True).order_by(Tedarikci.name).all()
    if request.method == 'POST':
        for kalem in talep.kalemler:
            prefix = f'kalem_{kalem.id}_'
            br_fiyat = request.form.get(prefix + 'br_fiyat')
            miktar = request.form.get(prefix + 'miktar')
            kalem.tedarikci_id = request.form.get(prefix + 'tedarikci_id') or None
            kalem.br_fiyat = float(br_fiyat) if br_fiyat else None
            if miktar:
                kalem.miktar = float(miktar)
            kalem.toplam_fiyat = round(kalem.br_fiyat * kalem.miktar, 2) if kalem.br_fiyat and kalem.miktar else None
            kalem.para_birimi = request.form.get(prefix + 'para_birimi', 'TL')
            vade = request.form.get(prefix + 'vade_gun')
            termin = request.form.get(prefix + 'termin_gun')
            kalem.vade_gun = int(vade) if vade else None
            kalem.termin_gun = int(termin) if termin else None
        # Onaylı/yolda siparişlerde durumu düşürme, sadece bekleyen/fiyatlandırılmış için güncelle
        if talep.durum not in ('onaylandi', 'yolda', 'teslim_alindi'):
            talep.durum = 'fiyatlandirildi'
        db.session.commit()
        flash('Fiyatlar güncellendi.', 'success')
        return redirect(url_for('satin_alma.panel'))
    return render_template('fiyatlandir.html', talep=talep, tedarikciler=tedarikciler)

@satin_alma.route('/kalem/<int:kalem_id>/duzenle', methods=['POST'])
@login_required
def kalem_duzenle(kalem_id):
    kalem = db.get_or_404(TalepKalem, kalem_id)
    talep = kalem.talep
    is_satinalma = current_user.role in ['satinalma', 'admin']
    is_talep_eden = talep.talep_eden_id == current_user.id

    if not is_satinalma and not is_talep_eden:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('main.talep_detay', talep_id=talep.id))

    malzeme_adi = request.form.get('malzeme_adi', '').strip()
    if malzeme_adi:
        kalem.malzeme_adi = malzeme_adi

    marka_model = request.form.get('marka_model', '').strip()
    kalem.marka_model = marka_model or kalem.marka_model

    malzeme_turu = request.form.get('malzeme_turu', '').strip()
    if malzeme_turu:
        kalem.malzeme_turu = malzeme_turu

    birim = request.form.get('birim', '').strip()
    if birim:
        kalem.birim = birim

    hedef = request.form.get('hedef', '').strip()
    if hedef in ['siparis', 'stok', 'proje']:
        kalem.hedef = hedef

    miktar_str = request.form.get('miktar', '').strip()
    if miktar_str:
        try:
            kalem.miktar = float(miktar_str)
        except ValueError:
            flash('Geçersiz miktar.', 'danger')
            return redirect(url_for('main.talep_detay', talep_id=talep.id))

    if is_satinalma:
        br_fiyat_str = request.form.get('br_fiyat', '').strip()
        if br_fiyat_str:
            try:
                kalem.br_fiyat = float(br_fiyat_str)
            except ValueError:
                flash('Geçersiz birim fiyat.', 'danger')
                return redirect(url_for('main.talep_detay', talep_id=talep.id))

    if kalem.br_fiyat and kalem.miktar:
        kalem.toplam_fiyat = round(kalem.br_fiyat * kalem.miktar, 2)

    db.session.commit()
    flash('Kalem güncellendi.', 'success')
    next_url = request.form.get('next', '')
    return redirect(next_url if next_url else url_for('main.talep_detay', talep_id=talep.id))

@satin_alma.route('/siparis/<int:talep_id>')
@login_required
@role_required('satinalma', 'admin')
def siparis_ozet(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    gruplar = {}
    atamasiz = []
    for kalem in talep.kalemler:
        if kalem.tedarikci:
            tid = kalem.tedarikci_id
            if tid not in gruplar:
                gruplar[tid] = {'tedarikci': kalem.tedarikci, 'kalemler': [], 'toplam': 0}
            gruplar[tid]['kalemler'].append(kalem)
            if kalem.toplam_fiyat:
                gruplar[tid]['toplam'] += kalem.toplam_fiyat
        else:
            atamasiz.append(kalem)
    cc_emails = [u.email for u in User.query.filter(
        User.role == 'satinalma', User.id != current_user.id, User.is_active == True
    ).all()]
    cc = ','.join(cc_emails)
    # Orijinal (Almanca) isim haritası
    tum_malzeme_adlari = {k.malzeme_adi for g in gruplar.values() for k in g['kalemler']}
    tum_malzeme_adlari |= {k.malzeme_adi for k in atamasiz}
    orijinal_map = {}
    for adi in tum_malzeme_adlari:
        m = Malzeme.query.filter(Malzeme.malzeme_adi == adi, Malzeme.aciklama.isnot(None)).first()
        if m and m.aciklama and 'Orijinal: ' in m.aciklama:
            orijinal_map[adi] = m.aciklama.split('Orijinal: ')[1].split(' | ')[0]
    return render_template('siparis_ozet.html', talep=talep, gruplar=gruplar, atamasiz=atamasiz, cc=cc, orijinal_map=orijinal_map)

@satin_alma.route('/siparis/<int:talep_id>/excel/<int:tedarikci_id>')
@login_required
@role_required('satinalma', 'admin')
def siparis_excel(talep_id, tedarikci_id):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    talep = db.get_or_404(TalepFormu, talep_id)
    tedarikci = db.get_or_404(Tedarikci, tedarikci_id)
    kalemler = [k for k in talep.kalemler if k.tedarikci_id == tedarikci_id]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sipariş'

    yesil = PatternFill(fill_type='solid', fgColor='2d7a3a')
    acik = PatternFill(fill_type='solid', fgColor='f0f7f1')
    kalin = Font(bold=True)
    beyaz_kalin = Font(bold=True, color='FFFFFF')
    ince = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )
    orta = Alignment(horizontal='center', vertical='center')

    # Başlık
    ws.merge_cells('A1:J1')
    ws['A1'] = 'ERLAU — SİPARİŞ FORMU'
    ws['A1'].font = Font(bold=True, size=14, color='2d7a3a')
    ws['A1'].alignment = orta
    ws.row_dimensions[1].height = 28

    # Bilgi satırları
    bilgiler = [
        ('Sipariş No', talep.siparis_no), ('Tarih', datetime.utcnow().strftime('%d.%m.%Y')),
        ('Tedarikçi', tedarikci.name), ('İletişim Kişisi', tedarikci.iletisim_kisi or '-'),
        ('E-posta', tedarikci.email or '-'), ('Telefon', tedarikci.telefon or '-'),
    ]
    for i, (label, val) in enumerate(bilgiler, 2):
        ws.cell(row=i, column=1, value=label).font = kalin
        ws.cell(row=i, column=2, value=val)
    ws.row_dimensions[8].height = 8

    # Tablo başlığı
    basliklar = ['#', 'Malzeme Adı', 'Marka/Model', 'Tür', 'Birim', 'Miktar', 'Br. Fiyat', 'Toplam', 'Para Birimi', 'Termin (gün)']
    for col, b in enumerate(basliklar, 1):
        c = ws.cell(row=9, column=col, value=b)
        c.font = beyaz_kalin
        c.fill = yesil
        c.alignment = orta
        c.border = ince

    genislikler = [4, 30, 20, 15, 8, 8, 12, 12, 12, 12]
    for i, g in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(i)].width = g

    # Kalemler
    toplam_genel = 0
    for idx, k in enumerate(kalemler, 1):
        row = 9 + idx
        fill = acik if idx % 2 == 0 else PatternFill()
        satirlar = [idx, k.malzeme_adi, k.marka_model or '', k.malzeme_turu or '',
                    k.birim or '', k.miktar or '', k.br_fiyat or '', k.toplam_fiyat or '', k.para_birimi or '', k.termin_gun or '']
        for col, val in enumerate(satirlar, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = ince
            if fill.fill_type:
                c.fill = fill
        if k.toplam_fiyat:
            toplam_genel += k.toplam_fiyat

    # Toplam satırı
    t_row = 9 + len(kalemler) + 1
    ws.cell(row=t_row, column=7, value='TOPLAM').font = kalin
    ws.cell(row=t_row, column=8, value=toplam_genel).font = kalin
    ws.cell(row=t_row, column=9, value=kalemler[0].para_birimi if kalemler else '').font = kalin

    # Not alanı
    ws.cell(row=t_row+2, column=1, value='Not / Açıklama:').font = kalin
    ws.merge_cells(f'B{t_row+2}:J{t_row+3}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import make_response
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    tr_map = str.maketrans('çğıöşüÇĞİÖŞÜ', 'cgiosucgiosu')
    guvenli_ad = tedarikci.name.translate(tr_map).replace(' ', '_')
    from urllib.parse import quote
    utf8_ad = quote(f'siparis_{talep.siparis_no}_{tedarikci.name}.xlsx')
    resp.headers['Content-Disposition'] = f"attachment; filename=\"siparis_{talep.siparis_no}_{guvenli_ad}.xlsx\"; filename*=UTF-8''{utf8_ad}"
    return resp

@satin_alma.route('/yolda/<int:talep_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def yolda(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    talep.durum = 'yolda'
    db.session.commit()
    flash('Sipariş yolda olarak işaretlendi.', 'success')
    return redirect(url_for('satin_alma.panel'))

@satin_alma.route('/teslim/<int:talep_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teslim(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    talep.durum = 'teslim_alindi'
    db.session.commit()
    flash('Sipariş teslim alındı olarak işaretlendi.', 'success')
    return redirect(url_for('satin_alma.panel'))

@satin_alma.route('/durum/<int:talep_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def durum_guncelle(talep_id):
    talep = db.get_or_404(TalepFormu, talep_id)
    yeni_durum = request.form.get('durum')
    gecerli_durumlar = ['bekliyor', 'fiyatlandirildi', 'onaylandi', 'yolda', 'teslim_alindi', 'iptal']
    if yeni_durum in gecerli_durumlar:
        talep.durum = yeni_durum
        if yeni_durum == 'yolda' and not talep.yolda_tarihi:
            talep.yolda_tarihi = datetime.utcnow()
        if yeni_durum == 'teslim_alindi':
            now = datetime.utcnow()
            for kalem in talep.kalemler:
                kalem.son_alinma_tarihi = now
                kalem.son_siparis_no = talep.siparis_no
        db.session.commit()
        flash('Durum güncellendi.', 'success')
    return redirect(url_for('satin_alma.panel'))

@main.route('/api/son-alim')
@login_required
def son_alim_api():
    malzeme_adi = request.args.get('malzeme_adi', '').strip()
    if not malzeme_adi:
        return jsonify(None)
    from sqlalchemy import func
    kalem = (TalepKalem.query
             .filter(func.lower(TalepKalem.malzeme_adi) == func.lower(malzeme_adi))
             .filter(TalepKalem.son_alinma_tarihi.isnot(None))
             .order_by(TalepKalem.son_alinma_tarihi.desc())
             .first())
    if kalem:
        return jsonify({
            'tarih': kalem.son_alinma_tarihi.strftime('%d.%m.%Y'),
            'siparis_no': kalem.son_siparis_no or ''
        })
    return jsonify(None)

@main.route('/api/kullanim-sikligi')
@login_required
def kullanim_sikligi_api():
    from sqlalchemy import func
    malzeme_adi = request.args.get('malzeme_adi', '').strip()
    if not malzeme_adi:
        return jsonify(None)
    toplam = (TalepKalem.query
              .join(TalepFormu)
              .filter(func.lower(TalepKalem.malzeme_adi) == func.lower(malzeme_adi))
              .filter(TalepFormu.durum == 'teslim_alindi')
              .count())
    return jsonify({'toplam': toplam})

# ─── TEKLİF SİSTEMİ ──────────────────────────────────────────────────────────

@satin_alma.route('/teklifler')
@login_required
@role_required('satinalma', 'admin')
def teklif_listesi():
    gruplar = (TeklifGrubu.query
               .options(selectinload(TeklifGrubu.talep_kalem), selectinload(TeklifGrubu.kalemler))
               .order_by(TeklifGrubu.created_at.desc())
               .all())
    return render_template('teklifler.html', gruplar=gruplar)


@satin_alma.route('/teklif/yeni/<int:kalem_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_yeni(kalem_id):
    kalem = TalepKalem.query.get_or_404(kalem_id)
    mevcut = TeklifGrubu.query.filter_by(talep_kalem_id=kalem_id).first()
    if mevcut:
        return redirect(url_for('satin_alma.teklif_detay', grup_id=mevcut.id))
    grup = TeklifGrubu(teklif_no=generate_teklif_no(), talep_kalem_id=kalem_id)
    db.session.add(grup)
    db.session.commit()
    return redirect(url_for('satin_alma.teklif_detay', grup_id=grup.id))


@satin_alma.route('/teklif/toplu-yeni', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_toplu_yeni():
    import uuid as _uuid
    kalem_ids = request.form.getlist('kalem_ids')
    if not kalem_ids:
        flash('Kalem seçilmedi.', 'warning')
        return redirect(request.referrer or url_for('satin_alma.panel'))

    batch_id = str(_uuid.uuid4())
    for kid in kalem_ids:
        try:
            kid = int(kid)
        except ValueError:
            continue
        kalem = TalepKalem.query.get(kid)
        if not kalem:
            continue
        mevcut = TeklifGrubu.query.filter_by(talep_kalem_id=kid).first()
        if mevcut:
            if not mevcut.batch_id:
                mevcut.batch_id = batch_id
            else:
                batch_id = mevcut.batch_id
            continue
        grup = TeklifGrubu(teklif_no=generate_teklif_no(), talep_kalem_id=kid, batch_id=batch_id)
        db.session.add(grup)
    db.session.commit()
    return redirect(url_for('satin_alma.teklif_toplu', batch_id=batch_id))


@satin_alma.route('/teklif/toplu/<batch_id>')
@login_required
@role_required('satinalma', 'admin')
def teklif_toplu(batch_id):
    gruplar = (TeklifGrubu.query
               .filter_by(batch_id=batch_id)
               .options(selectinload(TeklifGrubu.talep_kalem),
                        selectinload(TeklifGrubu.kalemler).selectinload(TeklifKalem.tedarikci))
               .all())
    if not gruplar:
        flash('Toplu teklif bulunamadı.', 'warning')
        return redirect(url_for('satin_alma.panel'))
    tedarikciler = Tedarikci.query.filter_by(is_active=True).order_by(Tedarikci.name).all()
    talep = gruplar[0].talep_kalem.talep if gruplar[0].talep_kalem else None
    return render_template('toplu_teklif.html', gruplar=gruplar, tedarikciler=tedarikciler,
                           talep=talep, batch_id=batch_id)


@satin_alma.route('/teklif/toplu/<batch_id>/ekle', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_toplu_ekle(batch_id):
    gruplar = TeklifGrubu.query.filter_by(batch_id=batch_id).all()
    tedarikci_id = request.form.get('tedarikci_id') or None
    para_birimi = request.form.get('para_birimi', 'TL')
    vade_gun = request.form.get('vade_gun') or None
    kaynak = 'manuel'
    for grup in gruplar:
        fiyat_str = request.form.get(f'fiyat_{grup.id}', '').strip()
        if not fiyat_str:
            continue
        try:
            birim_fiyat = float(fiyat_str)
        except ValueError:
            continue
        kalem = TeklifKalem(
            grup_id=grup.id,
            tedarikci_id=int(tedarikci_id) if tedarikci_id else None,
            birim_fiyat=birim_fiyat,
            para_birimi=para_birimi,
            vade_gun=int(vade_gun) if vade_gun else None,
            kaynak=kaynak,
        )
        db.session.add(kalem)
    db.session.commit()
    flash('Teklifler kaydedildi.', 'success')
    return redirect(url_for('satin_alma.teklif_toplu', batch_id=batch_id))


@satin_alma.route('/teklif/toplu/<batch_id>/excel')
@login_required
@role_required('satinalma', 'admin')
def teklif_toplu_excel(batch_id):
    """Toplu RFQ Excel — tüm batch kalemleri tek sayfada."""
    import io
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    gruplar = (TeklifGrubu.query
               .filter_by(batch_id=batch_id)
               .options(selectinload(TeklifGrubu.talep_kalem))
               .all())
    if not gruplar:
        flash('Toplu teklif bulunamadı.', 'warning')
        return redirect(url_for('satin_alma.panel'))

    talep = gruplar[0].talep_kalem.talep if gruplar[0].talep_kalem else None
    siparis_no = talep.siparis_no if talep else ''

    YESIL = "1B5E20"; ACIK_YESIL = "C8E6C9"; GRI = "F5F5F5"
    KOYU_GRI = "455A64"; BEYAZ = "FFFFFF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teklif Talebi"

    def stil(ws, cell, bold=False, size=11, color=None, bg=None, align='left',
             wrap=False, border=False, italic=False):
        c = ws[cell] if isinstance(cell, str) else cell
        c.font = Font(bold=bold, size=size, color=color or "000000", name="Calibri", italic=italic)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if border:
            thin = Side(style='thin', color='BDBDBD')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, w in [('A',5),('B',30),('C',16),('D',10),('E',10),('F',16),('G',14),('H',14),('I',22)]:
        ws.column_dimensions[col].width = w

    # Başlık
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 14

    ws.merge_cells('B2:D4')
    ws['B2'] = 'ERLAU'
    ws['B2'].font = Font(bold=True, size=28, color=YESIL, name='Calibri')
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B2'].fill = PatternFill("solid", fgColor=BEYAZ)

    ws.merge_cells('E2:I2')
    ws['E2'] = 'TEKLİF TALEBİ / REQUEST FOR QUOTATION'
    stil(ws, 'E2', bold=True, size=14, color=BEYAZ, bg=YESIL, align='center')

    ws.merge_cells('E3:I3')
    ws['E3'] = f'Tarih / Date: {date.today().strftime("%d.%m.%Y")}'
    stil(ws, 'E3', size=10, color=KOYU_GRI, bg=GRI, align='right')

    ws.merge_cells('E4:I4')
    ws['E4'] = f'Sipariş: {siparis_no}  |  Kalem Sayısı: {len(gruplar)}'
    stil(ws, 'E4', size=10, color=KOYU_GRI, bg=GRI, align='right')

    ws.merge_cells('B6:I6')
    ws['B6'] = 'ERLAU Metal San ve Tic LTD ŞTİ | erlau.com.tr'
    stil(ws, 'B6', size=9, color='757575', italic=True)

    ws.row_dimensions[9].height = 10
    ws.row_dimensions[10].height = 26
    ws.merge_cells('B10:I10')
    ws['B10'] = 'TALEP EDİLEN MALZEMELER / REQUESTED ITEMS'
    stil(ws, 'B10', bold=True, size=12, color=BEYAZ, bg=YESIL, align='center')

    hdrs = ['#', 'Malzeme Adı', 'Marka/Model', 'Miktar', 'Birim', 'Teknik Resim', 'Standart', 'Açıklama']
    ws.row_dimensions[11].height = 20
    for i, h in enumerate(hdrs):
        c = ws.cell(row=11, column=i+1, value=h)
        stil(ws, c, bold=True, size=9, color=BEYAZ, bg=KOYU_GRI, align='center', border=True)

    for idx, grup in enumerate(gruplar):
        k = grup.talep_kalem
        row = 12 + idx
        ws.row_dimensions[row].height = 30
        vals = [idx+1, k.malzeme_adi if k else '', k.marka_model or '' if k else '',
                k.miktar or '' if k else '', k.birim or '' if k else '',
                k.teknik_resim_kodu or '' if k else '', k.standart or '' if k else '',
                k.aciklama or '' if k else '']
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=i+1, value=v)
            bg = ACIK_YESIL if i == 1 else GRI
            stil(ws, c, size=10, bg=bg, align='center' if i not in [1,2,7] else 'left', wrap=True, border=True)

    # Tedarikçi doldurma alanı
    teklif_row_start = 13 + len(gruplar)
    ws.row_dimensions[teklif_row_start - 1].height = 10
    ws.row_dimensions[teklif_row_start].height = 26
    ws.merge_cells(f'B{teklif_row_start}:I{teklif_row_start}')
    ws[f'B{teklif_row_start}'] = 'TEDARİKÇİ TEKLİF ALANI  ← Lütfen doldurunuz / Please fill in'
    stil(ws, f'B{teklif_row_start}', bold=True, size=11, color=YESIL, bg='FFF9C4', align='center')

    th_row = teklif_row_start + 1
    ws.row_dimensions[th_row].height = 32
    t_hdrs = ['#', 'Malzeme Adı', 'Birim Fiyat\n(Unit Price)', 'Para Birimi\n(Currency)',
              'Toplam\n(Total)', 'Vade (gün)\n(Payment Days)', 'Termin (gün)\n(Lead Days)',
              'Teklif Geçerlilik\n(Validity)', 'Notlar / Notes']
    for i, h in enumerate(t_hdrs):
        c = ws.cell(row=th_row, column=i+1, value=h)
        stil(ws, c, bold=True, size=9, color=BEYAZ, bg="1565C0", align='center', wrap=True, border=True)

    for idx, grup in enumerate(gruplar):
        k = grup.talep_kalem
        row = th_row + 1 + idx
        ws.row_dimensions[row].height = 24
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.value = idx+1 if col == 1 else (k.malzeme_adi if col == 2 and k else '')
            stil(ws, c, size=10, bg="E3F2FD" if col > 2 else GRI, align='center', border=True)

    note_row = th_row + 1 + len(gruplar) + 1
    ws.merge_cells(f'B{note_row}:I{note_row}')
    ws[f'B{note_row}'] = f'Lütfen teklifinizi {siparis_no} referans numarası ile gönderin. | ERLAU Satınalma Departmanı'
    stil(ws, f'B{note_row}', size=9, color='757575', italic=True)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    dosya_adi = f'Toplu_Teklif_{siparis_no or batch_id[:8]}.xlsx'
    return send_file(output, as_attachment=True, download_name=dosya_adi,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@satin_alma.route('/teklif/<int:grup_id>')
@login_required
@role_required('satinalma', 'admin')
def teklif_detay(grup_id):
    grup = (TeklifGrubu.query
            .options(selectinload(TeklifGrubu.kalemler).selectinload(TeklifKalem.tedarikci),
                     selectinload(TeklifGrubu.talep_kalem))
            .get_or_404(grup_id))
    tedarikciler = Tedarikci.query.filter_by(is_active=True).order_by(Tedarikci.name).all()
    return render_template('teklif_detay.html', grup=grup, tedarikciler=tedarikciler)


@satin_alma.route('/teklif/<int:grup_id>/ekle', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_kalem_ekle(grup_id):
    grup = TeklifGrubu.query.get_or_404(grup_id)
    birim_fiyat = request.form.get('birim_fiyat')
    vade_gun = request.form.get('vade_gun')
    kaynak = request.form.get('kaynak', 'manuel')
    if kaynak not in ('manuel', 'pdf', 'excel', 'mail'):
        kaynak = 'manuel'
    kalem = TeklifKalem(
        grup_id=grup_id,
        tedarikci_id=request.form.get('tedarikci_id') or None,
        birim_fiyat=float(birim_fiyat) if birim_fiyat else None,
        para_birimi=request.form.get('para_birimi', 'TL'),
        vade_gun=int(vade_gun) if vade_gun else None,
        notlar=request.form.get('notlar'),
        kaynak=kaynak,
    )
    db.session.add(kalem)
    if grup.durum == 'bekliyor':
        grup.durum = 'teklif_alindi'
    db.session.commit()
    return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))


@satin_alma.route('/teklif/kalem/<int:kalem_id>/red-nedeni', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_kalem_red_nedeni(kalem_id):
    kalem = TeklifKalem.query.get_or_404(kalem_id)
    kalem.red_nedeni = request.form.get('red_nedeni', '').strip() or None
    db.session.commit()
    return jsonify({'ok': True})


@satin_alma.route('/teklif/<int:grup_id>/sec/<int:kalem_id>', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_sec(grup_id, kalem_id):
    grup = TeklifGrubu.query.get_or_404(grup_id)
    # Seçilmeyen alınan tekliflerde red nedeni zorunlu
    eksik = [k for k in grup.kalemler
             if k.id != kalem_id and k.birim_fiyat is not None and not k.red_nedeni]
    if eksik:
        tedarikci_listesi = ', '.join(k.tedarikci.name if k.tedarikci else '?' for k in eksik)
        flash(f'Seçilmeyen teklifler için red nedeni girilmedi: {tedarikci_listesi}', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))
    for k in grup.kalemler:
        k.secildi = False
    kazanan = TeklifKalem.query.get_or_404(kalem_id)
    kazanan.secildi = True
    grup.durum = 'secildi'
    talep_kalem = grup.talep_kalem
    talep_kalem.br_fiyat = kazanan.birim_fiyat
    talep_kalem.para_birimi = kazanan.para_birimi
    talep_kalem.vade_gun = kazanan.vade_gun
    talep_kalem.tedarikci_id = kazanan.tedarikci_id
    if kazanan.birim_fiyat and talep_kalem.miktar:
        talep_kalem.toplam_fiyat = kazanan.birim_fiyat * talep_kalem.miktar
    db.session.commit()
    flash('Teklif seçildi, sipariş kalemi güncellendi.', 'success')
    return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))


@satin_alma.route('/teklif/kalem/<int:kalem_id>/sil', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_kalem_sil(kalem_id):
    kalem = TeklifKalem.query.get_or_404(kalem_id)
    grup_id = kalem.grup_id
    db.session.delete(kalem)
    db.session.commit()
    return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))


@satin_alma.route('/teklif/<int:grup_id>/ai-yukle', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_ai_yukle(grup_id):
    import os, tempfile
    from app.teklif_ai import teklif_oku
    grup = TeklifGrubu.query.get_or_404(grup_id)
    dosya = request.files.get('teklif_pdf')
    if not dosya or not dosya.filename.lower().endswith('.pdf'):
        flash('Lütfen geçerli bir PDF dosyası yükleyin.', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))

    malzeme_adi = grup.talep_kalem.malzeme_adi if grup.talep_kalem else None
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    try:
        dosya.save(tmp.name)
        veri = teklif_oku(tmp.name, malzeme_adi=malzeme_adi)
    except Exception as e:
        flash(f'AI analiz hatası: {e}', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))
    finally:
        os.unlink(tmp.name)

    tedarikci_id = None
    tedarikci_adi = veri.get('tedarikci_adi') or ''
    if tedarikci_adi:
        from sqlalchemy import func
        t = Tedarikci.query.filter(
            func.lower(Tedarikci.name).contains(tedarikci_adi.lower()[:20])
        ).first()
        if t:
            tedarikci_id = t.id

    kalem = TeklifKalem(
        grup_id=grup_id,
        tedarikci_id=tedarikci_id,
        birim_fiyat=veri.get('birim_fiyat'),
        para_birimi=veri.get('para_birimi') or 'TL',
        vade_gun=veri.get('vade_gun'),
        notlar=(veri.get('notlar') or '') + (f' [AI — tedarikçi: {tedarikci_adi}]' if tedarikci_adi and not tedarikci_id else ''),
        kaynak='pdf',
    )
    db.session.add(kalem)
    if grup.durum == 'bekliyor':
        grup.durum = 'teklif_alindi'
    db.session.commit()
    flash(f'AI teklif analizi tamamlandı — güven skoru: {veri.get("guven_skoru", 0):.0%}', 'success')
    return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))


# ─── YENİ TEKLİF ÖZELLİKLERİ ────────────────────────────────────────────────

@satin_alma.route('/kalem/<int:kalem_id>/cogalt', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def kalem_cogalt(kalem_id):
    """Bir talep kalemini en fazla 4 kopyaya çoğaltır (farklı tedarikçi için)."""
    kaynak = TalepKalem.query.get_or_404(kalem_id)
    # parent_kalem_id'si olan kopyayı çoğaltmaya izin verme
    ana_id = kaynak.parent_kalem_id or kaynak.id
    mevcut_kopya = TalepKalem.query.filter_by(parent_kalem_id=ana_id).count()
    if mevcut_kopya >= 3:  # orijinal + 3 kopya = 4
        return jsonify({'ok': False, 'hata': 'Bir kalem için en fazla 3 kopya oluşturulabilir.'}), 400
    kopya = TalepKalem(
        talep_id=kaynak.talep_id,
        parent_kalem_id=ana_id,
        malzeme_adi=kaynak.malzeme_adi,
        marka_model=kaynak.marka_model,
        malzeme_turu=kaynak.malzeme_turu,
        birim=kaynak.birim,
        miktar=kaynak.miktar,
        hedef=kaynak.hedef,
        kullanim_amaci=kaynak.kullanim_amaci,
        kullanilan_alan=kaynak.kullanilan_alan,
        proje_makine=kaynak.proje_makine,
        aciklama=kaynak.aciklama,
        teknik_resim_kodu=kaynak.teknik_resim_kodu,
        standart=kaynak.standart,
    )
    db.session.add(kopya)
    db.session.commit()
    return jsonify({'ok': True, 'kopya_id': kopya.id})


@satin_alma.route('/kalem/<int:kalem_id>/kopya-sil', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def kalem_kopya_sil(kalem_id):
    kalem = TalepKalem.query.get_or_404(kalem_id)
    if not kalem.parent_kalem_id:
        return jsonify({'ok': False, 'hata': 'Orijinal kalem silinemez'}), 400
    for grup in list(kalem.teklif_gruplari):
        db.session.delete(grup)
    db.session.flush()
    db.session.delete(kalem)
    db.session.commit()
    return jsonify({'ok': True})


@satin_alma.route('/teklif/<int:grup_id>/excel')
@login_required
@role_required('satinalma', 'admin')
def teklif_excel(grup_id):
    """Tedarikçiye gönderilecek RFQ Excel şablonu oluşturur."""
    import io, os
    from flask import send_file
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from datetime import date

    grup = TeklifGrubu.query.get_or_404(grup_id)
    kalem = grup.talep_kalem
    talep = kalem.talep if kalem else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teklif Talebi"

    # Renk paleti
    YESIL = "1B5E20"
    ACIK_YESIL = "C8E6C9"
    GRI = "F5F5F5"
    KOYU_GRI = "455A64"
    BEYAZ = "FFFFFF"

    def stil(ws, cell, bold=False, size=11, color=None, bg=None, align='left',
             wrap=False, border=False, italic=False):
        c = ws[cell] if isinstance(cell, str) else cell
        c.font = Font(bold=bold, size=size, color=color or "000000",
                      name="Calibri", italic=italic)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center',
                                 wrap_text=wrap)
        if border:
            thin = Side(style='thin', color='BDBDBD')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Sütun genişlikleri
    for col, w in [('A',5),('B',28),('C',14),('D',10),('E',16),('F',14),
                   ('G',14),('H',14),('I',22)]:
        ws.column_dimensions[col].width = w

    # Başlık alanı (1-5. satırlar)
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 14

    ws.merge_cells('B2:D4')
    ws['B2'] = 'ERLAU'
    ws['B2'].font = Font(bold=True, size=28, color=YESIL, name='Calibri')
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B2'].fill = PatternFill("solid", fgColor=BEYAZ)

    ws.merge_cells('E2:I2')
    ws['E2'] = 'TEKLİF TALEBİ / REQUEST FOR QUOTATION'
    stil(ws, 'E2', bold=True, size=14, color=BEYAZ, bg=YESIL, align='center')

    ws.merge_cells('E3:I3')
    ws['E3'] = f'Tarih / Date: {date.today().strftime("%d.%m.%Y")}'
    stil(ws, 'E3', size=10, color=KOYU_GRI, bg=GRI, align='right')

    ws.merge_cells('E4:I4')
    teklif_no = grup.teklif_no or 'TKL-XXX'
    siparis_no = talep.siparis_no if talep else ''
    ws['E4'] = f'Teklif No / RFQ No: {teklif_no}  |  Sipariş: {siparis_no}'
    stil(ws, 'E4', size=10, color=KOYU_GRI, bg=GRI, align='right')

    # Firma Bilgisi
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 14

    ws.merge_cells('B6:I6')
    ws['B6'] = 'ERLAU Metal San ve Tic LTD ŞTİ | erlau.com.tr'
    stil(ws, 'B6', size=9, color='757575', italic=True)

    # Boşluk
    ws.row_dimensions[9].height = 10

    # Malzeme Başlığı
    ws.row_dimensions[10].height = 26
    ws.merge_cells('B10:I10')
    ws['B10'] = 'TALEP EDİLEN MALZEME / REQUESTED ITEM'
    stil(ws, 'B10', bold=True, size=12, color=BEYAZ, bg=YESIL, align='center')

    # Malzeme Detayları başlıkları
    headers = ['#', 'Malzeme Adı / Material', 'Marka/Model', 'Miktar', 'Birim',
               'Teknik Resim', 'Standart', 'Açıklama']
    ws.row_dimensions[11].height = 20
    for i, h in enumerate(headers):
        c = ws.cell(row=11, column=i+1, value=h)
        stil(ws, c, bold=True, size=9, color=BEYAZ, bg=KOYU_GRI,
             align='center', border=True)

    # Malzeme Satırı
    ws.row_dimensions[12].height = 40
    vals = [
        1,
        kalem.malzeme_adi if kalem else '',
        kalem.marka_model or '' if kalem else '',
        kalem.miktar or '',
        kalem.birim or '' if kalem else '',
        kalem.teknik_resim_kodu or '' if kalem else '',
        kalem.standart or '' if kalem else '',
        kalem.aciklama or '' if kalem else '',
    ]
    for i, v in enumerate(vals):
        c = ws.cell(row=12, column=i+1, value=v)
        bg = ACIK_YESIL if i == 1 else GRI
        stil(ws, c, size=10, bg=bg, align='center' if i not in [1,2,7] else 'left',
             wrap=True, border=True)

    # Boşluk
    ws.row_dimensions[13].height = 10

    # TEDARİKÇİ DOLDURMA ALANI
    ws.row_dimensions[14].height = 26
    ws.merge_cells('B14:I14')
    ws['B14'] = 'TEDARİKÇİ TEKLİF ALANI / SUPPLIER QUOTATION AREA  ← Lütfen doldurunuz / Please fill in'
    stil(ws, 'B14', bold=True, size=11, color=YESIL, bg='FFF9C4', align='center')

    teklif_headers = ['#', 'Malzeme Adı', 'Birim Fiyat\n(Unit Price)', 'Para Birimi\n(Currency)',
                       'Toplam\n(Total)', 'Vade (gün)\n(Payment Days)',
                       'Termin (gün)\n(Lead Days)', 'Teklif Geçerlilik\n(Validity)', 'Notlar / Notes']
    ws.row_dimensions[15].height = 32
    for i, h in enumerate(teklif_headers):
        c = ws.cell(row=15, column=i+1, value=h)
        stil(ws, c, bold=True, size=9, color=BEYAZ, bg="1565C0",
             align='center', wrap=True, border=True)

    # Tedarikçi doldurma satırı (1 satır)
    ws.row_dimensions[16].height = 24
    for col in range(1, 10):
        c = ws.cell(row=16, column=col)
        c.value = 1 if col == 1 else (kalem.malzeme_adi if col == 2 and kalem else '')
        stil(ws, c, size=10, bg="E3F2FD" if col > 2 else GRI,
             align='center', border=True)

    # Son not
    ws.row_dimensions[17].height = 10
    ws.row_dimensions[18].height = 18
    ws.merge_cells('B18:I18')
    ws['B18'] = f'Lütfen teklifinizi {teklif_no} referans numarası ile gönderin. | Please reply with reference no: {teklif_no}'
    stil(ws, 'B18', size=9, color='757575', italic=True)

    ws.row_dimensions[19].height = 16
    ws.merge_cells('B19:I19')
    ws['B19'] = 'ERLAU Metal San ve Tic LTD ŞTİ | Satınalma Departmanı | satinalma@erlau.com.tr'
    stil(ws, 'B19', size=9, color='757575', italic=True)

    # Print area
    ws.print_area = 'A1:J20'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Konu başlığını kaydet
    if not grup.konu_basligi:
        malzeme_kisa = (kalem.malzeme_adi[:30] if kalem else 'Malzeme').replace(' ', '_')
        grup.konu_basligi = f'[ERLAU TEKLIF] {teklif_no} - {kalem.malzeme_adi[:40] if kalem else ""}'
        db.session.commit()

    dosya_adi = f'Teklif_Talebi_{teklif_no}.xlsx'
    return send_file(output, as_attachment=True, download_name=dosya_adi,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@satin_alma.route('/teklif/<int:grup_id>/mailto')
@login_required
@role_required('satinalma', 'admin')
def teklif_mailto(grup_id):
    """Tedarikçi için Outlook taslağı oluşturur (mailto: redirect)."""
    import urllib.parse
    grup = TeklifGrubu.query.get_or_404(grup_id)
    kalem = grup.talep_kalem
    talep = kalem.talep if kalem else None
    tedarikci_id = request.args.get('tedarikci_id', type=int)

    tedarikci = Tedarikci.query.get(tedarikci_id) if tedarikci_id else None
    email = tedarikci.email if tedarikci else ''

    konu = grup.konu_basligi or f'[ERLAU TEKLIF] {grup.teklif_no} - {kalem.malzeme_adi[:40] if kalem else ""}'
    if not grup.konu_basligi:
        grup.konu_basligi = konu
        db.session.commit()

    malzeme_adi = kalem.malzeme_adi if kalem else '-'
    miktar = f"{kalem.miktar} {kalem.birim or ''}" if kalem and kalem.miktar else '-'
    siparis_no = talep.siparis_no if talep else '-'
    tedarikci_adi = tedarikci.name if tedarikci else 'Sayın Yetkili'

    govde = f"""Sayın {tedarikci_adi},

Aşağıda belirtilen malzeme/malzemeler için fiyat teklifinizi bekliyoruz.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TALEP BİLGİSİ / REQUEST INFO
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Referans No  : {konu}
Sipariş No   : {siparis_no}
Malzeme      : {malzeme_adi}
Miktar       : {miktar}
{f"Marka/Model  : {kalem.marka_model}" if kalem and kalem.marka_model else ""}
{f"Teknik Resim : {kalem.teknik_resim_kodu}" if kalem and kalem.teknik_resim_kodu else ""}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Lütfen teklifinizde aşağıdakileri belirtiniz:
  ✓ Birim fiyat ve para birimi
  ✓ Vade gün sayısı
  ✓ Termin / Teslim süresi
  ✓ Teklif geçerlilik süresi

Ekte RFQ (Teklif Talep) Excel formu bulunmaktadır.
Teklifinizi bu form üzerinden veya mail yanıtı olarak iletebilirsiniz.

Teşekkürler,
Erlau Satınalma Departmanı
"""

    mailto_url = f"mailto:{email}?subject={urllib.parse.quote(konu)}&body={urllib.parse.quote(govde)}"
    from flask import redirect as flask_redirect
    return flask_redirect(mailto_url)


@satin_alma.route('/teklif/<int:grup_id>/mail-oku', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_mail_oku(grup_id):
    """AI ile gelen teklif mailini parse eder."""
    grup = TeklifGrubu.query.get_or_404(grup_id)
    kalem = grup.talep_kalem
    mail_icerik = request.form.get('mail_icerik', '').strip()
    if not mail_icerik:
        return jsonify({'ok': False, 'hata': 'Mail içeriği boş'}), 400
    try:
        import anthropic, os, json as _json
        client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'), timeout=10.0)
        prompt = f"""Aşağıdaki tedarikçi teklif mailini analiz et ve JSON döndür.

Talep edilen malzeme: {kalem.malzeme_adi if kalem else 'bilinmiyor'}
Talep miktarı: {f'{kalem.miktar} {kalem.birim}' if kalem and kalem.miktar else 'bilinmiyor'}

Mail içeriği:
---
{mail_icerik[:3000]}
---

Sadece JSON döndür, başka hiçbir şey yazma:
{{
  "tedarikci_adi": "string veya null",
  "tedarikci_email": "string veya null",
  "birim_fiyat": number veya null,
  "para_birimi": "TL|EUR|USD",
  "vade_gun": number veya null,
  "termin_gun": number veya null,
  "teklif_gecerlilik_gun": number veya null,
  "notlar": "string",
  "guven_skoru": 0.0-1.0
}}"""
        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=500,
            messages=[{'role': 'user', 'content': prompt}]
        )
        metin = msg.content[0].text.strip()
        import re
        m = re.search(r'\{.*\}', metin, re.DOTALL)
        veri = _json.loads(m.group()) if m else {}
    except Exception as e:
        return jsonify({'ok': False, 'hata': str(e)}), 500

    # Tedarikçi eşleştir
    tedarikci_id = None
    if veri.get('tedarikci_email'):
        t = Tedarikci.query.filter_by(email=veri['tedarikci_email']).first()
        if t: tedarikci_id = t.id
    if not tedarikci_id and veri.get('tedarikci_adi'):
        from sqlalchemy import func
        t = Tedarikci.query.filter(
            func.lower(Tedarikci.name).contains(veri['tedarikci_adi'].lower()[:20])
        ).first()
        if t: tedarikci_id = t.id

    return jsonify({
        'ok': True,
        'tedarikci_id': tedarikci_id,
        'tedarikci_adi': veri.get('tedarikci_adi'),
        'birim_fiyat': veri.get('birim_fiyat'),
        'para_birimi': veri.get('para_birimi', 'TL'),
        'vade_gun': veri.get('vade_gun'),
        'termin_gun': veri.get('termin_gun'),
        'notlar': veri.get('notlar', ''),
        'guven': veri.get('guven_skoru', 0.5)
    })


@satin_alma.route('/teklif/<int:grup_id>/ai-tavsiye')
@login_required
@role_required('satinalma', 'admin')
def teklif_ai_tavsiye(grup_id):
    """AI en iyi teklifi analiz eder ve tavsiye verir."""
    grup = TeklifGrubu.query.get_or_404(grup_id)
    if len(grup.kalemler) < 2:
        return jsonify({'ok': False, 'tavsiye': 'Karşılaştırma için en az 2 teklif gereklidir.'})
    kalem = grup.talep_kalem
    miktar = kalem.miktar if kalem and kalem.miktar else 1
    teklif_listesi_str = ""
    for i, tk in enumerate(grup.kalemler):
        toplam = (tk.birim_fiyat or 0) * miktar
        teklif_listesi_str += f"{i+1}. {tk.tedarikci.name if tk.tedarikci else 'Bilinmiyor'}: {tk.birim_fiyat} {tk.para_birimi}, Vade: {tk.vade_gun or '?'} gün, Toplam: {toplam:.2f} {tk.para_birimi}, Not: {tk.notlar or '-'}\n"
    try:
        import anthropic, os
        client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'), timeout=10.0)
        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=300,
            messages=[{'role': 'user', 'content': f"""Satın alma teklifleri karşılaştırması:
Malzeme: {kalem.malzeme_adi if kalem else 'bilinmiyor'}, Miktar: {miktar} {kalem.birim if kalem else ''}

{teklif_listesi_str}

Kısa ve net değerlendir (3-4 cümle): En uygun seçenek hangisi ve neden? TL cinsinden toplam maliyet ve vade süresini de göz önünde bulundur. Türkçe yaz."""}]
        )
        tavsiye = msg.content[0].text.strip()
    except Exception as e:
        tavsiye = f'AI analiz yapılamadı: {e}'
    return jsonify({'ok': True, 'tavsiye': tavsiye})


# ─── SİPARİŞ FORMU (PO) ──────────────────────────────────────────────────────

def _generate_po_no():
    from datetime import date
    prefix = f"PO-{date.today().strftime('%Y%m%d')}"
    son = TeklifGrubu.query.filter(TeklifGrubu.po_no.like(f'{prefix}%')).count()
    return f"{prefix}-{son+1:03d}"


def _build_po_pdf_bytes(po_no, talep_no, gonderen_adi, tedarikci,
                        malzeme_adi, marka_model, miktar, birim,
                        birim_fiyat, para_birimi, vade_gun, notlar):
    """Ortak PO PDF üretici. bytes döner."""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, HRFlowable
    from reportlab.graphics.renderPDF import Drawing
    from reportlab.graphics.shapes import String
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    import io as _io

    _register_fonts()
    FONT     = 'DejaVu'      if _fonts_registered else 'Helvetica'
    FONT_B   = 'DejaVu-Bold' if _fonts_registered else 'Helvetica-Bold'

    buffer = _io.BytesIO()
    PAGE_W_PT = A4[0] - 3*cm
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=2*cm)

    N    = ParagraphStyle('n',  fontName=FONT,   fontSize=9,  leading=13)
    B    = ParagraphStyle('b',  fontName=FONT_B, fontSize=9,  leading=13)
    BIG  = ParagraphStyle('bg', fontName=FONT_B, fontSize=14, leading=18)
    GRAY = ParagraphStyle('g',  fontName=FONT,   fontSize=8,  leading=11,
                          textColor=colors.grey)
    elements = []

    # Başlık
    logo = Drawing(120, 36)
    logo.add(String(0, 12, 'ERLAU', fontSize=26, fontName='Helvetica-Bold',
                    fillColor=colors.HexColor('#3a8a00')))
    logo.add(String(1,  3, 'EINE MARKE DER RUD GRUPPE', fontSize=6,
                    fontName='Helvetica-Bold', fillColor=colors.black))
    hdr = Table([[logo, Paragraph('SİPARİŞ FORMU', BIG)]],
                colWidths=[5*cm, PAGE_W_PT - 5*cm])
    hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (1,0), (1,0),  'RIGHT'),
    ]))
    elements += [hdr, HRFlowable(width='100%', thickness=2,
                                  color=colors.HexColor('#3a8a00'), spaceAfter=8)]

    # PO Bilgileri + Tedarikçi
    po_tarih = datetime.utcnow().strftime('%d.%m.%Y')
    sol = Table([
        [Paragraph('<b>Sipariş No (PO):</b>', B), Paragraph(po_no,        N)],
        [Paragraph('<b>Talep No:</b>',         B), Paragraph(talep_no,    N)],
        [Paragraph('<b>Tarih:</b>',             B), Paragraph(po_tarih,    N)],
        [Paragraph('<b>Hazırlayan:</b>',        B), Paragraph(gonderen_adi, N)],
    ], colWidths=[4.5*cm, 5.5*cm])
    sol.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),FONT),
                              ('FONTSIZE',(0,0),(-1,-1),9),
                              ('PADDING',(0,0),(-1,-1),3)]))

    ted_html = f"<b>{tedarikci.name if tedarikci else '—'}</b>"
    if tedarikci:
        if tedarikci.unvan:    ted_html += f"<br/>{tedarikci.unvan}"
        if tedarikci.adres:    ted_html += f"<br/>{tedarikci.adres}"
        if tedarikci.email:    ted_html += f"<br/>{tedarikci.email}"
        if tedarikci.telefon:  ted_html += f"<br/>{tedarikci.telefon}"
        if tedarikci.vergi_no: ted_html += f"<br/>Vergi No: {tedarikci.vergi_no}"
    sag = Table([[Paragraph('<b>TEDARİKÇİ</b>', B)],
                 [Paragraph(ted_html, N)]], colWidths=[8*cm])
    sag.setStyle(TableStyle([
        ('BOX',        (0,0),(-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0),(0,0),   colors.HexColor('#f0fdf4')),
        ('PADDING',    (0,0),(-1,-1), 5),
    ]))
    iki = Table([[sol, sag]], colWidths=[10*cm, PAGE_W_PT - 10*cm])
    iki.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),
                              ('PADDING',(0,0),(-1,-1),0)]))
    elements += [iki, Spacer(1, 0.5*cm)]

    # Malzeme tablosu
    toplam = (birim_fiyat or 0) * (miktar or 0)
    mal = Table([
        ['#', 'Malzeme Adı', 'Marka/Model', 'Miktar', 'Birim',
         'Br. Fiyat', 'Para Birimi', 'Toplam'],
        ['1', malzeme_adi or '-', marka_model or '-',
         str(miktar or '-'), birim or '-',
         f"{birim_fiyat:.2f}" if birim_fiyat else '-',
         para_birimi or 'TL',
         f"{toplam:.2f}" if birim_fiyat and miktar else '-'],
    ], colWidths=[0.8*cm, 5.5*cm, 3*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm, 2.2*cm])
    mal.setStyle(TableStyle([
        ('FONTNAME',    (0,0),(-1,-1), FONT),
        ('FONTNAME',    (0,0),(-1,0),  FONT_B),
        ('FONTSIZE',    (0,0),(-1,-1), 9),
        ('BACKGROUND',  (0,0),(-1,0),  colors.HexColor('#2d7a3a')),
        ('TEXTCOLOR',   (0,0),(-1,0),  colors.white),
        ('GRID',        (0,0),(-1,-1), 0.5, colors.grey),
        ('PADDING',     (0,0),(-1,-1), 5),
        ('ALIGN',       (3,0),(-1,-1), 'CENTER'),
        ('ALIGN',       (5,1),(-1,-1), 'RIGHT'),
    ]))
    elements += [mal, Spacer(1, 0.3*cm)]

    # Toplam satırı
    tot = Table([['','','','','','',
                  Paragraph('<b>GENEL TOPLAM:</b>', B),
                  Paragraph(f"<b>{toplam:.2f} {para_birimi or 'TL'}</b>", B)]],
                colWidths=[0.8*cm, 5.5*cm, 3*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm, 2.2*cm])
    tot.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),FONT),
                              ('FONTSIZE',(0,0),(-1,-1),9),
                              ('ALIGN',(6,0),(7,0),'RIGHT'),
                              ('PADDING',(0,0),(-1,-1),4)]))
    elements += [tot, Spacer(1, 0.4*cm)]

    # Koşullar
    kos = Table([
        [Paragraph('<b>Ödeme Vadesi:</b>', B),
         Paragraph(f"{vade_gun or '-'} gün", N),
         Paragraph('<b>Notlar:</b>', B),
         Paragraph(notlar or '-', N)],
    ], colWidths=[3.5*cm, 4*cm, 2.5*cm, PAGE_W_PT - 10*cm])
    kos.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),FONT), ('FONTSIZE',(0,0),(-1,-1),9),
        ('BOX',(0,0),(-1,-1),0.5,colors.grey),
        ('INNERGRID',(0,0),(-1,-1),0.3,colors.HexColor('#e5e7eb')),
        ('PADDING',(0,0),(-1,-1),5),
    ]))
    elements += [kos, Spacer(1, 0.4*cm)]

    # Teslimat adresi
    adr = Table([
        [Paragraph('<b>Teslimat Adresi:</b>', B),
         Paragraph('ERLAU Metal San ve Tic LTD ŞTİ<br/>'
                   'Organize Sanayi Bölgesi, Bursa / TÜRKİYE<br/>'
                   'satinalma@erlau.com.tr', N)]
    ], colWidths=[3.5*cm, PAGE_W_PT - 3.5*cm])
    adr.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),FONT), ('FONTSIZE',(0,0),(-1,-1),9),
        ('BOX',(0,0),(-1,-1),0.5,colors.grey), ('PADDING',(0,0),(-1,-1),5),
        ('BACKGROUND',(0,0),(0,0),colors.HexColor('#f9fafb')),
    ]))
    elements += [adr, Spacer(1, 1*cm)]

    # İmza
    imz = Table([
        ['Hazırlayan', 'Onaylayan'],
        [f'\n\n\n{gonderen_adi}', '\n\n\n'],
        ['İmza / Tarih', 'İmza / Tarih'],
    ], colWidths=[PAGE_W_PT/2, PAGE_W_PT/2])
    imz.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),FONT), ('FONTSIZE',(0,0),(-1,-1),9),
        ('FONTNAME',(0,0),(-1,0),FONT_B),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
        ('PADDING',(0,0),(-1,-1),6),
    ]))
    elements.append(imz)

    doc.build(elements)
    return buffer.getvalue()


def _build_po_eml(to_email, subject, body, pdf_bytes, pdf_filename):
    """PDF ekli EML dosyası üretir. Kullanıcı çift tıklayınca Outlook açılır."""
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    from email.utils import formatdate
    msg = MIMEMultipart('mixed')
    msg['From']    = 'satinalma@erlau.com.tr'
    msg['To']      = to_email or ''
    msg['Subject'] = subject
    msg['Date']    = formatdate(localtime=True)
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    pdf_part = MIMEApplication(pdf_bytes, _subtype='pdf')
    pdf_part.add_header('Content-Disposition', 'attachment', filename=pdf_filename)
    msg.attach(pdf_part)
    return msg.as_bytes()


@satin_alma.route('/teklif/<int:grup_id>/po-pdf')
@login_required
@role_required('satinalma', 'admin')
def teklif_po_pdf(grup_id):
    grup = TeklifGrubu.query.get_or_404(grup_id)
    if grup.durum != 'secildi':
        flash('Önce bir teklif seçilmelidir.', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))
    kazanan = next((k for k in grup.kalemler if k.secildi), None)
    if not kazanan:
        flash('Seçili teklif bulunamadı.', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))
    if not grup.po_no:
        grup.po_no = _generate_po_no()
        db.session.commit()
    tk = grup.talep_kalem
    talep_no = tk.talep.siparis_no if tk and tk.talep else '-'
    pdf = _build_po_pdf_bytes(
        po_no=grup.po_no, talep_no=talep_no, gonderen_adi=current_user.name,
        tedarikci=kazanan.tedarikci,
        malzeme_adi=tk.malzeme_adi if tk else '-',
        marka_model=tk.marka_model if tk else '-',
        miktar=tk.miktar if tk else 0, birim=tk.birim if tk else '-',
        birim_fiyat=kazanan.birim_fiyat, para_birimi=kazanan.para_birimi,
        vade_gun=kazanan.vade_gun, notlar=kazanan.notlar,
    )
    resp = make_response(pdf)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename=PO_{grup.po_no}.pdf'
    return resp


@satin_alma.route('/teklif/<int:grup_id>/po-eml')
@login_required
@role_required('satinalma', 'admin')
def teklif_po_eml(grup_id):
    """PDF ekli Outlook taslağı (EML) indir — teklif akışı."""
    grup = TeklifGrubu.query.get_or_404(grup_id)
    if grup.durum != 'secildi':
        flash('Önce bir teklif seçilmelidir.', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))
    kazanan = next((k for k in grup.kalemler if k.secildi), None)
    if not kazanan:
        flash('Seçili teklif yok.', 'danger')
        return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))
    if not grup.po_no:
        grup.po_no = _generate_po_no()
        db.session.commit()
    tk = grup.talep_kalem
    talep_no = tk.talep.siparis_no if tk and tk.talep else '-'
    tedarikci = kazanan.tedarikci
    malzeme = tk.malzeme_adi if tk else '-'
    pdf_bytes = _build_po_pdf_bytes(
        po_no=grup.po_no, talep_no=talep_no, gonderen_adi=current_user.name,
        tedarikci=tedarikci,
        malzeme_adi=malzeme, marka_model=tk.marka_model if tk else '-',
        miktar=tk.miktar if tk else 0, birim=tk.birim if tk else '-',
        birim_fiyat=kazanan.birim_fiyat, para_birimi=kazanan.para_birimi,
        vade_gun=kazanan.vade_gun, notlar=kazanan.notlar,
    )
    konu = f"[ERLAU SİPARİŞ] {grup.po_no} - {malzeme[:40]}"
    to_email = tedarikci.email if tedarikci else ''
    ted_adi = tedarikci.name if tedarikci else 'Sayın Yetkili'
    miktar_str = f"{tk.miktar} {tk.birim or ''}" if tk and tk.miktar else '-'
    fiyat_str = f"{kazanan.birim_fiyat:.2f} {kazanan.para_birimi}" if kazanan.birim_fiyat else '-'
    govde = (f"Sayın {ted_adi},\n\nAşağıda detayları yer alan siparişimizi iletiyoruz.\n\n"
             f"{'━'*34}\nSİPARİŞ BİLGİSİ / PURCHASE ORDER\n{'━'*34}\n"
             f"PO Numarası : {grup.po_no}\nTalep No    : {talep_no}\n"
             f"Malzeme     : {malzeme}\nMiktar      : {miktar_str}\n"
             f"Birim Fiyat : {fiyat_str}\nVade        : {kazanan.vade_gun or '-'} gün\n"
             f"{'━'*34}\n\nSipariş formu ekte yer almaktadır.\n"
             f"Onayınızı ve tahmini teslimat tarihinizi bekliyoruz.\n\n"
             f"Saygılarımızla,\n{current_user.name}\nErlau Satınalma Departmanı\nsatinalma@erlau.com.tr")
    eml = _build_po_eml(to_email, konu, govde, pdf_bytes, f"PO_{grup.po_no}.pdf")
    resp = make_response(eml)
    resp.headers['Content-Type'] = 'message/rfc822'
    resp.headers['Content-Disposition'] = f'attachment; filename="SiparisFormu_{grup.po_no}.eml"'
    return resp


@satin_alma.route('/kalem/<int:kalem_id>/po-eml')
@login_required
@role_required('satinalma', 'admin')
def kalem_po_eml(kalem_id):
    """PDF ekli Outlook taslağı (EML) — teklif almadan direkt sipariş."""
    kalem = TalepKalem.query.get_or_404(kalem_id)
    if not kalem.tedarikci_id or not kalem.br_fiyat:
        flash('Tedarikçi ve fiyat girilmiş olmalıdır.', 'danger')
        talep_id = kalem.talep_id
        return redirect(url_for('satin_alma.fiyatlandir', talep_id=talep_id))
    tedarikci = kalem.tedarikci
    talep = kalem.talep
    from datetime import date
    po_no = f"PO-{date.today().strftime('%Y%m%d')}-{kalem.id}"
    talep_no = talep.siparis_no if talep else '-'
    pdf_bytes = _build_po_pdf_bytes(
        po_no=po_no, talep_no=talep_no, gonderen_adi=current_user.name,
        tedarikci=tedarikci,
        malzeme_adi=kalem.malzeme_adi, marka_model=kalem.marka_model,
        miktar=kalem.miktar, birim=kalem.birim,
        birim_fiyat=kalem.br_fiyat, para_birimi=kalem.para_birimi or 'TL',
        vade_gun=kalem.vade_gun, notlar=kalem.aciklama,
    )
    konu = f"[ERLAU SİPARİŞ] {po_no} - {kalem.malzeme_adi[:40]}"
    to_email = tedarikci.email if tedarikci else ''
    ted_adi = tedarikci.name if tedarikci else 'Sayın Yetkili'
    miktar_str = f"{kalem.miktar} {kalem.birim or ''}" if kalem.miktar else '-'
    fiyat_str = f"{kalem.br_fiyat:.2f} {kalem.para_birimi or 'TL'}"
    govde = (f"Sayın {ted_adi},\n\nAşağıda detayları yer alan siparişimizi iletiyoruz.\n\n"
             f"{'━'*34}\nSİPARİŞ BİLGİSİ / PURCHASE ORDER\n{'━'*34}\n"
             f"PO Numarası : {po_no}\nTalep No    : {talep_no}\n"
             f"Malzeme     : {kalem.malzeme_adi}\nMiktar      : {miktar_str}\n"
             f"Birim Fiyat : {fiyat_str}\nVade        : {kalem.vade_gun or '-'} gün\n"
             f"{'━'*34}\n\nSipariş formu ekte yer almaktadır.\n"
             f"Onayınızı ve tahmini teslimat tarihinizi bekliyoruz.\n\n"
             f"Saygılarımızla,\n{current_user.name}\nErlau Satınalma Departmanı\nsatinalma@erlau.com.tr")
    eml = _build_po_eml(to_email, konu, govde, pdf_bytes, f"PO_{po_no}.pdf")
    resp = make_response(eml)
    resp.headers['Content-Type'] = 'message/rfc822'
    resp.headers['Content-Disposition'] = f'attachment; filename="SiparisFormu_{po_no}.eml"'
    return resp


@satin_alma.route('/teklif/<int:grup_id>/po-gonder', methods=['POST'])
@login_required
@role_required('satinalma', 'admin')
def teklif_po_gonder(grup_id):
    grup = TeklifGrubu.query.get_or_404(grup_id)
    if not grup.po_no:
        grup.po_no = _generate_po_no()
    grup.po_tarihi = datetime.utcnow()
    grup.po_gonderen_id = current_user.id
    db.session.commit()
    flash(f'Sipariş {grup.po_no} gönderildi olarak kaydedildi.', 'success')
    return redirect(url_for('satin_alma.teklif_detay', grup_id=grup_id))


# ─── MUHASEBE ────────────────────────────────────────────────────────────────

IZINLI_MUHASEBE = ['muhasebe', 'satinalma', 'admin']

@muhasebe.route('/faturalar')
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_listesi():
    durum_filtre = request.args.get('durum', '')
    q = request.args.get('q', '').strip()
    query = Fatura.query
    if durum_filtre:
        query = query.filter_by(durum=durum_filtre)
    if q:
        from sqlalchemy import or_
        query = query.filter(or_(
            Fatura.fatura_no.ilike(f'%{q}%'),
            Fatura.tedarikci_adi_ham.ilike(f'%{q}%'),
        ))
    faturalar = query.order_by(Fatura.yukleme_tarihi.desc()).all()
    ozet = {
        'toplam': len(faturalar),
        'bekliyor': sum(1 for f in faturalar if f.durum == 'bekliyor'),
        'odendi': sum(1 for f in faturalar if f.durum == 'odendi'),
        'iptal_iade': sum(1 for f in faturalar if f.durum in ['iptal', 'iade']),
        'toplam_tutar': sum(f.genel_toplam or 0 for f in faturalar if f.durum not in ['iptal', 'iade']),
    }
    return render_template('fatura_listesi.html', faturalar=faturalar, ozet=ozet,
                           durum_filtre=durum_filtre, q=q, today=date.today())

@muhasebe.route('/fatura/yukle', methods=['GET', 'POST'])
@login_required
@role_required('muhasebe', 'admin')
def fatura_yukle():
    if request.method == 'POST':
        dosya = request.files.get('pdf_dosya')
        if not dosya or not dosya.filename.endswith('.pdf'):
            flash('Geçerli bir PDF dosyası seçin.', 'danger')
            return redirect(request.url)
        import os
        from werkzeug.utils import secure_filename
        dosya_adi = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{secure_filename(dosya.filename)}"
        kayit_dizini = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'faturalar')
        os.makedirs(kayit_dizini, exist_ok=True)
        tam_yol = os.path.join(kayit_dizini, dosya_adi)
        dosya.save(tam_yol)

        try:
            from app.fatura_ai import pdf_oku
            # Tedarikçi hafızası var mı kontrol et
            ted_id = request.form.get('tedarikci_id')
            ornek = None
            if ted_id:
                sablon = TedarikciSablon.query.filter_by(tedarikci_id=int(ted_id)).first()
                if sablon:
                    ornek = sablon.ornek_json
            ai_sonuc = pdf_oku(tam_yol, ornek)
        except Exception as e:
            os.remove(tam_yol)
            flash(f'AI analizi başarısız: {e}', 'danger')
            return redirect(request.url)

        import json
        from app.tcmb import get_kur
        para_birimi = ai_sonuc.get('para_birimi', 'TL')
        genel_toplam = ai_sonuc.get('genel_toplam')
        fatura_turu = request.form.get('fatura_turu', 'normal')
        ana_fatura_id = request.form.get('ana_fatura_id') or None

        # TCMB kuru otomatik çek
        fatura_kuru = None
        tl_karsiligi = None
        if para_birimi != 'TL':
            fatura_kuru = get_kur(para_birimi)
            if fatura_kuru and genel_toplam:
                tl_karsiligi = round(float(genel_toplam) * fatura_kuru, 2)

        fatura = Fatura(
            fatura_no=ai_sonuc.get('fatura_no'),
            tedarikci_adi_ham=ai_sonuc.get('tedarikci_adi'),
            ara_toplam=ai_sonuc.get('ara_toplam'),
            iskonto_tutari=ai_sonuc.get('iskonto_tutari'),
            iskonto_orani=ai_sonuc.get('iskonto_orani'),
            kdv_tutari=ai_sonuc.get('kdv_tutari'),
            genel_toplam=genel_toplam,
            para_birimi=para_birimi,
            dosya_yolu=dosya_adi,
            ai_ham_veri=json.dumps(ai_sonuc, ensure_ascii=False),
            ai_guvenskoru=ai_sonuc.get('guven_skoru', 0.5),
            yukleyen_id=current_user.id,
            durum='bekliyor',
            fatura_turu=fatura_turu,
            ana_fatura_id=int(ana_fatura_id) if ana_fatura_id else None,
            fatura_kuru=fatura_kuru,
            tl_karsiligi=tl_karsiligi,
        )
        if ai_sonuc.get('fatura_tarihi'):
            try:
                from datetime import date
                fatura.fatura_tarihi = date.fromisoformat(ai_sonuc['fatura_tarihi'])
            except: pass
        if ai_sonuc.get('vade_tarihi'):
            try:
                from datetime import date
                fatura.vade_tarihi = date.fromisoformat(ai_sonuc['vade_tarihi'])
            except: pass
        if ted_id:
            fatura.tedarikci_id = int(ted_id)
        db.session.add(fatura)
        db.session.flush()

        for k in ai_sonuc.get('kalemler', []):
            fk = FaturaKalem(
                fatura_id=fatura.id,
                malzeme_adi=k.get('malzeme_adi'),
                miktar=k.get('miktar'),
                birim=k.get('birim'),
                liste_fiyati=k.get('liste_fiyati'),
                iskonto_orani=k.get('iskonto_orani'),
                iskonto_tutari=k.get('iskonto_tutari'),
                br_fiyat=k.get('br_fiyat'),
                kdv_orani=k.get('kdv_orani'),
                toplam_fiyat=k.get('toplam_fiyat'),
            )
            db.session.add(fk)

        db.session.commit()
        flash(f'Fatura yüklendi ve AI analiz etti. Lütfen kontrol edin.', 'success')
        return redirect(url_for('muhasebe.fatura_detay', fatura_id=fatura.id))

    from app.tcmb import kur_listesi
    tedarikciler = Tedarikci.query.filter_by(is_active=True).order_by(Tedarikci.name).all()
    diger_faturalar = Fatura.query.filter_by(fatura_turu='normal').order_by(Fatura.yukleme_tarihi.desc()).limit(100).all()
    return render_template('fatura_yukle.html', tedarikciler=tedarikciler,
                           guncel_kurlar=kur_listesi(), diger_faturalar=diger_faturalar)

@muhasebe.route('/fatura/<int:fatura_id>')
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_detay(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    talepler = TalepFormu.query.filter(
        TalepFormu.durum.in_(['onaylandi', 'yolda', 'teslim_alindi'])
    ).order_by(TalepFormu.created_at.desc()).limit(50).all()
    tedarikciler = Tedarikci.query.filter_by(is_active=True).order_by(Tedarikci.name).all()
    talep_kalemleri = fatura.talep.kalemler if fatura.talep else []
    from datetime import date
    return render_template('fatura_detay.html', fatura=fatura, talepler=talepler,
                           tedarikciler=tedarikciler, today=date.today(),
                           talep_kalemleri=talep_kalemleri)

@muhasebe.route('/fatura/<int:fatura_id>/kalem-esles', methods=['POST'])
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_kalem_esles_manuel(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    from app.models import TalepKalem
    from app.fatura_ai import net_birim_fiyat, hafizaya_kaydet

    for fk in fatura.kalemler:
        tk_id_str = request.form.get(f'kalem_{fk.id}_talep_kalem_id', '')
        tk_id = int(tk_id_str) if tk_id_str.isdigit() else None
        fk.talep_kalem_id = tk_id

        if tk_id:
            tk = TalepKalem.query.get(tk_id)
            if tk:
                fatura_net = net_birim_fiyat({
                    'br_fiyat': fk.br_fiyat,
                    'liste_fiyati': fk.liste_fiyati,
                    'iskonto_orani': fk.iskonto_orani,
                    'iskonto_tutari': fk.iskonto_tutari,
                    'miktar': fk.miktar,
                    'toplam_fiyat': fk.toplam_fiyat,
                    'kdv_orani': fk.kdv_orani,
                })
                siparis_net = tk.br_fiyat or 0
                if siparis_net > 0 and fatura_net > 0:
                    fark = abs(fatura_net - siparis_net)
                    yuzde = fark / siparis_net * 100
                    if yuzde > 5:
                        fk.eslesme_durumu = 'fiyat_farki'
                        fk.eslesme_notu = (f"Siparişteki net br. fiyat: {siparis_net:.2f} | "
                                           f"Faturadaki net br. fiyat: {fatura_net:.2f} | Fark: %{yuzde:.1f}")
                    else:
                        fk.eslesme_durumu = 'eslesti'
                        fk.eslesme_notu = 'Manuel eşleştirme'
                else:
                    fk.eslesme_durumu = 'eslesti'
                    fk.eslesme_notu = 'Manuel eşleştirme'
                if fk.malzeme_adi and tk.malzeme_adi:
                    hafizaya_kaydet(fk.malzeme_adi, tk.malzeme_adi)
        else:
            fk.eslesme_durumu = 'eslesmiyor'
            fk.eslesme_notu = ''

    db.session.commit()
    flash('Kalem eşleştirmeleri kaydedildi ve hafızaya öğretildi.', 'success')
    return redirect(url_for('muhasebe.fatura_detay', fatura_id=fatura_id))


@muhasebe.route('/fatura/<int:fatura_id>/guncelle', methods=['POST'])
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_guncelle(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    fatura.fatura_no = request.form.get('fatura_no')
    fatura.tedarikci_adi_ham = request.form.get('tedarikci_adi_ham')
    fatura.ara_toplam = request.form.get('ara_toplam') or None
    fatura.kdv_tutari = request.form.get('kdv_tutari') or None
    fatura.genel_toplam = request.form.get('genel_toplam') or None
    fatura.para_birimi = request.form.get('para_birimi', 'TL')
    fatura.notlar = request.form.get('notlar')
    ted_id = request.form.get('tedarikci_id')
    fatura.tedarikci_id = int(ted_id) if ted_id else None
    tarih_str = request.form.get('fatura_tarihi')
    vade_str = request.form.get('vade_tarihi')
    if tarih_str:
        try:
            from datetime import date
            fatura.fatura_tarihi = date.fromisoformat(tarih_str)
        except: pass
    if vade_str:
        try:
            from datetime import date
            fatura.vade_tarihi = date.fromisoformat(vade_str)
        except: pass

    # Kalemleri güncelle
    for kalem in fatura.kalemler:
        prefix = f'kalem_{kalem.id}_'
        kalem.malzeme_adi = request.form.get(prefix + 'malzeme_adi', kalem.malzeme_adi)
        kalem.miktar = request.form.get(prefix + 'miktar') or kalem.miktar
        kalem.br_fiyat = request.form.get(prefix + 'br_fiyat') or kalem.br_fiyat
        kalem.toplam_fiyat = request.form.get(prefix + 'toplam_fiyat') or kalem.toplam_fiyat

    # Tedarikçi hafızasını güncelle
    if fatura.tedarikci_id:
        import json
        sablon = TedarikciSablon.query.filter_by(tedarikci_id=fatura.tedarikci_id).first()
        if not sablon:
            sablon = TedarikciSablon(tedarikci_id=fatura.tedarikci_id)
            db.session.add(sablon)
        sablon.ornek_json = json.dumps({
            'fatura_no': fatura.fatura_no,
            'fatura_tarihi': str(fatura.fatura_tarihi),
            'tedarikci_adi': fatura.tedarikci_adi_ham,
            'genel_toplam': fatura.genel_toplam,
            'para_birimi': fatura.para_birimi,
            'kalemler': [{'malzeme_adi': k.malzeme_adi, 'miktar': k.miktar,
                          'birim': k.birim, 'br_fiyat': k.br_fiyat} for k in fatura.kalemler]
        }, ensure_ascii=False)
        sablon.guncelleme_tarihi = datetime.utcnow()

    db.session.commit()
    flash('Fatura güncellendi ve tedarikçi hafızası iyileştirildi.', 'success')
    return redirect(url_for('muhasebe.fatura_detay', fatura_id=fatura.id))

@muhasebe.route('/fatura/<int:fatura_id>/durum', methods=['POST'])
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_durum(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    yeni_durum = request.form.get('durum')
    if yeni_durum in ['bekliyor', 'onaylandi', 'odendi', 'iptal', 'iade']:
        fatura.durum = yeni_durum
        if yeni_durum == 'odendi':
            from datetime import date
            fatura.odeme_tarihi = date.today()
            # Ödeme kuru işle (döviz faturalar için)
            odeme_kuru_str = request.form.get('odeme_kuru', '').strip()
            if odeme_kuru_str and fatura.para_birimi != 'TL':
                try:
                    fatura.odeme_kuru = float(odeme_kuru_str.replace(',', '.'))
                    if fatura.genel_toplam:
                        fatura.odenen_tl = round(fatura.genel_toplam * fatura.odeme_kuru, 2)
                except ValueError:
                    pass
            elif fatura.para_birimi == 'TL':
                fatura.odenen_tl = fatura.genel_toplam
        # Onaylanan faturalarda eşleşmeleri hafızaya kaydet
        if yeni_durum in ['onaylandi', 'odendi']:
            from app.fatura_ai import hafizaya_kaydet
            from app.models import TalepKalem
            for kalem in fatura.kalemler:
                if kalem.eslesme_durumu in ['eslesti', 'fiyat_farki'] and kalem.talep_kalem_id and kalem.malzeme_adi:
                    tk = TalepKalem.query.get(kalem.talep_kalem_id)
                    if tk:
                        hafizaya_kaydet(kalem.malzeme_adi, tk.malzeme_adi)
        db.session.commit()
        flash(f'Fatura durumu güncellendi: {yeni_durum}', 'success')
    return redirect(url_for('muhasebe.fatura_detay', fatura_id=fatura.id))

@muhasebe.route('/fatura/<int:fatura_id>/esles', methods=['POST'])
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_esles(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    talep_id = request.form.get('talep_id')
    if talep_id:
        talep = db.get_or_404(TalepFormu, int(talep_id))
        fatura.talep_id = talep.id
        from app.fatura_ai import siparis_eslestir
        eslesme = siparis_eslestir(
            [{
                'malzeme_adi': k.malzeme_adi,
                'miktar': k.miktar,
                'birim': k.birim,
                'liste_fiyati': k.liste_fiyati,
                'iskonto_orani': k.iskonto_orani,
                'iskonto_tutari': k.iskonto_tutari,
                'br_fiyat': k.br_fiyat,
                'kdv_orani': k.kdv_orani,
                'toplam_fiyat': k.toplam_fiyat,
            } for k in fatura.kalemler],
            talep.kalemler
        )
        from app.fatura_ai import hafizaya_kaydet
        for i, kalem in enumerate(fatura.kalemler):
            if i < len(eslesme):
                kalem.talep_kalem_id = eslesme[i]['talep_kalem_id']
                kalem.eslesme_durumu = eslesme[i]['eslesme_durumu']
                kalem.eslesme_notu = eslesme[i]['eslesme_notu']
                # Eşleşen kalemleri hafızaya kaydet
                if eslesme[i]['talep_kalem_id'] and kalem.malzeme_adi:
                    from app.models import TalepKalem
                    tk = TalepKalem.query.get(eslesme[i]['talep_kalem_id'])
                    if tk:
                        hafizaya_kaydet(kalem.malzeme_adi, tk.malzeme_adi)
        db.session.commit()
        eslesen = sum(1 for e in eslesme if e['eslesme_durumu'] in ['eslesti', 'fiyat_farki'])
        toplam = len(eslesme)
        if eslesen == 0:
            flash(f'{talep.siparis_no} ile eşleştirildi — ancak hiçbir kalem uyuşmadı. Doğru siparişi seçtiğinizden emin olun.', 'warning')
        elif eslesen == toplam:
            flash(f'{talep.siparis_no} ile eşleştirildi. Tüm kalemler uyuştu ✓', 'success')
        else:
            flash(f'{talep.siparis_no} ile eşleştirildi. {eslesen}/{toplam} kalem uyuştu.', 'warning')
    return redirect(url_for('muhasebe.fatura_detay', fatura_id=fatura.id))

@muhasebe.route('/fatura/<int:fatura_id>/sil', methods=['POST'])
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_sil(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    import os
    if fatura.dosya_yolu:
        tam_yol = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'faturalar', fatura.dosya_yolu)
        if os.path.exists(tam_yol):
            os.remove(tam_yol)
    db.session.delete(fatura)
    db.session.commit()
    flash('Fatura silindi.', 'warning')
    return redirect(url_for('muhasebe.fatura_listesi'))

@muhasebe.route('/fatura/<int:fatura_id>/pdf')
@login_required
@role_required('muhasebe', 'satinalma', 'admin')
def fatura_pdf_indir(fatura_id):
    fatura = db.get_or_404(Fatura, fatura_id)
    import os
    from flask import send_file
    tam_yol = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'faturalar', fatura.dosya_yolu)
    return send_file(tam_yol, as_attachment=False, mimetype='application/pdf')


# ---------------------------------------------------------------------------
# G-004: MALZEME LİSTESİ
# ---------------------------------------------------------------------------

@admin.route('/malzemeler')
@login_required
def malzeme_listesi():
    malzemeler = Malzeme.query.order_by(Malzeme.stok_kodu).all()
    return render_template('admin/malzeme_listesi.html', malzemeler=malzemeler)

@admin.route('/malzeme/ekle', methods=['POST'])
@login_required
def malzeme_ekle():
    if not has_permission(current_user, 'list.manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403
    adi = request.form.get('malzeme_adi', '').strip()
    if not adi:
        return jsonify({'ok': False, 'hata': 'Malzeme adı boş olamaz'}), 400
    m = Malzeme(
        stok_kodu=generate_stok_kodu(),
        malzeme_adi=adi,
        birim=request.form.get('birim', '').strip(),
        kategori=request.form.get('kategori', '').strip(),
        aciklama=request.form.get('aciklama', '').strip(),
    )
    db.session.add(m)
    db.session.commit()
    return jsonify({'ok': True, 'id': m.id, 'stok_kodu': m.stok_kodu})

@admin.route('/malzeme/<int:m_id>/duzenle', methods=['POST'])
@login_required
def malzeme_duzenle(m_id):
    if not has_permission(current_user, 'list.manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403
    m = db.get_or_404(Malzeme, m_id)
    alan = request.form.get('alan')
    deger = request.form.get('deger', '').strip()
    if alan in ('malzeme_adi', 'birim', 'kategori', 'aciklama', 'kullanim_notu'):
        setattr(m, alan, deger)
        db.session.commit()
    return jsonify({'ok': True})

@admin.route('/malzeme/<int:m_id>/sil', methods=['POST'])
@login_required
def malzeme_sil(m_id):
    if not has_permission(current_user, 'list.manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403
    m = db.get_or_404(Malzeme, m_id)
    m.is_active = False
    db.session.commit()
    return jsonify({'ok': True})


# ---------------------------------------------------------------------------
# G-004: ÜRÜN LİSTESİ
# ---------------------------------------------------------------------------

@admin.route('/urunler')
@login_required
def urun_listesi():
    urunler = Urun.query.order_by(Urun.urun_kodu).all()
    return render_template('admin/urun_listesi.html', urunler=urunler)

@admin.route('/urun/ekle', methods=['POST'])
@login_required
def urun_ekle():
    if not has_permission(current_user, 'list.manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403
    adi = request.form.get('urun_adi', '').strip()
    if not adi:
        return jsonify({'ok': False, 'hata': 'Ürün adı boş olamaz'}), 400
    u = Urun(
        urun_kodu=generate_urun_kodu(),
        urun_adi=adi,
        proje=request.form.get('proje', '').strip(),
        makine=request.form.get('makine', '').strip(),
        aciklama=request.form.get('aciklama', '').strip(),
    )
    db.session.add(u)
    db.session.commit()
    return jsonify({'ok': True, 'id': u.id, 'urun_kodu': u.urun_kodu})

@admin.route('/urun/<int:u_id>/duzenle', methods=['POST'])
@login_required
def urun_duzenle(u_id):
    if not has_permission(current_user, 'list.manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403
    u = db.get_or_404(Urun, u_id)
    alan = request.form.get('alan')
    deger = request.form.get('deger', '').strip()
    if alan in ('urun_adi', 'proje', 'makine', 'aciklama'):
        setattr(u, alan, deger)
        db.session.commit()
    return jsonify({'ok': True})

@admin.route('/urun/<int:u_id>/sil', methods=['POST'])
@login_required
def urun_sil(u_id):
    if not has_permission(current_user, 'list.manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403
    u = db.get_or_404(Urun, u_id)
    u.is_active = False
    db.session.commit()
    return jsonify({'ok': True})


# ---------------------------------------------------------------------------
# B-008: ADMIN SUNUCU DASHBOARD
# ---------------------------------------------------------------------------

@admin.route('/sunucu')
@login_required
@role_required('admin')
def sunucu_dashboard():
    import psutil, subprocess
    from datetime import timedelta

    # CPU
    cpu_yuzde = psutil.cpu_percent(interval=0.5)
    cpu_sayisi = psutil.cpu_count()

    # RAM
    ram = psutil.virtual_memory()
    ram_toplam_gb = round(ram.total / 1024**3, 1)
    ram_kullanan_gb = round(ram.used / 1024**3, 1)
    ram_yuzde = ram.percent

    # Disk
    disk = psutil.disk_usage('/')
    disk_toplam_gb = round(disk.total / 1024**3, 1)
    disk_kullanan_gb = round(disk.used / 1024**3, 1)
    disk_yuzde = disk.percent

    # Uptime
    boot_ts = psutil.boot_time()
    import time
    uptime_sn = int(time.time() - boot_ts)
    uptime_str = str(timedelta(seconds=uptime_sn))

    # Son 30 dk aktif kullanıcılar (son_giris bazlı)
    otuz_dk_once = datetime.utcnow() - timedelta(minutes=30)
    aktif_kullanicilar = User.query.filter(
        User.son_giris >= otuz_dk_once,
        User.is_active == True
    ).order_by(User.son_giris.desc()).all()

    # Erlau servis durumu
    try:
        result = subprocess.run(['systemctl', 'is-active', 'erlau'],
                                capture_output=True, text=True, timeout=3)
        servis_durum = result.stdout.strip()
    except Exception:
        servis_durum = 'bilinmiyor'

    return render_template('admin/sunucu_dashboard.html',
        cpu_yuzde=cpu_yuzde, cpu_sayisi=cpu_sayisi,
        ram_toplam_gb=ram_toplam_gb, ram_kullanan_gb=ram_kullanan_gb, ram_yuzde=ram_yuzde,
        disk_toplam_gb=disk_toplam_gb, disk_kullanan_gb=disk_kullanan_gb, disk_yuzde=disk_yuzde,
        uptime_str=uptime_str, aktif_kullanicilar=aktif_kullanicilar,
        servis_durum=servis_durum)


@admin.route('/sunucu/durum')
@login_required
@role_required('admin')
def sunucu_durum_api():
    """AJAX: Canlı CPU/RAM/disk verileri döner."""
    import psutil
    cpu_yuzde = psutil.cpu_percent(interval=0.3)
    ram = psutil.virtual_memory()
    disk = psutil.disk_usage('/')
    return jsonify({
        'cpu': cpu_yuzde,
        'ram': ram.percent,
        'ram_kullanan': round(ram.used / 1024**3, 1),
        'ram_toplam': round(ram.total / 1024**3, 1),
        'disk': disk.percent,
        'disk_kullanan': round(disk.used / 1024**3, 1),
        'disk_toplam': round(disk.total / 1024**3, 1),
    })


@admin.route('/sunucu/restart', methods=['POST'])
@login_required
@role_required('admin')
def sunucu_restart():
    """Şifre doğrulandıktan sonra erlau servisini yeniden başlatır."""
    import subprocess
    from werkzeug.security import check_password_hash
    sifre = request.form.get('sifre', '')
    if not check_password_hash(current_user.password, sifre):
        return jsonify({'ok': False, 'hata': 'Şifre hatalı'}), 403
    try:
        subprocess.Popen(['systemctl', 'restart', 'erlau'])
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'hata': str(e)}), 500


# ---------------------------------------------------------------------------
# G-004: AUTOCOMPLETE API (G-003 için)
# ---------------------------------------------------------------------------

api = Blueprint('api', __name__, url_prefix='/api')

@api.route('/malzeme-ara')
@login_required
def malzeme_ara():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    sonuclar = Malzeme.query.filter(
        db.or_(
            Malzeme.malzeme_adi.ilike(f'%{q}%'),
            Malzeme.stok_kodu.ilike(f'%{q}%'),
            Malzeme.aciklama.ilike(f'%{q}%')
        ),
        Malzeme.is_active == True
    ).order_by(
        db.case((Malzeme.stok_kodu == q, 0), else_=1),
        Malzeme.malzeme_adi
    ).limit(12).all()
    return jsonify([{
        'id': m.id,
        'stok_kodu': m.stok_kodu,
        'malzeme_adi': m.malzeme_adi,
        'birim': m.birim or '',
        'kategori': m.kategori or '',
        'kullanim_notu': m.kullanim_notu or ''
    } for m in sonuclar])

@api.route('/malzeme-oneri', methods=['POST'])
@login_required
def malzeme_oneri():
    """AI destekli malzeme önerisi — yeni talep formunda kullanılır."""
    data = request.get_json() or {}
    malzeme_adi = data.get('malzeme_adi', '').strip()
    proje_makine = data.get('proje_makine', '').strip()
    kullanim_amaci = data.get('kullanim_amaci', '').strip()

    if len(malzeme_adi) < 2:
        return jsonify({'oneri': None})

    # Önce veritabanında benzer malzemeleri bul
    benzerler = Malzeme.query.filter(
        db.or_(
            Malzeme.malzeme_adi.ilike(f'%{malzeme_adi}%'),
            Malzeme.aciklama.ilike(f'%{malzeme_adi}%')
        ),
        Malzeme.is_active == True
    ).limit(5).all()

    benzer_liste = [
        f"{m.stok_kodu} — {m.malzeme_adi}"
        + (f" ({m.kullanim_notu[:80]})" if m.kullanim_notu else "")
        for m in benzerler
    ]

    try:
        import anthropic, os
        client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'), timeout=10.0)
        prompt = f"""Satın alma sisteminde yeni talep oluşturuluyor.

Girilen malzeme adı: "{malzeme_adi}"
Proje/Makine: "{proje_makine or 'belirtilmemiş'}"
Kullanım amacı: "{kullanim_amaci or 'belirtilmemiş'}"

Sistemdeki benzer malzemeler:
{chr(10).join(benzer_liste) if benzer_liste else 'Benzer malzeme bulunamadı'}

Kısa ve pratik bir yanıt ver (max 2 cümle):
1. Bu malzeme için önerilen standart isim ve özellikler neler olmalı?
2. Sistemdeki benzer malzemelerden biri kullanılabilir mi?

Türkçe yanıtla, teknik ve özlü ol."""

        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=200,
            messages=[{'role': 'user', 'content': prompt}]
        )
        oneri_metni = msg.content[0].text.strip()
    except Exception:
        oneri_metni = None

    return jsonify({
        'oneri': oneri_metni,
        'benzerler': [{'stok_kodu': m.stok_kodu, 'malzeme_adi': m.malzeme_adi,
                       'birim': m.birim or '', 'kullanim_notu': m.kullanim_notu or ''}
                      for m in benzerler]
    })


@api.route('/urun-ara')
@login_required
def urun_ara():
    q = request.args.get('q', '').strip()
    if len(q) < 3:
        return jsonify([])
    sonuclar = Urun.query.filter(
        Urun.urun_adi.ilike(f'%{q}%'),
        Urun.is_active == True
    ).limit(10).all()
    return jsonify([{
        'id': u.id,
        'urun_kodu': u.urun_kodu,
        'urun_adi': u.urun_adi,
        'proje': u.proje or '',
        'makine': u.makine or ''
    } for u in sonuclar])

@api.route('/proje-makine-ara')
@login_required
def proje_makine_ara():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    sonuclar = db.session.query(TalepKalem.proje_makine)\
        .filter(TalepKalem.proje_makine.isnot(None))\
        .filter(TalepKalem.proje_makine != '')\
        .filter(TalepKalem.proje_makine.ilike(f'%{q}%'))\
        .distinct()\
        .order_by(TalepKalem.proje_makine)\
        .limit(12).all()
    return jsonify([r[0] for r in sonuclar])


@api.route('/urun-ekle', methods=['POST'])
@login_required
def urun_otomatik_ekle():
    adi = (request.json or {}).get('urun_adi', '').strip()
    if not adi:
        return jsonify({'ok': False}), 400
    mevcut = Urun.query.filter_by(urun_adi=adi).first()
    if mevcut:
        return jsonify({'ok': True, 'id': mevcut.id, 'urun_kodu': mevcut.urun_kodu})
    u = Urun(urun_kodu=generate_urun_kodu(), urun_adi=adi)
    db.session.add(u)
    db.session.commit()
    return jsonify({'ok': True, 'id': u.id, 'urun_kodu': u.urun_kodu})


# ---------------------------------------------------------------------------
# G-006: ÜRETİM MODÜLܽ
# ---------------------------------------------------------------------------

def uretim_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        endpoint_permission = ENDPOINT_PERMISSIONS.get(request.endpoint, 'production.dashboard')
        production_profiles = ('uretim_operatoru', 'cnc_operatoru', 'uretim_sorumlusu', 'uretim_yoneticisi')
        has_profile_access = getattr(current_user, 'job_profile', '') in production_profiles
        if not current_user.is_authenticated or (not has_profile_access and current_user.role not in ('uretim', 'departman_yoneticisi', 'admin', 'gm') and not has_permission(current_user, endpoint_permission)):
            flash('Üretim modülüne erişim yetkiniz yok.', 'danger')
            return redirect(url_for('auth.portal'))
        return f(*args, **kwargs)
    return decorated


def _production_manager_user():
    return getattr(current_user, 'job_profile', '') in ('admin', 'gm', 'uretim_sorumlusu', 'uretim_yoneticisi') or current_user.role in ('admin', 'gm')


def _station_scope_ids():
    if getattr(current_user, 'scope_type', '') != 'assigned_station':
        return None
    ids = [row.station_id for row in getattr(current_user, 'kapsam_istasyonlari', [])]
    if not ids and getattr(current_user, 'uretim_personeli', None) and current_user.uretim_personeli.istasyon_id:
        ids = [current_user.uretim_personeli.istasyon_id]
    return ids


def _apply_station_scope(query, column):
    ids = _station_scope_ids()
    if ids is None:
        return query
    if not ids:
        return query.filter(False)
    return query.filter(column.in_(ids))

@uretim.route('/')
@login_required
@uretim_required
def uretim_dashboard():
    from sqlalchemy import func
    bugun = date.today()
    hafta_baslangic = bugun - timedelta(days=bugun.weekday())
    hafta_bitis = hafta_baslangic + timedelta(days=4)
    aktif = UretimPlani.query.filter_by(durum='aktif').order_by(UretimPlani.id.desc()).first()
    istasyonlar = _apply_station_scope(
        IsIstasyonu.query.filter_by(is_active=True),
        IsIstasyonu.id
    ).order_by(IsIstasyonu.istasyon_adi).all()
    visible_station_ids = [i.id for i in istasyonlar]
    plan_q = UretimPlaniSatir.query.join(UretimPlani).filter(
        UretimPlaniSatir.tarih == bugun,
        UretimPlani.durum == 'aktif'
    )
    if visible_station_ids:
        plan_q = plan_q.filter(UretimPlaniSatir.istasyon_id.in_(visible_station_ids))
    elif _station_scope_ids() is not None:
        plan_q = plan_q.filter(False)
    plan_satirlari = plan_q.all()
    satir_ids = [s.id for s in plan_satirlari]
    kayit_map = {}
    if satir_ids:
        rows = db.session.query(
            UretimKaydi.plan_satir_id,
            func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
        ).filter(UretimKaydi.plan_satir_id.in_(satir_ids)).group_by(UretimKaydi.plan_satir_id).all()
        kayit_map = {satir_id: {'gerceklesen_adet': int(toplam or 0)} for satir_id, toplam in rows}

    toplam_planlanan = sum((s.planlanan_adet or 0) + (getattr(s, 'devir_adet', 0) or 0) for s in plan_satirlari)
    toplam_gerceklesen = sum(k['gerceklesen_adet'] for k in kayit_map.values())

    personel_map = {}
    for p in UretimPersoneli.query.filter_by(is_active=True).all():
        if p.istasyon_id and p.istasyon_id not in personel_map:
            personel_map[p.istasyon_id] = p.tam_ad

    istasyon_durumlar = []
    for istasyon in istasyonlar:
        istasyon_satirlari = [s for s in plan_satirlari if s.istasyon_id == istasyon.id]
        hedef = sum((s.planlanan_adet or 0) + (getattr(s, 'devir_adet', 0) or 0) for s in istasyon_satirlari)
        gerceklesen = sum(kayit_map.get(s.id, {}).get('gerceklesen_adet', 0) for s in istasyon_satirlari)
        yuzde = round(gerceklesen / hedef * 100, 0) if hedef else 0
        renk = '#16a34a' if yuzde >= 100 else '#f59e0b' if yuzde >= 70 else '#dc2626'
        istasyon_durumlar.append({
            'istasyon': istasyon.istasyon_adi,
            'personel_adi': personel_map.get(istasyon.id),
            'hedef': hedef,
            'gerceklesen': gerceklesen,
            'fark': gerceklesen - hedef,
            'yuzde': int(yuzde),
            'renk': renk,
        })

    haftalik_ilerleme = []
    gun_adlari = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    if aktif:
        hafta_baslangic = aktif.baslangic_tarihi or hafta_baslangic
    for idx, gun_adi in enumerate(gun_adlari):
        gun = hafta_baslangic + timedelta(days=idx)
        gun_satirlari = UretimPlaniSatir.query.join(UretimPlani).filter(
            UretimPlaniSatir.tarih == gun,
            UretimPlani.durum == 'aktif'
        ).all()
        gun_ids = [s.id for s in gun_satirlari]
        gun_gerceklesen = 0
        if gun_ids:
            gun_gerceklesen = db.session.query(func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)).filter(UretimKaydi.plan_satir_id.in_(gun_ids)).scalar() or 0
        gun_hedef = sum((s.planlanan_adet or 0) + (getattr(s, 'devir_adet', 0) or 0) for s in gun_satirlari)
        haftalik_ilerleme.append({
            'tarih': gun,
            'gun_adi': gun_adi,
            'toplam_hedef': int(gun_hedef or 0),
            'toplam_gerceklesen': int(gun_gerceklesen or 0),
            'yuzde': round(int(gun_gerceklesen or 0) / int(gun_hedef or 1) * 100, 0) if gun_hedef else 0,
        })

    personel_rows_q = db.session.query(
        UretimPersoneli.id,
        UretimPersoneli.ad,
        UretimPersoneli.soyad,
        IsIstasyonu.istasyon_adi,
        func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0),
        func.count(func.distinct(UretimKaydi.tarih))
    ).join(UretimKaydi, UretimKaydi.uretim_personeli_id == UretimPersoneli.id)\
     .outerjoin(IsIstasyonu, IsIstasyonu.id == UretimPersoneli.istasyon_id)\
     .filter(UretimKaydi.tarih >= hafta_baslangic, UretimKaydi.tarih <= hafta_bitis)
    personel_rows_q = _apply_station_scope(personel_rows_q, UretimKaydi.istasyon_id)
    personel_rows = personel_rows_q.group_by(UretimPersoneli.id, UretimPersoneli.ad, UretimPersoneli.soyad, IsIstasyonu.istasyon_adi)\
     .order_by(func.sum(UretimKaydi.gerceklesen_adet).desc()).limit(10).all()
    personel_siralamasi = []
    for row in personel_rows:
        toplam_hafta = int(row[4] or 0)
        gun_sayisi = int(row[5] or 0)
        personel_siralamasi.append({
            'ad': f"{row[1]} {row[2] or ''}".strip(),
            'istasyon': row[3] or '—',
            'toplam_hafta': toplam_hafta,
            'gun_ort': round(toplam_hafta / gun_sayisi, 1) if gun_sayisi else 0,
        })

    ariza_q = ArizaKaydi.query.filter(ArizaKaydi.tarih >= hafta_baslangic, ArizaKaydi.tarih <= hafta_bitis)
    ariza_q = _apply_station_scope(ariza_q, ArizaKaydi.istasyon_id)
    arizalar = ariza_q.all()
    toplam_dakika = 0
    for ariza in arizalar:
        if ariza.baslangic_saati and ariza.bitis_saati:
            bas = datetime.combine(ariza.tarih, ariza.baslangic_saati)
            bit = datetime.combine(ariza.tarih, ariza.bitis_saati)
            if bit > bas:
                toplam_dakika += int((bit - bas).total_seconds() // 60)
    en_cok = db.session.query(IsIstasyonu.istasyon_adi, func.count(ArizaKaydi.id)).join(ArizaKaydi, ArizaKaydi.istasyon_id == IsIstasyonu.id)\
        .filter(ArizaKaydi.tarih >= hafta_baslangic, ArizaKaydi.tarih <= hafta_bitis)\
        .group_by(IsIstasyonu.istasyon_adi).order_by(func.count(ArizaKaydi.id).desc()).first()
    ariza_ozet = {
        'bu_hafta_adet': len(arizalar),
        'bu_hafta_sure_saat': round(toplam_dakika / 60, 1),
        'en_cok_ariza_istasyon': en_cok[0] if en_cok else '—',
    }

    acik_dof_sayisi = 0
    geciken_aksiyon_sayisi = 0
    try:
        acik_dof_sayisi = DOF.query.filter(DOF.hedef_departman == 'Üretim', DOF.durum.in_(('acik', 'isleniyor', 'gecikti'))).count()
        geciken_aksiyon_sayisi = db.session.query(DOFAksiyon).join(DOF).filter(
            DOF.hedef_departman == 'Üretim',
            DOFAksiyon.durum != 'tamamlandi',
            DOFAksiyon.planlanan_tarih < bugun
        ).count()
    except Exception:
        acik_dof_sayisi = 0
        geciken_aksiyon_sayisi = 0

    return render_template('uretim/dashboard.html',
        istasyonlar=istasyonlar, istasyon_durumlar=istasyon_durumlar,
        plan_satirlari=plan_satirlari,
        kayit_map=kayit_map, bugun=bugun,
        haftalik_ilerleme=haftalik_ilerleme,
        personel_siralamasi=personel_siralamasi,
        ariza_ozet=ariza_ozet,
        toplam_hedef=toplam_planlanan,
        toplam_gerceklesen=toplam_gerceklesen,
        toplam_planlanan=toplam_planlanan,
        acik_dof_sayisi=acik_dof_sayisi,
        geciken_aksiyon_sayisi=geciken_aksiyon_sayisi)

@uretim.route('/giris', methods=['GET', 'POST'])
@login_required
@uretim_required
def uretim_giris():
    if request.method == 'POST':
        station_scope_ids = _station_scope_ids()
        mode = request.form.get('mode', 'tek')
        tarih_str = request.form.get('tarih', str(date.today()))
        try:
            kayit_tarihi = date.fromisoformat(tarih_str)
        except ValueError:
            kayit_tarihi = date.today()

        if mode == 'batch':
            satir_ids = request.form.getlist('plan_satir_id[]')
            gerceklesenler = request.form.getlist('gerceklesen[]')
            personel_ids = request.form.getlist('uretim_personeli_id[]')
            fire_adetler = request.form.getlist('fire_adet[]')
            hurda_adetler = request.form.getlist('hurda_adet[]')
            eklenen = 0
            for i, satir_id in enumerate(satir_ids):
                gercek = int(gerceklesenler[i] or 0) if i < len(gerceklesenler) else 0
                if gercek <= 0:
                    continue
                personel_id = personel_ids[i] if i < len(personel_ids) else None
                fire = int(fire_adetler[i] or 0) if i < len(fire_adetler) else 0
                hurda = int(hurda_adetler[i] or 0) if i < len(hurda_adetler) else 0
                satir = UretimPlaniSatir.query.get(int(satir_id))
                if not satir:
                    continue
                if station_scope_ids is not None and satir.istasyon_id not in station_scope_ids:
                    continue
                mevcut = UretimKaydi.query.filter_by(plan_satir_id=satir.id, tarih=kayit_tarihi).first()
                if mevcut:
                    mevcut.gerceklesen_adet = gercek
                    mevcut.fire_adet = fire
                    mevcut.hurda_adet = hurda
                    mevcut.uretim_personeli_id = personel_id or None
                    mevcut.giren_personel_id = current_user.id
                else:
                    db.session.add(UretimKaydi(
                        plan_satir_id=satir.id,
                        istasyon_id=satir.istasyon_id,
                        urun_id=satir.urun_id,
                        tarih=kayit_tarihi,
                        gerceklesen_adet=gercek,
                        fire_adet=fire,
                        hurda_adet=hurda,
                        uretim_personeli_id=personel_id or None,
                        giren_personel_id=current_user.id,
                    ))
                eklenen += 1
            db.session.commit()
            flash(f'{eklenen} üretim kaydı kaydedildi.', 'success')
        else:
            istasyon_id = request.form.get('istasyon_id', type=int)
            if station_scope_ids is not None and istasyon_id not in station_scope_ids:
                flash('Bu istasyona üretim kaydı girme yetkiniz yok.', 'danger')
                return redirect(url_for('uretim.uretim_giris'))
            db.session.add(UretimKaydi(
                istasyon_id=istasyon_id,
                urun_id=request.form.get('urun_id') or None,
                tarih=kayit_tarihi,
                gerceklesen_adet=int(request.form.get('gerceklesen_adet_tek', 0) or 0),
                fire_adet=int(request.form.get('fire_adet_tek', 0) or 0),
                hurda_adet=int(request.form.get('hurda_adet_tek', 0) or 0),
                uretim_personeli_id=request.form.get('uretim_personeli_id_tek') or None,
                giren_personel_id=current_user.id,
            ))
            db.session.commit()
            flash('Üretim kaydı eklendi.', 'success')
        return redirect(url_for('uretim.uretim_dashboard'))

    tarih_str = request.args.get('tarih', str(date.today()))
    try:
        tarih = date.fromisoformat(tarih_str)
    except ValueError:
        tarih = date.today()
    istasyonlar = _apply_station_scope(IsIstasyonu.query.filter_by(is_active=True), IsIstasyonu.id).all()
    visible_station_ids = [i.id for i in istasyonlar]
    urunler = Urun.query.filter_by(is_active=True).order_by(Urun.urun_adi).all()
    personel_q = UretimPersoneli.query.filter_by(is_active=True)
    if visible_station_ids:
        personel_q = personel_q.filter((UretimPersoneli.istasyon_id.in_(visible_station_ids)) | (UretimPersoneli.istasyon_id == None))
    elif _station_scope_ids() is not None:
        personel_q = personel_q.filter(False)
    personeller = personel_q.order_by(UretimPersoneli.ad, UretimPersoneli.soyad).all()
    plan_q = UretimPlaniSatir.query.join(UretimPlani).filter(
        UretimPlaniSatir.tarih == tarih,
        UretimPlani.durum == 'aktif'
    )
    if visible_station_ids:
        plan_q = plan_q.filter(UretimPlaniSatir.istasyon_id.in_(visible_station_ids))
    elif _station_scope_ids() is not None:
        plan_q = plan_q.filter(False)
    plan_satirlari = plan_q.all()
    return render_template('uretim/giris.html',
        istasyonlar=istasyonlar, urunler=urunler,
        personeller=personeller,
        plan_satirlari=plan_satirlari, bugun=tarih, tarih=tarih_str)

@uretim.route('/ariza', methods=['GET', 'POST'])
@login_required
@uretim_required
def ariza_kaydi():
    if request.method == 'POST':
        station_scope_ids = _station_scope_ids()
        istasyon_id = request.form.get('istasyon_id', type=int)
        if station_scope_ids is not None and istasyon_id not in station_scope_ids:
            flash('Bu istasyona arıza kaydı girme yetkiniz yok.', 'danger')
            return redirect(url_for('uretim.ariza_kaydi'))
        from datetime import time as dtime
        bas = request.form.get('baslangic_saati')
        bit = request.form.get('bitis_saati')
        a = ArizaKaydi(
            istasyon_id=istasyon_id,
            tarih=date.fromisoformat(request.form.get('tarih', str(date.today()))),
            baslangic_saati=dtime.fromisoformat(bas) if bas else None,
            bitis_saati=dtime.fromisoformat(bit) if bit else None,
            aciklama=request.form.get('aciklama', ''),
            giren_personel_id=current_user.id,
        )
        db.session.add(a)
        db.session.commit()
        flash('Arıza kaydı oluşturuldu.', 'success')
        return redirect(url_for('uretim.uretim_dashboard'))
    istasyonlar = _apply_station_scope(IsIstasyonu.query.filter_by(is_active=True), IsIstasyonu.id).all()
    return render_template('uretim/ariza.html', istasyonlar=istasyonlar, bugun=date.today())

@uretim.route('/personel')
@login_required
@uretim_required
def personel_listesi():
    if getattr(current_user, 'job_profile', '') in ('uretim_operatoru', 'cnc_operatoru'):
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.uretim_dashboard'))
    if not (_production_manager_user() or has_permission(current_user, 'production.staff_view')):
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.uretim_dashboard'))
    from sqlalchemy import func
    bugun = date.today()
    hafta_baslangic = bugun - timedelta(days=bugun.weekday())
    ay_baslangic = date(bugun.year, bugun.month, 1)

    hafta_rows = db.session.query(
        UretimKaydi.uretim_personeli_id,
        func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
    ).filter(
        UretimKaydi.tarih >= hafta_baslangic,
        UretimKaydi.uretim_personeli_id.isnot(None)
    ).group_by(UretimKaydi.uretim_personeli_id).all()

    ay_rows = db.session.query(
        UretimKaydi.uretim_personeli_id,
        func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
    ).filter(
        UretimKaydi.tarih >= ay_baslangic,
        UretimKaydi.uretim_personeli_id.isnot(None)
    ).group_by(UretimKaydi.uretim_personeli_id).all()

    hafta_map = {personel_id: int(toplam or 0) for personel_id, toplam in hafta_rows}
    ay_map = {personel_id: int(toplam or 0) for personel_id, toplam in ay_rows}
    personeller = UretimPersoneli.query.order_by(UretimPersoneli.is_active.desc(), UretimPersoneli.ad, UretimPersoneli.soyad).all()
    for personel in personeller:
        personel.bu_hafta = hafta_map.get(personel.id, 0)
        personel.bu_ay = ay_map.get(personel.id, 0)
    istasyonlar = IsIstasyonu.query.filter_by(is_active=True).order_by(IsIstasyonu.istasyon_adi).all()
    return render_template('uretim/personel.html', personeller=personeller, istasyonlar=istasyonlar)

@uretim.route('/personel/ekle', methods=['POST'])
@login_required
@uretim_required
def personel_ekle():
    if not (_production_manager_user() or has_permission(current_user, 'production.staff_manage')):
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.personel_listesi'))
    ad = request.form.get('ad', '').strip()
    soyad = request.form.get('soyad', '').strip()
    sicil_no = request.form.get('sicil_no', '').strip() or None
    istasyon_id = request.form.get('istasyon_id') or None
    if not ad:
        flash('Ad alanı zorunludur.', 'danger')
        return redirect(url_for('uretim.personel_listesi'))
    if sicil_no and UretimPersoneli.query.filter_by(sicil_no=sicil_no).first():
        flash('Bu sicil numarası zaten kayıtlı.', 'danger')
        return redirect(url_for('uretim.personel_listesi'))
    db.session.add(UretimPersoneli(ad=ad, soyad=soyad, sicil_no=sicil_no, istasyon_id=istasyon_id))
    db.session.commit()
    flash('Personel eklendi.', 'success')
    return redirect(url_for('uretim.personel_listesi'))

@uretim.route('/personel/<int:id>/duzenle', methods=['POST'])
@login_required
@uretim_required
def personel_duzenle(id):
    if not (_production_manager_user() or has_permission(current_user, 'production.staff_manage')):
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.personel_listesi'))
    personel = db.get_or_404(UretimPersoneli, id)
    ad = request.form.get('ad', '').strip()
    sicil_no = request.form.get('sicil_no', '').strip() or None
    if not ad:
        flash('Ad alanı zorunludur.', 'danger')
        return redirect(url_for('uretim.personel_listesi'))
    if sicil_no:
        mevcut = UretimPersoneli.query.filter(UretimPersoneli.sicil_no == sicil_no, UretimPersoneli.id != id).first()
        if mevcut:
            flash('Bu sicil numarası başka bir personele ait.', 'danger')
            return redirect(url_for('uretim.personel_listesi'))
    personel.ad = ad
    personel.soyad = request.form.get('soyad', '').strip()
    personel.sicil_no = sicil_no
    personel.istasyon_id = request.form.get('istasyon_id') or None
    db.session.commit()
    flash('Personel güncellendi.', 'success')
    return redirect(url_for('uretim.personel_listesi'))

@uretim.route('/personel/<int:id>/pasif', methods=['POST'])
@login_required
@uretim_required
def personel_pasif(id):
    if not (_production_manager_user() or has_permission(current_user, 'production.staff_manage')):
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.personel_listesi'))
    personel = db.get_or_404(UretimPersoneli, id)
    personel.is_active = False
    db.session.commit()
    flash('Personel pasife alındı.', 'success')
    return redirect(url_for('uretim.personel_listesi'))

@uretim.route('/personel/<int:id>')
@login_required
@uretim_required
def personel_detay(id):
    from sqlalchemy import func
    personel = db.get_or_404(UretimPersoneli, id)
    bugun = date.today()
    hafta_baslangic = bugun - timedelta(days=bugun.weekday())
    ay_baslangic = date(bugun.year, bugun.month, 1)
    otuz_gun_once = bugun - timedelta(days=29)

    bu_hafta_toplam = db.session.query(func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)).filter(
        UretimKaydi.uretim_personeli_id == personel.id,
        UretimKaydi.tarih >= hafta_baslangic
    ).scalar() or 0
    bu_ay_toplam = db.session.query(func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)).filter(
        UretimKaydi.uretim_personeli_id == personel.id,
        UretimKaydi.tarih >= ay_baslangic
    ).scalar() or 0
    calisilan_gun = db.session.query(func.count(func.distinct(UretimKaydi.tarih))).filter(
        UretimKaydi.uretim_personeli_id == personel.id,
        UretimKaydi.tarih >= ay_baslangic
    ).scalar() or 0
    gunluk_ortalama = float(bu_ay_toplam) / calisilan_gun if calisilan_gun else 0

    gun_rows = db.session.query(
        UretimKaydi.tarih,
        func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
    ).filter(
        UretimKaydi.uretim_personeli_id == personel.id,
        UretimKaydi.tarih >= otuz_gun_once
    ).group_by(UretimKaydi.tarih).order_by(UretimKaydi.tarih).all()
    gun_map = {tarih: int(toplam or 0) for tarih, toplam in gun_rows}
    performans_30gun = [
        {'tarih': (otuz_gun_once + timedelta(days=i)).isoformat(), 'toplam_adet': gun_map.get(otuz_gun_once + timedelta(days=i), 0)}
        for i in range(30)
    ]

    urun_rows = db.session.query(
        Urun.urun_adi,
        func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
    ).join(UretimKaydi, UretimKaydi.urun_id == Urun.id).filter(
        UretimKaydi.uretim_personeli_id == personel.id,
        UretimKaydi.tarih >= ay_baslangic
    ).group_by(Urun.urun_adi).order_by(func.sum(UretimKaydi.gerceklesen_adet).desc()).all()
    urun_dagilimi = [
        {'urun_adi': urun_adi, 'toplam': int(toplam or 0), 'oran': round((float(toplam or 0) / float(bu_ay_toplam) * 100), 1) if bu_ay_toplam else 0}
        for urun_adi, toplam in urun_rows
    ]
    son_kayitlar = UretimKaydi.query.filter_by(uretim_personeli_id=personel.id).order_by(UretimKaydi.tarih.desc(), UretimKaydi.id.desc()).limit(10).all()
    return render_template('uretim/personel_detay.html',
        personel=personel,
        bu_hafta_toplam=int(bu_hafta_toplam or 0),
        bu_ay_toplam=int(bu_ay_toplam or 0),
        gunluk_ortalama=gunluk_ortalama,
        performans_30gun=performans_30gun,
        urun_dagilimi=urun_dagilimi,
        son_kayitlar=son_kayitlar)

@uretim.route('/istasyonlar', methods=['GET', 'POST'])
@login_required
def istasyon_yonetim():
    if not (_production_manager_user() or has_permission(current_user, 'production.station_manage')):
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.uretim_dashboard'))
    if request.method == 'POST':
        eylem = request.form.get('eylem')
        if eylem == 'ekle':
            from app.utils import generate_istasyon_kodu
            i = IsIstasyonu(
                istasyon_kodu=generate_istasyon_kodu(),
                istasyon_adi=request.form.get('istasyon_adi', '').strip(),
                aciklama=request.form.get('aciklama', '').strip(),
            )
            db.session.add(i); db.session.commit()
            flash('İstasyon eklendi.', 'success')
        elif eylem == 'sil':
            i = db.get_or_404(IsIstasyonu, request.form.get('istasyon_id'))
            i.is_active = False; db.session.commit()
            flash('İstasyon pasif yapıldı.', 'warning')
        return redirect(url_for('uretim.istasyon_yonetim'))
    istasyonlar = IsIstasyonu.query.order_by(IsIstasyonu.istasyon_kodu).all()
    return render_template('uretim/istasyonlar.html', istasyonlar=istasyonlar)

@uretim.route('/raporlar')
@login_required
def uretim_raporlar():
    if getattr(current_user, 'job_profile', '') in ('uretim_operatoru', 'cnc_operatoru'):
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('uretim.uretim_dashboard'))
    if not (_production_manager_user() or has_permission(current_user, 'production.reports')):
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    baslangic = request.args.get('baslangic', str(date.today() - timedelta(days=6)))
    bitis = request.args.get('bitis', str(date.today()))
    istasyon_id = request.args.get('istasyon_id', '')
    try:
        bas_dt = date.fromisoformat(baslangic)
        bit_dt = date.fromisoformat(bitis)
    except ValueError:
        bas_dt = date.today() - timedelta(days=6)
        bit_dt = date.today()

    kayit_q = UretimKaydi.query.filter(
        UretimKaydi.tarih >= bas_dt,
        UretimKaydi.tarih <= bit_dt
    )
    plan_q = UretimPlaniSatir.query.filter(
        UretimPlaniSatir.tarih >= bas_dt,
        UretimPlaniSatir.tarih <= bit_dt
    )
    if istasyon_id:
        kayit_q = kayit_q.filter(UretimKaydi.istasyon_id == istasyon_id)
        plan_q = plan_q.filter(UretimPlaniSatir.istasyon_id == istasyon_id)
    kayit_q = _apply_station_scope(kayit_q, UretimKaydi.istasyon_id)
    plan_q = _apply_station_scope(plan_q, UretimPlaniSatir.istasyon_id)

    kayitlar = kayit_q.order_by(UretimKaydi.tarih).all()
    plan_satirlari = plan_q.all()
    istasyonlar = IsIstasyonu.query.filter_by(is_active=True).all()

    toplam_planlanan = sum(s.planlanan_adet for s in plan_satirlari)
    toplam_gerceklesen = sum(k.gerceklesen_adet for k in kayitlar)
    toplam_fire = sum(k.fire_adet for k in kayitlar)

    return render_template('uretim/raporlar.html',
        kayitlar=kayitlar, plan_satirlari=plan_satirlari,
        istasyonlar=istasyonlar, istasyon_id=istasyon_id,
        baslangic=baslangic, bitis=bitis,
        toplam_planlanan=toplam_planlanan,
        toplam_gerceklesen=toplam_gerceklesen,
        toplam_fire=toplam_fire)


# ---------------------------------------------------------------------------
# G-008: PLANLAMA MODÜLÜ
# ---------------------------------------------------------------------------

def _planlama_ozet_satirlari(plan):
    if not plan:
        return []
    ozet = {}
    gun_anahtarlari = ['pzt', 'sal', 'car', 'per', 'cum']
    hafta_baslangic = plan.baslangic_tarihi
    satir_ids = [satir.id for satir in plan.satirlar]
    gerceklesen_map = {}
    if satir_ids:
        from sqlalchemy import func
        rows = db.session.query(
            UretimKaydi.plan_satir_id,
            func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
        ).filter(UretimKaydi.plan_satir_id.in_(satir_ids)).group_by(UretimKaydi.plan_satir_id).all()
        gerceklesen_map = {satir_id: int(toplam or 0) for satir_id, toplam in rows}
    for satir in plan.satirlar:
        key = (satir.urun_id, satir.istasyon_id)
        row = ozet.setdefault(key, {
            'urun': satir.urun,
            'istasyon': satir.istasyon,
            'pzt': 0,
            'sal': 0,
            'car': 0,
            'per': 0,
            'cum': 0,
            'toplam': 0,
            'gerceklesen': 0,
            'kalan': 0,
            'yuzde': 0,
        })
        if hafta_baslangic and satir.tarih:
            gun_index = (satir.tarih - hafta_baslangic).days
            if 0 <= gun_index < len(gun_anahtarlari):
                row[gun_anahtarlari[gun_index]] += satir.planlanan_adet or 0
        row['toplam'] += satir.planlanan_adet or 0
        row['gerceklesen'] += gerceklesen_map.get(satir.id, 0)
    for row in ozet.values():
        row['kalan'] = max((row['toplam'] or 0) - (row['gerceklesen'] or 0), 0)
        row['yuzde'] = round((row['gerceklesen'] or 0) / (row['toplam'] or 1) * 100, 1) if row['toplam'] else 0
    return list(ozet.values())

def _planlama_dashboard_verileri(aktif):
    from sqlalchemy import func
    bugun = date.today()
    haftalik_veri = []
    istasyon_kapasite = []
    kalite_ozet = {'ok_adet': 0, 'nok_adet': 0, 'ok_oran': 0, 'hurda_adet': 0, 'tamir_adet': 0}
    toplam_hedef = 0
    toplam_gerceklesen = 0
    toplam_devir = 0

    if aktif:
        satirlar = list(aktif.satirlar)
        satir_ids = [satir.id for satir in satirlar]
        gerceklesen_map = {}
        if satir_ids:
            rows = db.session.query(
                UretimKaydi.plan_satir_id,
                func.coalesce(func.sum(UretimKaydi.gerceklesen_adet), 0)
            ).filter(UretimKaydi.plan_satir_id.in_(satir_ids)).group_by(UretimKaydi.plan_satir_id).all()
            gerceklesen_map = {satir_id: int(toplam or 0) for satir_id, toplam in rows}

        toplam_hedef = sum(s.planlanan_adet or 0 for s in satirlar)
        toplam_devir = sum(getattr(s, 'devir_adet', 0) or 0 for s in satirlar)
        toplam_gerceklesen = sum(gerceklesen_map.values())

        gun_adlari = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
        hafta_baslangic = aktif.baslangic_tarihi or (bugun - timedelta(days=bugun.weekday()))
        for idx, gun_adi in enumerate(gun_adlari):
            gun = hafta_baslangic + timedelta(days=idx)
            gun_satirlari = [s for s in satirlar if s.tarih == gun]
            haftalik_veri.append({
                'gun_adi': gun_adi,
                'planlanan': sum(s.planlanan_adet or 0 for s in gun_satirlari),
                'gerceklesen': sum(gerceklesen_map.get(s.id, 0) for s in gun_satirlari),
                'devir': sum(getattr(s, 'devir_adet', 0) or 0 for s in gun_satirlari),
            })

        istasyon_map = {}
        for satir in satirlar:
            key = satir.istasyon_id or 0
            row = istasyon_map.setdefault(key, {
                'istasyon_adi': satir.istasyon.istasyon_adi if satir.istasyon else '—',
                'planlanan': 0,
                'gerceklesen': 0,
                'oran': 0,
            })
            row['planlanan'] += satir.planlanan_adet or 0
            row['gerceklesen'] += gerceklesen_map.get(satir.id, 0)
        for row in istasyon_map.values():
            row['oran'] = round(row['gerceklesen'] / row['planlanan'] * 100, 1) if row['planlanan'] else 0
            istasyon_kapasite.append(row)
        istasyon_kapasite.sort(key=lambda r: r['istasyon_adi'])

        try:
            bas = aktif.baslangic_tarihi
            bit = aktif.bitis_tarihi
            if bas and bit:
                ok_adet = db.session.query(func.coalesce(func.sum(KaliteKontrol.ok_adet), 0)).filter(KaliteKontrol.tarih >= bas, KaliteKontrol.tarih <= bit).scalar() or 0
                nok_adet = db.session.query(func.coalesce(func.sum(KaliteKontrol.nok_adet), 0)).filter(KaliteKontrol.tarih >= bas, KaliteKontrol.tarih <= bit).scalar() or 0
                hurda_adet = db.session.query(func.coalesce(func.sum(KaliteKontrol.nok_adet), 0)).filter(KaliteKontrol.tarih >= bas, KaliteKontrol.tarih <= bit, KaliteKontrol.nok_akibet == 'hurda').scalar() or 0
                tamir_adet = db.session.query(func.coalesce(func.sum(KaliteKontrol.nok_adet), 0)).filter(KaliteKontrol.tarih >= bas, KaliteKontrol.tarih <= bit, KaliteKontrol.nok_akibet == 'tamir').scalar() or 0
                toplam_kontrol = int(ok_adet or 0) + int(nok_adet or 0)
                kalite_ozet = {
                    'ok_adet': int(ok_adet or 0),
                    'nok_adet': int(nok_adet or 0),
                    'ok_oran': round(int(ok_adet or 0) / toplam_kontrol * 100, 1) if toplam_kontrol else 0,
                    'hurda_adet': int(hurda_adet or 0),
                    'tamir_adet': int(tamir_adet or 0),
                }
        except Exception:
            kalite_ozet = {'ok_adet': 0, 'nok_adet': 0, 'ok_oran': 0, 'hurda_adet': 0, 'tamir_adet': 0}

    acik_dof_sayisi = 0
    geciken_aksiyon_sayisi = 0
    try:
        planlama_dept = ('Planlama', 'Planlama ve Tedarik Zinciri')
        acik_dof_sayisi = DOF.query.filter(
            DOF.hedef_departman.in_(planlama_dept),
            DOF.durum.in_(('acik', 'isleniyor', 'gecikti'))
        ).count()
        geciken_aksiyon_sayisi = db.session.query(DOFAksiyon).join(DOF).filter(
            DOF.hedef_departman.in_(planlama_dept),
            DOFAksiyon.durum != 'tamamlandi',
            DOFAksiyon.planlanan_tarih < bugun
        ).count()
    except Exception:
        acik_dof_sayisi = 0
        geciken_aksiyon_sayisi = 0

    tamamlanma_yuzdesi = round(toplam_gerceklesen / toplam_hedef * 100, 1) if toplam_hedef else 0
    return {
        'haftalik_veri': haftalik_veri,
        'istasyon_kapasite': istasyon_kapasite,
        'kalite_ozet': kalite_ozet,
        'toplam_hedef': toplam_hedef,
        'toplam_gerceklesen': toplam_gerceklesen,
        'tamamlanma_yuzdesi': tamamlanma_yuzdesi,
        'toplam_devir': toplam_devir,
        'bugun': bugun,
        'acik_dof_sayisi': acik_dof_sayisi,
        'geciken_aksiyon_sayisi': geciken_aksiyon_sayisi,
    }

@planlama.route('/')
@login_required
def planlama_dashboard():
    if current_user.role not in ('planlama', 'departman_yoneticisi', 'admin') and not has_permission(current_user, 'planning.dashboard'):
        flash('Erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    aktif = UretimPlani.query.filter_by(durum='aktif').order_by(UretimPlani.id.desc()).first()
    planlar = UretimPlani.query.order_by(UretimPlani.id.desc()).limit(5).all()
    dashboard_verileri = _planlama_dashboard_verileri(aktif)
    return render_template('planlama/dashboard.html',
        aktif=aktif, planlar=planlar, aktif_ozet=_planlama_ozet_satirlari(aktif),
        **dashboard_verileri)

@planlama.route('/gun-sonu-devir', methods=['POST'])
@login_required
def gun_sonu_devir():
    if current_user.role not in ('planlama', 'departman_yoneticisi', 'admin') and not has_permission(current_user, 'planning.carryover'):
        flash('Erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    bugun = date.today()
    aktif = UretimPlani.query.filter_by(durum='aktif').order_by(UretimPlani.id.desc()).first()
    if not aktif:
        flash('Aktif plan bulunmuyor.', 'danger')
        return redirect(url_for('planlama.planlama_dashboard'))

    bugun_satirlari = UretimPlaniSatir.query.filter_by(plan_id=aktif.id, tarih=bugun).all()
    olusturulan = 0
    for satir in bugun_satirlari:
        gerceklesen = db.session.query(db.func.coalesce(db.func.sum(UretimKaydi.gerceklesen_adet), 0)).filter_by(plan_satir_id=satir.id).scalar() or 0
        hedef = (satir.planlanan_adet or 0) + (getattr(satir, 'devir_adet', 0) or 0)
        kalan = max(hedef - int(gerceklesen or 0), 0)
        if kalan <= 0:
            continue
        sonraki_gun = devir_gunu(bugun)
        mevcut_devir = UretimPlaniSatir.query.filter_by(
            plan_id=aktif.id,
            urun_id=satir.urun_id,
            istasyon_id=satir.istasyon_id,
            tarih=sonraki_gun,
            kaynak='devir'
        ).first()
        if mevcut_devir:
            mevcut_devir.planlanan_adet = (mevcut_devir.planlanan_adet or 0) + kalan
            mevcut_devir.devir_adet = (mevcut_devir.devir_adet or 0) + kalan
        else:
            db.session.add(UretimPlaniSatir(
                plan_id=aktif.id,
                urun_id=satir.urun_id,
                istasyon_id=satir.istasyon_id,
                tarih=sonraki_gun,
                planlanan_adet=kalan,
                devir_adet=kalan,
                kaynak='devir',
            ))
        olusturulan += kalan
    db.session.commit()
    flash(f'Gün sonu devir işlemi tamamlandı. Devreden adet: {olusturulan}', 'success')
    return redirect(url_for('planlama.planlama_dashboard'))

@planlama.route('/yeni', methods=['GET', 'POST'])
@login_required
def yeni_plan():
    if current_user.role not in ('planlama', 'departman_yoneticisi', 'admin') and not has_permission(current_user, 'planning.create'):
        flash('Erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    if request.method == 'POST':
        eylem = request.form.get('eylem', 'taslak')
        baslangic = date.fromisoformat(request.form.get('baslangic_tarihi'))
        plan = UretimPlani(
            plan_no=generate_plan_no(),
            hafta=int(request.form.get('hafta')),
            yil=int(request.form.get('yil')),
            baslangic_tarihi=baslangic,
            bitis_tarihi=date.fromisoformat(request.form.get('bitis_tarihi')),
            planlayan_id=current_user.id,
            durum='aktif' if eylem == 'aktif' else 'taslak',
        )
        db.session.add(plan)
        db.session.flush()

        satirlar = request.form.getlist('urun_id[]')
        istasyonlar_list = request.form.getlist('istasyon_id[]')
        gunler = ['pzt', 'sal', 'car', 'per', 'cum']
        for idx, (uid, iid) in enumerate(zip(satirlar, istasyonlar_list)):
            if not uid:
                continue
            for g_idx, gun in enumerate(gunler):
                adetler = request.form.getlist(f'adet_{gun}[]')
                adet = int(adetler[idx] or 0) if idx < len(adetler) else 0
                if adet > 0:
                    satir = UretimPlaniSatir(
                        plan_id=plan.id,
                        urun_id=int(uid),
                        istasyon_id=int(iid) if iid else None,
                        tarih=baslangic + timedelta(days=g_idx),
                        planlanan_adet=adet,
                    )
                    db.session.add(satir)
        db.session.commit()
        flash(f'Plan {plan.plan_no} {"aktive edildi" if plan.durum == "aktif" else "taslak olarak kaydedildi"}.', 'success')
        return redirect(url_for('planlama.planlama_dashboard'))

    istasyonlar = IsIstasyonu.query.filter_by(is_active=True).all()
    urunler = Urun.query.filter_by(is_active=True).order_by(Urun.urun_adi).all()
    bugun = date.today()
    pzt = bugun - timedelta(days=bugun.weekday())
    return render_template('planlama/yeni_plan.html',
        istasyonlar=istasyonlar, urunler=urunler, pzt=pzt, cum=pzt + timedelta(days=4))

@planlama.route('/planlar')
@login_required
def plan_listesi():
    if current_user.role not in ('planlama', 'departman_yoneticisi', 'admin', 'gm') and not has_permission(current_user, 'planning.list'):
        flash('Erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    planlar = UretimPlani.query.order_by(UretimPlani.id.desc()).all()
    return render_template('planlama/plan_listesi.html', planlar=planlar)


# ---------------------------------------------------------------------------
# G-007: BAKIM MODÜLܽ
# ---------------------------------------------------------------------------

def bakim_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        endpoint_permission = ENDPOINT_PERMISSIONS.get(request.endpoint, 'maintenance.dashboard')
        if not current_user.is_authenticated or not has_permission(current_user, endpoint_permission):
            flash('Bakım modülüne erişim yetkiniz yok.', 'danger')
            return redirect(url_for('auth.portal'))
        return f(*args, **kwargs)
    return decorated

@bakim.route('/')
@login_required
@bakim_required
def bakim_dashboard():
    from datetime import timedelta
    from sqlalchemy import func
    bugun = date.today()
    yedi_gun = bugun + timedelta(days=7)
    otuz_gun_once = bugun - timedelta(days=30)
    otuz_gun_sonra = bugun + timedelta(days=30)

    # Bu ay / geçen ay sınırları
    bu_ay_bas = bugun.replace(day=1)
    gec_ay_son = bu_ay_bas - timedelta(days=1)
    gec_ay_bas = gec_ay_son.replace(day=1)

    bugunki_kayitlar = BakimKaydi.query.filter_by(tarih=bugun).order_by(BakimKaydi.id.desc()).all()
    yaklasan = BakimPlani.query.filter(
        BakimPlani.sonraki_bakim_tarihi <= yedi_gun,
        BakimPlani.sonraki_bakim_tarihi >= bugun,
        BakimPlani.is_active == True
    ).order_by(BakimPlani.sonraki_bakim_tarihi).all()

    # --- Arıza istatistikleri ---
    bu_ay_ariza = BakimKaydi.query.filter(
        BakimKaydi.bakim_turu == 'ariza',
        BakimKaydi.tarih >= bu_ay_bas
    ).count()
    gec_ay_ariza = BakimKaydi.query.filter(
        BakimKaydi.bakim_turu == 'ariza',
        BakimKaydi.tarih >= gec_ay_bas,
        BakimKaydi.tarih <= gec_ay_son
    ).count()

    # --- Bu ay toplam bakım süresi (dakika → saat) ---
    bu_ay_sure_dk = db.session.query(func.sum(BakimKaydi.sure_dakika)).filter(
        BakimKaydi.tarih >= bu_ay_bas
    ).scalar() or 0
    bu_ay_sure_saat = round(bu_ay_sure_dk / 60, 1)

    # --- Arıza oranı (bu ay) ---
    bu_ay_toplam = BakimKaydi.query.filter(BakimKaydi.tarih >= bu_ay_bas).count()
    ariza_oran = round(bu_ay_ariza / bu_ay_toplam * 100) if bu_ay_toplam else 0

    # --- En sık arıza yapan makineler (son 30 gün, top 5) ---
    en_cok_ariza = db.session.query(
        Makine.makine_adi,
        func.count(BakimKaydi.id).label('adet')
    ).join(BakimKaydi, BakimKaydi.makine_id == Makine.id).filter(
        BakimKaydi.bakim_turu == 'ariza',
        BakimKaydi.tarih >= otuz_gun_once
    ).group_by(Makine.id, Makine.makine_adi).order_by(
        func.count(BakimKaydi.id).desc()
    ).limit(5).all()

    # --- Periyodik kontrol uyarıları (gecikmiş + 30 gün içinde) ---
    periyodik_uyari = PeriyodikKontrol.query.filter(
        PeriyodikKontrol.durum == 'aktif',
        PeriyodikKontrol.sonraki_kontrol_tarihi != None,
        PeriyodikKontrol.sonraki_kontrol_tarihi <= otuz_gun_sonra
    ).order_by(PeriyodikKontrol.sonraki_kontrol_tarihi).limit(10).all()

    try:
        acik_dof_sayisi = DOF.query.filter(
            DOF.hedef_departman == 'Bakım',
            DOF.durum.in_(('acik', 'isleniyor', 'gecikti'))
        ).count()
        geciken_aksiyon_sayisi = db.session.query(DOFAksiyon).join(DOF).filter(
            DOF.hedef_departman == 'Bakım',
            DOFAksiyon.durum != 'tamamlandi',
            DOFAksiyon.planlanan_tarih < bugun
        ).count()
    except Exception:
        acik_dof_sayisi = 0
        geciken_aksiyon_sayisi = 0

    return render_template('bakim/dashboard.html',
        bugunki_kayitlar=bugunki_kayitlar, yaklasan=yaklasan, bugun=bugun,
        acik_dof_sayisi=acik_dof_sayisi, geciken_aksiyon_sayisi=geciken_aksiyon_sayisi,
        bu_ay_ariza=bu_ay_ariza, gec_ay_ariza=gec_ay_ariza,
        bu_ay_sure_saat=bu_ay_sure_saat, ariza_oran=ariza_oran,
        en_cok_ariza=en_cok_ariza, periyodik_uyari=periyodik_uyari,
        otuz_gun_sonra=otuz_gun_sonra)

@bakim.route('/kayit', methods=['GET', 'POST'])
@login_required
@bakim_required
def bakim_kayit():
    if request.method == 'POST':
        from datetime import timedelta
        makine_id = request.form.get('makine_id')
        plan_id = request.form.get('bakim_plani_id') or None
        tarih_str = request.form.get('tarih', str(date.today()))
        k = BakimKaydi(
            makine_id=makine_id,
            bakim_plani_id=plan_id,
            bakim_turu=request.form.get('bakim_turu', 'gunluk'),
            tarih=date.fromisoformat(tarih_str),
            yapilan_isler=request.form.get('yapilan_isler', ''),
            sure_dakika=request.form.get('sure_dakika') or None,
            giren_personel_id=current_user.id,
        )
        db.session.add(k)
        if plan_id:
            plan = BakimPlani.query.get(int(plan_id))
            if plan:
                plan.son_bakim_tarihi = date.fromisoformat(tarih_str)
                plan.sonraki_bakim_tarihi = date.fromisoformat(tarih_str) + timedelta(days=plan.periyot_gun)
        db.session.commit()
        flash('Bakım kaydı eklendi.', 'success')
        return redirect(url_for('bakim.bakim_dashboard'))
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    return render_template('bakim/kayit.html', makineler=makineler, bugun=date.today())

@bakim.route('/makine/<int:makine_id>/planlar')
@login_required
@bakim_required
def makine_planlari(makine_id):
    return jsonify([{
        'id': p.id, 'bakim_adi': p.bakim_adi,
        'periyot_gun': p.periyot_gun,
        'sonraki': str(p.sonraki_bakim_tarihi) if p.sonraki_bakim_tarihi else ''
    } for p in BakimPlani.query.filter_by(makine_id=makine_id, is_active=True).all()])

@bakim.route('/makineler', methods=['GET', 'POST'])
@login_required
@bakim_required
def makine_listesi():
    if request.method == 'POST':
        eylem = request.form.get('eylem')
        if eylem == 'ekle':
            from app.utils import generate_makine_kodu
            m = Makine(
                makine_kodu=generate_makine_kodu(),
                makine_adi=request.form.get('makine_adi', '').strip(),
                marka=request.form.get('marka', '').strip(),
                model=request.form.get('model', '').strip(),
                seri_no=request.form.get('seri_no', '').strip(),
            )
            db.session.add(m); db.session.commit()
            flash('Makine eklendi.', 'success')
        elif eylem == 'sil':
            m = db.get_or_404(Makine, request.form.get('makine_id'))
            m.is_active = False; db.session.commit()
            flash('Makine pasif yapıldı.', 'warning')
        return redirect(url_for('bakim.makine_listesi'))
    makineler = Makine.query.order_by(Makine.makine_kodu).all()
    return render_template('bakim/makineler.html', makineler=makineler)

@bakim.route('/plan', methods=['GET', 'POST'])
@login_required
@bakim_required
def bakim_plan():
    if request.method == 'POST':
        from datetime import timedelta
        eylem = request.form.get('eylem')
        if eylem == 'ekle':
            son_str = request.form.get('son_bakim_tarihi')
            periyot = int(request.form.get('periyot_gun', 30))
            son_dt = date.fromisoformat(son_str) if son_str else None
            sonraki = (son_dt + timedelta(days=periyot)) if son_dt else None
            p = BakimPlani(
                makine_id=request.form.get('makine_id'),
                bakim_adi=request.form.get('bakim_adi', '').strip(),
                periyot_gun=periyot,
                son_bakim_tarihi=son_dt,
                sonraki_bakim_tarihi=sonraki,
                aciklama=request.form.get('aciklama', '').strip(),
            )
            db.session.add(p); db.session.commit()
            flash('Bakım planı eklendi.', 'success')
        elif eylem == 'sil':
            p = db.get_or_404(BakimPlani, request.form.get('plan_id'))
            p.is_active = False; db.session.commit()
        return redirect(url_for('bakim.bakim_plan'))
    planlar = BakimPlani.query.filter_by(is_active=True).order_by(BakimPlani.sonraki_bakim_tarihi).all()
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    return render_template('bakim/plan.html', planlar=planlar, makineler=makineler)

@bakim.route('/takvim')
@login_required
@bakim_required
def bakim_takvim():
    import calendar
    yil = int(request.args.get('yil', date.today().year))
    ay = int(request.args.get('ay', date.today().month))
    _, gun_sayisi = calendar.monthrange(yil, ay)
    gunler = list(range(1, gun_sayisi + 1))
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    ay_bas = date(yil, ay, 1)
    ay_son = date(yil, ay, gun_sayisi)
    kayitlar = BakimKaydi.query.filter(
        BakimKaydi.tarih >= ay_bas,
        BakimKaydi.tarih <= ay_son
    ).all()
    kayit_map = {}
    for k in kayitlar:
        kayit_map.setdefault(k.makine_id, set()).add(k.tarih.day)
    return render_template('bakim/takvim.html',
        makineler=makineler, gunler=gunler,
        kayit_map=kayit_map, yil=yil, ay=ay, gun_sayisi=gun_sayisi)

@bakim.route('/raporlar')
@login_required
def bakim_raporlar():
    if current_user.role not in ('bakim', 'departman_yoneticisi', 'gm', 'admin') and not has_permission(current_user, 'maintenance.calendar_report'):
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    baslangic = request.args.get('baslangic', str(date.today() - timedelta(days=29)))
    bitis = request.args.get('bitis', str(date.today()))
    makine_id = request.args.get('makine_id', '')
    try:
        bas_dt = date.fromisoformat(baslangic)
        bit_dt = date.fromisoformat(bitis)
    except ValueError:
        bas_dt = date.today() - timedelta(days=29)
        bit_dt = date.today()
    q = BakimKaydi.query.filter(BakimKaydi.tarih >= bas_dt, BakimKaydi.tarih <= bit_dt)
    if makine_id:
        q = q.filter(BakimKaydi.makine_id == makine_id)
    kayitlar = q.order_by(BakimKaydi.tarih.desc()).all()
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    return render_template('bakim/raporlar.html',
        kayitlar=kayitlar, makineler=makineler,
        makine_id=makine_id, baslangic=baslangic, bitis=bitis)


# ── PLANLİ BAKIM ────────────────────────────────────────────────────────────

@bakim.route('/program', methods=['GET', 'POST'])
@login_required
@bakim_required
def bakim_program():
    from app.models import PlanliBakim
    if request.method == 'POST':
        eylem = request.form.get('eylem', 'ekle')
        if eylem == 'ekle':
            bas_str = request.form.get('baslangic_tarihi')
            bit_str = request.form.get('bitis_tarihi') or bas_str
            bas_saat = request.form.get('baslangic_saati') or None
            bit_saat = request.form.get('bitis_saati') or None
            sure = request.form.get('sure_saat') or None
            from datetime import time as dtime
            def parse_time(s):
                if not s:
                    return None
                try:
                    h, m = s.split(':')
                    return dtime(int(h), int(m))
                except Exception:
                    return None
            p = PlanliBakim(
                makine_id=request.form.get('makine_id'),
                bakim_adi=request.form.get('bakim_adi', '').strip(),
                baslangic_tarihi=date.fromisoformat(bas_str),
                bitis_tarihi=date.fromisoformat(bit_str),
                baslangic_saati=parse_time(bas_saat),
                bitis_saati=parse_time(bit_saat),
                sure_saat=float(sure) if sure else None,
                notlar=request.form.get('notlar', '').strip() or None,
                planlayan_id=current_user.id,
            )
            db.session.add(p)
            db.session.commit()
            flash(f'Planlı bakım eklendi: {p.makine.makine_adi} — {p.baslangic_tarihi.strftime("%d.%m.%Y")}', 'success')
        elif eylem == 'durum':
            p = db.get_or_404(PlanliBakim, request.form.get('program_id'))
            p.durum = request.form.get('durum', p.durum)
            db.session.commit()
            flash('Durum güncellendi.', 'success')
        return redirect(url_for('bakim.bakim_program'))

    durum_filtre = request.args.get('durum', '')
    q = PlanliBakim.query
    if durum_filtre:
        q = q.filter_by(durum=durum_filtre)
    else:
        q = q.filter(PlanliBakim.durum.in_(['planli', 'devam_ediyor']))
    programlar = q.order_by(PlanliBakim.baslangic_tarihi).all()
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    return render_template('bakim/program.html',
        programlar=programlar, makineler=makineler,
        durum_filtre=durum_filtre, bugun=date.today())


@bakim.route('/program/<int:id>/durum', methods=['POST'])
@login_required
@bakim_required
def program_durum(id):
    from app.models import PlanliBakim
    p = db.get_or_404(PlanliBakim, id)
    p.durum = request.form.get('durum', p.durum)
    db.session.commit()
    flash('Durum güncellendi.', 'success')
    return redirect(url_for('bakim.bakim_program'))


# ── PERİYODİK KONTROL ───────────────────────────────────────────────────────

@bakim.route('/periyodik', methods=['GET', 'POST'])
@login_required
def periyodik_listesi():
    if not has_permission(current_user, 'maintenance.periodic'):
        flash('Erişim yetkiniz yok.', 'danger')
        return redirect(url_for('auth.portal'))
    from app.models import PeriyodikKontrol
    if request.method == 'POST':
        eylem = request.form.get('eylem', 'ekle')
        if eylem == 'ekle':
            son_str = request.form.get('son_kontrol_tarihi') or None
            periyot = int(request.form.get('periyot_ay', 12))
            son_dt = date.fromisoformat(son_str) if son_str else None
            from dateutil.relativedelta import relativedelta
            sonraki = (son_dt + relativedelta(months=periyot)) if son_dt else None
            pk = PeriyodikKontrol(
                makine_id=request.form.get('makine_id') or None,
                kontrol_adi=request.form.get('kontrol_adi', '').strip(),
                kontrol_turu=request.form.get('kontrol_turu', 'diger'),
                periyot_ay=periyot,
                son_kontrol_tarihi=son_dt,
                sonraki_kontrol_tarihi=sonraki,
                kontrol_eden_firma=request.form.get('kontrol_eden_firma', '').strip() or None,
                son_kontrol_notu=request.form.get('son_kontrol_notu', '').strip() or None,
                olusturan_id=current_user.id,
            )
            db.session.add(pk)
            db.session.commit()
            flash('Periyodik kontrol eklendi.', 'success')
        elif eylem == 'guncelle':
            pk = db.get_or_404(PeriyodikKontrol, request.form.get('kontrol_id'))
            son_str = request.form.get('son_kontrol_tarihi')
            if son_str:
                from dateutil.relativedelta import relativedelta
                son_dt = date.fromisoformat(son_str)
                pk.son_kontrol_tarihi = son_dt
                pk.sonraki_kontrol_tarihi = son_dt + relativedelta(months=pk.periyot_ay)
            pk.kontrol_eden_firma = request.form.get('kontrol_eden_firma', '').strip() or pk.kontrol_eden_firma
            pk.son_kontrol_notu = request.form.get('son_kontrol_notu', '').strip() or None
            db.session.commit()
            flash('Kontrol güncellendi.', 'success')
        elif eylem == 'pasif':
            pk = db.get_or_404(PeriyodikKontrol, request.form.get('kontrol_id'))
            pk.durum = 'pasif'
            db.session.commit()
            flash('Kontrol pasife alındı.', 'warning')
        return redirect(url_for('bakim.periyodik_listesi'))

    bugun = date.today()
    bir_ay_sonra = bugun + timedelta(days=30)
    kontroller = PeriyodikKontrol.query.filter_by(durum='aktif').order_by(PeriyodikKontrol.sonraki_kontrol_tarihi).all()
    yaklaşan = [k for k in kontroller if k.sonraki_kontrol_tarihi and k.sonraki_kontrol_tarihi <= bir_ay_sonra]
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    return render_template('bakim/periyodik.html',
        kontroller=kontroller, yaklaşan=yaklaşan,
        makineler=makineler, bugun=bugun, bir_ay_sonra=bir_ay_sonra)


# ── ARIZ TAMIR KAYIT (bakım personeli giriyor) ──────────────────────────────

@bakim.route('/ariza-tamir', methods=['GET', 'POST'])
@login_required
@bakim_required
def ariza_tamir_listesi():
    from datetime import time as dtime
    if request.method == 'POST':
        bas_saat_str = request.form.get('baslangic_saati') or None
        bit_saat_str = request.form.get('bitis_saati') or None
        def parse_time(s):
            if not s:
                return None
            try:
                h, m = s.split(':')
                return dtime(int(h), int(m))
            except Exception:
                return None
        sure = None
        bas_t = parse_time(bas_saat_str)
        bit_t = parse_time(bit_saat_str)
        if bas_t and bit_t:
            bas_min = bas_t.hour * 60 + bas_t.minute
            bit_min = bit_t.hour * 60 + bit_t.minute
            if bit_min > bas_min:
                sure = bit_min - bas_min
        ariza_devam = bool(request.form.get('ariza_devam_ediyor'))
        k = BakimKaydi(
            makine_id=request.form.get('makine_id'),
            bakim_turu='ariza',
            tarih=date.fromisoformat(request.form.get('tarih', str(date.today()))),
            yapilan_isler=request.form.get('yapilan_isler', '').strip(),
            sure_dakika=None if ariza_devam else (sure or (int(request.form.get('sure_dakika')) if request.form.get('sure_dakika') else None)),
            baslangic_saati=bas_t,
            bitis_saati=None if ariza_devam else bit_t,
            parca_kullanildi=bool(request.form.get('parca_kullanildi')),
            kullanilan_parcalar=request.form.get('kullanilan_parcalar', '').strip() or None,
            ariza_devam_ediyor=ariza_devam,
            giren_personel_id=current_user.id,
        )
        db.session.add(k)
        db.session.commit()
        flash('Arıza tamir kaydı eklendi.', 'success')
        return redirect(url_for('bakim.ariza_tamir_listesi'))

    bas = request.args.get('baslangic', str(date.today() - timedelta(days=29)))
    bit = request.args.get('bitis', str(date.today()))
    makine_id = request.args.get('makine_id', '')
    try:
        bas_dt = date.fromisoformat(bas)
        bit_dt = date.fromisoformat(bit)
    except ValueError:
        bas_dt = date.today() - timedelta(days=29)
        bit_dt = date.today()
    q = BakimKaydi.query.filter_by(bakim_turu='ariza').filter(
        BakimKaydi.tarih >= bas_dt, BakimKaydi.tarih <= bit_dt)
    if makine_id:
        q = q.filter_by(makine_id=makine_id)
    kayitlar = q.order_by(BakimKaydi.tarih.desc(), BakimKaydi.id.desc()).all()
    makineler = Makine.query.filter_by(is_active=True).order_by(Makine.makine_adi).all()
    return render_template('bakim/ariza_tamir.html',
        kayitlar=kayitlar, makineler=makineler,
        makine_id=makine_id, baslangic=bas, bitis=bit, bugun=date.today())


# ── API: Üretim planı için bakım uyarısı ───────────────────────────────────

@api.route('/bakim-kontrol')
@login_required
def api_bakim_kontrol():
    """Verilen istasyon ve tarihe bakım var mı? Üretim planı uyarısı için."""
    from app.models import PlanliBakim
    istasyon_id = request.args.get('istasyon_id', type=int)
    tarih_str = request.args.get('tarih', '')
    if not istasyon_id or not tarih_str:
        return jsonify({'uyari': False})
    try:
        tarih = date.fromisoformat(tarih_str)
    except ValueError:
        return jsonify({'uyari': False})
    planli = (PlanliBakim.query
        .join(Makine, Makine.id == PlanliBakim.makine_id)
        .filter(
            Makine.istasyon_id == istasyon_id,
            PlanliBakim.baslangic_tarihi <= tarih,
            PlanliBakim.bitis_tarihi >= tarih,
            PlanliBakim.durum.in_(['planli', 'devam_ediyor'])
        ).first())
    if planli:
        return jsonify({
            'uyari': True,
            'mesaj': f'{planli.makine.makine_adi} makinesi bu gün planlı bakımda ({planli.bakim_adi})'
        })
    return jsonify({'uyari': False})


# ---------------------------------------------------------------------------
# TEKNİK RESİM BLUEPRINT
# ---------------------------------------------------------------------------

@teknik_resim_bp.route('/')
@login_required
def teknik_resim_listesi():
    from sqlalchemy import func
    q = request.args.get('q', '').strip()
    edit_yetkili = has_permission(current_user, 'technical.manage')

    if q:
        # Arama: eşleşen dosyaları getir, klasöre göre grupla
        resimler = TeknikResim.query.filter(
            db.or_(
                TeknikResim.dosya_adi_gosterim.ilike(f'%{q}%'),
                TeknikResim.klasor.ilike(f'%{q}%'),
                TeknikResim.aciklama.ilike(f'%{q}%')
            )
        ).order_by(TeknikResim.klasor, TeknikResim.dosya_adi_gosterim).all()
        from collections import defaultdict
        gruplar = defaultdict(list)
        for r in resimler:
            gruplar[r.klasor or ''].append(r)
        return render_template('teknik_resimler.html',
                               mod='arama', gruplar=dict(sorted(gruplar.items())),
                               toplam=len(resimler), q=q, edit_yetkili=edit_yetkili)
    else:
        # Varsayılan: sadece klasör adları + dosya sayısı
        klasor_stats = db.session.query(
            TeknikResim.klasor,
            func.count(TeknikResim.id).label('sayi')
        ).group_by(TeknikResim.klasor).order_by(TeknikResim.klasor).all()
        toplam = db.session.query(func.count(TeknikResim.id)).scalar()
        return render_template('teknik_resimler.html',
                               mod='klasor', klasor_stats=klasor_stats,
                               toplam=toplam, q=q, edit_yetkili=edit_yetkili)


@teknik_resim_bp.route('/klasor-icerik')
@login_required
def klasor_icerik():
    klasor = request.args.get('k', '')
    if klasor == '__bos__':
        klasor = None
    if klasor is None:
        resimler = TeknikResim.query.filter(TeknikResim.klasor.is_(None))\
                    .order_by(TeknikResim.dosya_adi_gosterim).all()
    else:
        resimler = TeknikResim.query.filter_by(klasor=klasor)\
                    .order_by(TeknikResim.dosya_adi_gosterim).all()
    edit_yetkili = has_permission(current_user, 'technical.manage')
    return render_template('teknik_resim_satirlar.html', resimler=resimler, edit_yetkili=edit_yetkili)


@teknik_resim_bp.route('/yukle', methods=['POST'])
@login_required
@teknik_yetki_required
def teknik_resim_yukle():
    import os
    from werkzeug.utils import secure_filename

    dosya = request.files.get('pdf_dosya')
    if not dosya or not dosya.filename.lower().endswith('.pdf'):
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'hata': 'PDF değil'}), 400
        flash('Geçerli bir PDF dosyası seçin.', 'danger')
        return redirect(url_for('teknik_resim.teknik_resim_listesi'))

    klasor = request.form.get('klasor', '').strip()
    dosya_adi_gosterim = request.form.get('dosya_adi_gosterim', '').strip()
    aciklama = request.form.get('aciklama', '').strip()

    if not dosya_adi_gosterim:
        dosya_adi_gosterim = os.path.splitext(dosya.filename)[0]

    disk_adi = secure_filename(dosya.filename)
    kayit_dizini = os.path.join(os.path.dirname(__file__), 'static', 'teknik_resimler')
    os.makedirs(kayit_dizini, exist_ok=True)
    tam_yol = os.path.join(kayit_dizini, disk_adi)
    if os.path.exists(tam_yol):
        base, ext = os.path.splitext(disk_adi)
        disk_adi = f"{base}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{ext}"
        tam_yol = os.path.join(kayit_dizini, disk_adi)
    dosya.save(tam_yol)

    resim = TeknikResim(
        klasor=klasor or None,
        dosya_adi_gosterim=dosya_adi_gosterim,
        aciklama=aciklama or None,
        dosya_adi=disk_adi,
        yukleyen_id=current_user.id
    )
    db.session.add(resim)
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    flash('Teknik resim başarıyla yüklendi.', 'success')
    return redirect(url_for('teknik_resim.teknik_resim_listesi'))


@teknik_resim_bp.route('/goruntule/<int:resim_id>')
@login_required
def teknik_resim_goruntule(resim_id):
    import os
    from flask import send_from_directory
    resim = db.get_or_404(TeknikResim, resim_id)
    dizin = os.path.join(os.path.dirname(__file__), 'static', 'teknik_resimler')
    return send_from_directory(dizin, resim.dosya_adi, as_attachment=False)


@teknik_resim_bp.route('/indir/<int:resim_id>')
@login_required
def teknik_resim_indir(resim_id):
    import os
    from flask import send_from_directory
    resim = db.get_or_404(TeknikResim, resim_id)
    dizin = os.path.join(os.path.dirname(__file__), 'static', 'teknik_resimler')
    return send_from_directory(dizin, resim.dosya_adi, as_attachment=True, download_name=resim.dosya_adi)


@teknik_resim_bp.route('/sil/<int:resim_id>', methods=['POST'])
@login_required
@teknik_yetki_required
def teknik_resim_sil(resim_id):
    import os
    resim = db.get_or_404(TeknikResim, resim_id)
    dosya_yolu = os.path.join(os.path.dirname(__file__), 'static', 'teknik_resimler', resim.dosya_adi)
    if os.path.exists(dosya_yolu):
        os.remove(dosya_yolu)
    db.session.delete(resim)
    db.session.commit()
    flash('Teknik resim silindi.', 'success')
    return redirect(url_for('teknik_resim.teknik_resim_listesi'))
